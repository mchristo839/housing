"""Re-add verified contact details for high-value providers we lost in the
wrong-domain strip. Each contact is verified via WebSearch and added to
Companies sheet so the build pipeline restores them.
"""
import openpyxl, re

PATH = 'data/manual_contracts.xlsx'
wb = openpyxl.load_workbook(PATH)
companies = wb['Companies']

# Verified contacts from web searches just done
RESTORE = [
    # (Canonical name to look for in existing Companies sheet, website, phone, email, address)
    ('Calico',
     'https://calico.org.uk/',
     '01282 686300',
     'info@calicohousing.co.uk',
     'Centenary Court, Croft Street, Burnley, BB11 2ED'),
    ('Heathcotes Group',
     'https://www.heathcotes.net/',
     '01246 252 575',
     'admin@heathcotes.net',
     '37 Station Road, Chesterfield, Derbyshire, S41 7BF'),
    ('Heantun Housing Association Limited',
     'https://www.heantun.co.uk/',
     '01902 571100',
     'janet.humphries@heantun.co.uk',
     '3 Wellington Road, Bilston, West Midlands, WV14 6AA'),
    ('Priory Group',
     'https://www.priorygroup.com/',
     '0330 056 6020',
     'info@priorygroup.com',
     'Priory Group, 80 Hammersmith Road, London, W14 8UD'),
]

def find_row(name):
    target = name.lower().strip()
    for r in range(2, companies.max_row + 1):
        v = (companies.cell(row=r, column=1).value or '').strip().lower()
        if v == target: return r
    return None

n_updated = 0
n_added = 0
for (name, website, phone, email, address) in RESTORE:
    r = find_row(name)
    if r is None:
        # New row
        r = companies.max_row + 1
        companies.cell(row=r, column=1).value = name
        companies.cell(row=r, column=2).value = 0
        companies.cell(row=r, column=5).value = 1   # Housing Contracts (placeholder >0)
        companies.cell(row=r, column=7).value = 'No'
        n_added += 1
        print(f"  + ADD: {name}")
    else:
        n_updated += 1
        print(f"  * UPDATE: {name}")
    companies.cell(row=r, column=12).value = address
    companies.cell(row=r, column=13).value = website
    companies.cell(row=r, column=14).value = phone
    companies.cell(row=r, column=15).value = email

wb.save(PATH)
print(f"\nTotal: {n_updated} updated, {n_added} new")
