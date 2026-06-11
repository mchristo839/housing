"""
Import a marks CSV exported from REVIEW_PACK.html back into REVIEW_PACK.xlsx.

The HTML page's "Download my marks (CSV)" button produces a CSV with columns:
  row_index, supplier, council_raw, source_id, mark, notes

Usage:
  python scripts/import_marks_csv.py review_marks.csv
       Stamps Approve? + Notes columns in REVIEW_PACK.xlsx using row_index
       to align (1:1 with the order the HTML was built from).
       Falls back to (supplier, council_raw, source_id) triplet if the
       row_index is out of range (in case the pack was regenerated).
"""
import openpyxl, csv, sys, io, argparse, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

REVIEW = 'data/scraped/REVIEW_PACK.xlsx'
COL_APPROVE = 1
COL_NOTES   = 3
COL_SUPPLIER = 5
COL_SID      = 13
COL_COUNCIL  = 14


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('csv_path')
    a = ap.parse_args()

    if not os.path.exists(a.csv_path):
        print(f"ERR: {a.csv_path} not found")
        sys.exit(1)

    # Load marks
    marks = []   # list of dict
    with open(a.csv_path, encoding='utf-8-sig') as f:
        r = csv.DictReader(f)
        for row in r:
            marks.append({
                'idx': int(row.get('row_index', -1) or -1),
                'supplier': (row.get('supplier','') or '').strip(),
                'council':  (row.get('council_raw','') or '').strip(),
                'sid':      (row.get('source_id','') or '').strip(),
                'mark':     (row.get('mark','') or '').strip().upper(),
                'notes':    (row.get('notes','') or '').strip(),
            })
    print(f"Loaded {len(marks)} marks from {a.csv_path}")

    wb = openpyxl.load_workbook(REVIEW)
    ws = wb['Review']
    rows = list(ws.iter_rows(min_row=2))    # 0-based index aligned with HTML

    # build triplet lookup for fallback
    triplet_to_row = {}
    for i, row in enumerate(rows):
        sup = str(row[COL_SUPPLIER-1].value or '').strip()
        cnc = str(row[COL_COUNCIL-1].value or '').strip()
        sid = str(row[COL_SID-1].value or '').strip()
        triplet_to_row[(sup, cnc, sid)] = i

    applied = misses = 0
    for m in marks:
        target = None
        if 0 <= m['idx'] < len(rows):
            target = rows[m['idx']]
        if target is None:
            i = triplet_to_row.get((m['supplier'], m['council'], m['sid']))
            if i is not None:
                target = rows[i]
        if target is None:
            misses += 1
            continue
        if m['mark']:
            target[COL_APPROVE-1].value = m['mark']
        if m['notes']:
            target[COL_NOTES-1].value = m['notes']
        applied += 1

    wb.save(REVIEW)

    print(f"\nApplied : {applied}")
    print(f"Misses  : {misses}")

    # Status snapshot
    from collections import Counter
    c = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = str(row[0] or '').strip().upper()
        if v in ('Y','YES'):     k='Y'
        elif v in ('N','NO'):    k='N'
        elif v in ('?','MAYBE'): k='?'
        else:                    k='_'
        c[k] += 1
    print(f"\nREVIEW PACK NOW: Y:{c['Y']} · N:{c['N']} · ?:{c['?']} · unmarked:{c['_']} (total {sum(c.values())})")
    print(f"\nNext: tell Claude 'promote approved rows'")


if __name__ == '__main__':
    main()
