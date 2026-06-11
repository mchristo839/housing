"""Apply Firecrawl Mode B verdicts to VERIFIED.json + manual_contracts.xlsx.

KEEP verdicts get added to VERIFIED.json with the LLM-extracted services list.
Contact details (email/phone/address/CH number) get written to manual_contracts
so they survive monthly source refreshes.
DROP verdicts go into MANUAL_DROP_LIST.
"""
import json, time
from pathlib import Path
import openpyxl

FC_FILE   = Path('data/verification/firecrawl_analysed.json')
VERIFIED  = Path('data/verification/VERIFIED.json')
DROPS     = Path('data/MANUAL_DROP_LIST.json')
MANUAL_XL = Path('data/manual_contracts.xlsx')

fc = json.load(open(FC_FILE, encoding='utf-8'))
v  = json.load(open(VERIFIED, encoding='utf-8'))
d  = json.load(open(DROPS, encoding='utf-8'))

# Apply VERIFIED + DROPS to JSON files
new_v = 0; new_d = 0
for name, r in fc.items():
    if r.get('verdict') == 'KEEP':
        if name not in v:
            v[name] = {
                'verdict':     'KEEP',
                'step1':       'Y',
                'services':    ', '.join(r.get('housing_terms') or []),
                'verified_at': r.get('checked_at', time.strftime('%Y-%m-%d')),
                'notes':       f"Firecrawl LLM extraction; CH on site: {r.get('ch_number','')}",
            }
            new_v += 1
    elif r.get('verdict') == 'DROP':
        if name not in d:
            d.append(name); new_d += 1

json.dump(v, open(VERIFIED, 'w', encoding='utf-8'), indent=2)
json.dump(d, open(DROPS, 'w', encoding='utf-8'), indent=2)
print(f"VERIFIED:  +{new_v}  (total {len(v)})")
print(f"DROPS:     +{new_d}  (total {len(d)})")

# Apply contacts to manual_contracts.xlsx Companies sheet
wb = openpyxl.load_workbook(MANUAL_XL)
comp = wb['Companies']
idx = {}
for r in range(2, comp.max_row+1):
    nm = (comp.cell(row=r, column=1).value or '').strip().lower()
    if nm: idx[nm] = r

upd = add = 0
for name, r in fc.items():
    if r.get('verdict') != 'KEEP': continue
    nl = name.lower().strip()
    row_idx = idx.get(nl)
    if row_idx is None:
        row_idx = comp.max_row + 1
        comp.cell(row=row_idx, column=1).value = name
        comp.cell(row=row_idx, column=2).value = 0
        comp.cell(row=row_idx, column=5).value = 1
        comp.cell(row=row_idx, column=7).value = 'No'
        idx[nl] = row_idx
        add += 1
    else:
        upd += 1
    # Columns: 12=address 13=website 14=phone 15=email 16=contact page
    if r.get('address'):
        comp.cell(row=row_idx, column=12).value = r['address']
    if r.get('url'):
        comp.cell(row=row_idx, column=13).value = r['url']
    if r.get('phones'):
        comp.cell(row=row_idx, column=14).value = r['phones'][0]
    if r.get('emails'):
        comp.cell(row=row_idx, column=15).value = r['emails'][0]
    # Capture the Contact Us page URL as a fallback when no email is available.
    # The UI renders a "✎ contact form" icon when email is empty + contact_page exists.
    if r.get('contact_page') and not r.get('emails'):
        comp.cell(row=row_idx, column=16).value = r['contact_page']

wb.save(MANUAL_XL)
print(f"Companies sheet: {upd} updated, {add} added")
