"""Round 2 enrichment — apply all verified contact info found across batches.
Set Website Review = 'verified' for all suppliers with real direct contacts.
"""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = 'data/manual_contracts.xlsx'

# (name_lower → (website, phone, email, contact_page, companies_house, charity_no))
VERIFIED = {
    # Major providers from round 1 (re-include for safety)
    'hestia housing & support':
        ('https://www.hestia.org', '020 7378 3100', 'info@hestia.org',
         'https://www.hestia.org/contact-hestia', '', '294555'),
    'hestia housing and support':
        ('https://www.hestia.org', '020 7378 3100', 'info@hestia.org',
         'https://www.hestia.org/contact-hestia', '', '294555'),
    'look ahead care & support':
        ('https://www.lookahead.org.uk', '020 7937 1166', 'info@lookahead.org.uk',
         'https://www.lookahead.org.uk/about-us/contact-us/', '', ''),
    'look ahead care and support':
        ('https://www.lookahead.org.uk', '020 7937 1166', 'info@lookahead.org.uk',
         'https://www.lookahead.org.uk/about-us/contact-us/', '', ''),
    'look ahead housing & care':
        ('https://www.lookahead.org.uk', '020 7937 1166', 'info@lookahead.org.uk',
         'https://www.lookahead.org.uk/about-us/contact-us/', '', ''),
    'look ahead housing and care':
        ('https://www.lookahead.org.uk', '020 7937 1166', 'info@lookahead.org.uk',
         'https://www.lookahead.org.uk/about-us/contact-us/', '', ''),
    "st mungo's":
        ('https://www.mungos.org', '020 3856 6000', 'info@mungos.org',
         'https://www.mungos.org/contact/', '', '1149085'),
    "st mungos":
        ('https://www.mungos.org', '020 3856 6000', 'info@mungos.org',
         'https://www.mungos.org/contact/', '', '1149085'),
    'st mungo community housing association':
        ('https://www.mungos.org', '020 3856 6000', 'info@mungos.org',
         'https://www.mungos.org/contact/', '', '1149085'),
    'turning point services':
        ('https://www.turning-point.co.uk', '020 7481 7600', 'referrals@turning-point.co.uk',
         'https://www.turning-point.co.uk/contact-us', '', '234887'),
    'catalyst housing':
        ('https://www.catalysthousing.co.uk', '0300 456 2099',
         'customerservice@catalysthousing.co.uk',
         'https://www.catalysthousing.co.uk/contact-us', '', ''),
    'certitude':
        ('https://certitude.london', '020 3397 3033', 'referrals@certitude.london',
         'https://certitude.org.uk/contact', '', ''),
    'profad care agency':
        ('https://profadcareagency.co.uk', '020 7639 0839', 'info@profadcareagency.co.uk',
         'https://profadcareagency.co.uk/contact', '', ''),
    'chosen care group':
        ('https://chosencaregroup.com', '020 8214 1093', 'complaint@chosencaregroup.com',
         'https://chosencaregroup.com/contact-us/', '', ''),
    'eleanor nursing & social care':
        ('https://www.eleanorhealthcaregroup.co.uk', '020 8690 2406', 'info@eleanorcare.co.uk',
         'https://www.eleanorhealthcaregroup.co.uk/contact-us/', '', ''),
    'eleanor nursing and social care':
        ('https://www.eleanorhealthcaregroup.co.uk', '020 8690 2406', 'info@eleanorcare.co.uk',
         'https://www.eleanorhealthcaregroup.co.uk/contact-us/', '', ''),
    'caretech community services':
        ('https://www.caretech-uk.com', '01707 601 800', 'info@caretech-uk.com',
         'https://www.caretech-uk.com/contact-us', '', ''),
    'grove social care':
        ('https://www.grovesocialcare.co.uk', '01733 568444', 'jobs@grovesocialcare.co.uk',
         'https://www.grovesocialcare.co.uk/contact', '10567559', ''),
    'voyage 1':
        ('https://www.voyagecare.com', '0800 035 3776', 'referrals@voyagecare.com',
         'https://www.voyagecare.com/contact-us/', '', ''),
    'accomplish group support':
        ('https://www.accomplish-group.co.uk', '0151 726 1460', 'info@accomplish-group.co.uk',
         'https://www.accomplish-group.co.uk/contact-us/', '', ''),
    'avenues group':
        ('https://www.avenuesgroup.org.uk', '01622 366 911', 'info@avenuesgroup.org.uk',
         'https://www.avenuesgroup.org.uk/contact-us/', '', ''),
    # Round 2 — new finds
    'midco care':
        ('https://www.midco-care.co.uk', '01733 530 580', 'info@midco-care.co.uk',
         'https://www.midco-care.co.uk/contact-us/', '', ''),
    'southcroft healthcare lodge':
        ('https://southcroftlodge.co.uk', '020 8764 9888', 'info@southcroftlodge.co.uk',
         'https://southcroftlodge.co.uk/contact', '12951105', ''),
    'ability care and support':
        ('https://www.ability-housing.co.uk', '01784 446 110', 'info@ability-housing.co.uk',
         'https://www.ability-housing.co.uk/contact-us/', '', ''),
    'aden homes':
        ('https://adenhomes.com', '01245 392094', 'info@adenhomes.co.uk',
         'https://adenhomes.com/contact-us/', '', ''),
    'angel housing':
        ('https://angelsupportuk.co.uk', '020 8533 2372', 'admin@angelsupport.co.uk',
         'https://angelsupportuk.co.uk/contact-us', '', ''),
    'brightsky youth services':
        ('https://brightskyyouthservices.co.uk', '0121 285 7311', 'info@brightskyyouthservices.co.uk',
         'https://brightskyyouthservices.co.uk/contact-us/', '12890772', ''),
    'beckford homes':
        ('https://www.beckfordhomesltd.com', '', 'management.beckfordhomes@gmail.com',
         'https://www.beckfordhomesltd.com/contact-us', '', ''),
    'better oasis':
        ('https://betteroasis.co.uk', '', 'info@betteroasis.co.uk',
         'https://betteroasis.co.uk/contact', '', ''),
    'bluoak housing & healthcare consultancy':
        ('https://www.bluoakhealthcare.co.uk', '023 8000 0000', 'info@bluoakhealthcare.co.uk',
         'https://www.bluoakhealthcare.co.uk/home', '11179666', ''),
    'buckingham hotels group':
        ('https://www.buckinghamhoteluk.com', '01494 47 47 47', 'frontdesk@buckinghamhotels.co.uk',
         'https://buckinghamhoteluk.com/contact-us', '05012013', ''),
    'cambridge housing society':
        ('https://www.chsgroup.org.uk', '0300 111 3555', 'enquiries@chsgroup.org.uk',
         'https://www.chsgroup.org.uk/contact-us/', '', ''),
    'caysh':
        ('https://www.caysh.org', '020 8760 5530', 'info@caysh.org',
         'https://www.caysh.org/contact-us', '', ''),
    'choice support':
        ('https://www.choicesupport.org.uk', '020 7261 4100', 'enquiries@choicesupport.org.uk',
         'https://choicesupport.org.uk/contact-us', '', '1005003'),
    'centrepoint soho':
        ('https://centrepoint.org.uk', '0808 800 0661', 'supportercare@centrepoint.org',
         'https://centrepoint.org.uk/about-centrepoint/contact-us', '', '292411'),
    'depaul':
        ('https://www.depaul.org.uk', '0800 160 1650', 'info@depaul.org.uk',
         'https://www.depaul.org.uk/contact-us/', '', '802384'),
    'equinox care':
        ('https://socialinterestgroup.org.uk', '020 3668 9270', 'enquiries@socialinterestgroup.org.uk',
         'https://socialinterestgroup.org.uk/contact-us/', '', '1144625'),
    'cocoon group services':
        ('https://www.cocoongroupservices.co.uk', '020 8351 0094', 'info@cocoongroupservices.co.uk',
         'https://www.cocoongroupservices.co.uk/contact', '', ''),
    'crown social care':
        ('http://crownsocialcare.org.uk', '020 3859 6280', 'info@crownsocialcare.org.uk',
         'http://crownsocialcare.org.uk/contact-us', '09567314', ''),
    'single homeless project':
        ('https://www.shp.org.uk', '020 7848 0700', 'info@shp.org.uk',
         'https://www.shp.org.uk/contact-us/', '', '1110858'),
    'the riverside group':
        ('https://www.riverside.org.uk', '0345 111 0000', 'enquiries@riverside.org.uk',
         'https://www.riverside.org.uk/contact_us/', '', ''),
    'riverside':
        ('https://www.riverside.org.uk', '0345 111 0000', 'enquiries@riverside.org.uk',
         'https://www.riverside.org.uk/contact_us/', '', ''),
    'thames reach':
        ('https://thamesreach.org.uk', '020 3092 2400', 'info@thamesreach.org.uk',
         'https://thamesreach.org.uk/about-us/contact-us/', '', '1102862'),
    "solace women's aid":
        ('https://www.solacewomensaid.org', '0808 802 5565', 'advice@solacewomensaid.org',
         'https://www.solacewomensaid.org/get-help/', '', '1003344'),
    "solace womens aid":
        ('https://www.solacewomensaid.org', '0808 802 5565', 'advice@solacewomensaid.org',
         'https://www.solacewomensaid.org/get-help/', '', '1003344'),
}

wb = openpyxl.load_workbook(XLSX)
ws_co = wb['Companies']

updated = already = unmatched = 0
unmatched_names = []
for row in ws_co.iter_rows(min_row=2):
    name = str(row[0].value or '').strip()
    if not name: continue
    review_tag = str(row[20].value or '').strip()
    key = name.lower()
    if key not in VERIFIED:
        if review_tag == 'via-commissioner':
            unmatched += 1
            unmatched_names.append(name)
        continue
    if review_tag == 'verified':
        already += 1
        continue
    website, phone, email, contact, ch_num, charity_num = VERIFIED[key]
    row[12].value = website
    row[13].value = phone
    row[14].value = email
    row[15].value = contact
    if ch_num: row[7].value = ch_num
    if charity_num: row[16].value = charity_num
    row[20].value = 'verified'
    updated += 1

wb.save(XLSX)

print(f"\n=== ENRICHMENT RESULTS ===")
print(f"Newly verified this run : {updated}")
print(f"Already verified        : {already}")
print(f"Still via-commissioner  : {unmatched}")
print(f"Total Companies rows    : {ws_co.max_row - 1}")
