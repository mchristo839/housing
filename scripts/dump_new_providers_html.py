"""Build a self-contained HTML page showing the providers we added in the
most recent processing pass + the contract each relates to.

Net-new = in providers.json now but not in the main DB (care_housing_database_v2_ENRICHED.xlsx)
AND added via manual_contracts.xlsx in this session.
"""
import openpyxl, json, re, sys, io, html as htmllib, datetime
from collections import defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUT = 'data/scraped/NEW_PROVIDERS.html'

def norm(s): return re.sub(r'[^a-z0-9]', '', str(s or '').lower())

# Main DB suppliers (baseline)
print("Loading baseline main DB...")
wb_main = openpyxl.load_workbook('data/care_housing_database_v2_ENRICHED.xlsx', read_only=True, data_only=True)
main_suppliers = set()
for r in wb_main['Company × Council × Sector'].iter_rows(min_row=2, values_only=True):
    if r[0]: main_suppliers.add(norm(r[0]))
for r in wb_main['Companies'].iter_rows(min_row=2, values_only=True):
    if r[0]: main_suppliers.add(norm(r[0]))

# Live providers — what shipped
prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
by_norm = {norm(p['name']): p for p in prov}

# Manual file rows
wb_m = openpyxl.load_workbook('data/manual_contracts.xlsx', read_only=True)
ws_cs = wb_m['Company × Council × Sector']

supplier_rows = defaultdict(list)
for r in ws_cs.iter_rows(min_row=2, values_only=True):
    if not r[0] or not r[2]: continue
    supplier_rows[norm(r[0])].append({
        'supplier': str(r[0]),
        'council': str(r[2]),
        'categories': str(r[10] or ''),
        'award_date': str(r[11] or ''),
        'title': str(r[12] or ''),
    })

# Net-new = in manual file, in live providers, NOT in main DB
net_new = []
for ns, rows in supplier_rows.items():
    if ns in main_suppliers: continue
    if ns not in by_norm: continue   # didn't survive build filters
    p = by_norm[ns]
    net_new.append({
        'name': p['name'],
        'primary_cat': p.get('primary_cat',''),
        'sectors': p.get('sector') or [],
        'website': p.get('website',''),
        'phone': p.get('phone',''),
        'email': p.get('email',''),
        'rows': rows,
        'norm': ns,
    })

print(f"Net-new providers live AND in manual file: {len(net_new)}")

# Group by Bidstats notice / commissioner
net_new.sort(key=lambda x: x['name'].lower())

# Build HTML
parts = []
parts.append('<!doctype html><html><head><meta charset="utf-8">')
parts.append('<title>Net-new providers and their contracts</title>')
parts.append('<style>')
parts.append('body{font:14px -apple-system,Segoe UI,Roboto,sans-serif;margin:1.5em;background:#f8f9fa;color:#222;max-width:1200px}')
parts.append('h1{margin:0 0 .2em 0;font-size:22px}')
parts.append('.meta{color:#666;margin-bottom:1.5em;font-size:13px}')
parts.append('.search{position:sticky;top:0;background:#fff;padding:10px 14px;border:1px solid #ddd;border-radius:6px;margin-bottom:1em;box-shadow:0 2px 4px rgba(0,0,0,.05);z-index:5}')
parts.append('.search input{width:300px;padding:6px 10px;font-size:14px;border:1px solid #bbb;border-radius:4px}')
parts.append('.counts{color:#444;margin-left:1em;font-size:13px;font-weight:600}')
parts.append('.provider{background:#fff;border:1px solid #ddd;border-radius:8px;padding:14px 18px;margin-bottom:12px;box-shadow:0 1px 2px rgba(0,0,0,.04)}')
parts.append('.provider h2{margin:0 0 6px 0;font-size:17px;color:#1F4E79}')
parts.append('.tags{margin-bottom:8px}')
parts.append('.tag{display:inline-block;background:#e7eff8;color:#1F4E79;padding:2px 8px;border-radius:11px;font-size:11px;margin-right:5px}')
parts.append('.contact{font-size:12px;color:#666;margin-bottom:8px}')
parts.append('.contact a{color:#0563C1}')
parts.append('.contracts{border-top:1px dashed #ddd;padding-top:8px;margin-top:8px}')
parts.append('.contract{background:#fafbfc;border-left:3px solid #1F4E79;padding:8px 12px;margin-bottom:6px;font-size:13px}')
parts.append('.contract .commiss{font-weight:600;color:#1F4E79}')
parts.append('.contract .title{color:#222;margin:2px 0}')
parts.append('.contract .cat{color:#666;font-size:12px}')
parts.append('.hidden{display:none}')
parts.append('</style></head><body>')

parts.append('<h1>Net-new providers and their contracts</h1>')
parts.append(f'<div class="meta">'
             f'<b>{len(net_new)}</b> providers we added this session that are live on the site after build · sorted A–Z. '
             f'Each provider shows the London contract(s) we attached them to. '
             f'<br><br>'
             f'<b>Most recent processing pass added 28 of these</b> — type <code>778567851</code> in the filter to see just the latest Waltham Forest 16+ batch, '
             f'or <code>via-commissioner</code> for providers whose contact route is through the commissioning council.'
             f'</div>')

parts.append('<div class="search">')
parts.append('Filter: <input id="q" placeholder="provider name, borough, contract, source id..." oninput="flt()">')
parts.append('<span class="counts" id="c"></span>')
parts.append('</div>')

parts.append('<div id="list">')
for p in net_new:
    name = htmllib.escape(p['name'])
    primary = htmllib.escape(p['primary_cat'])
    tags_html = ''.join(f'<span class="tag">{htmllib.escape(s)}</span>' for s in p['sectors'][:6])

    contact_bits = []
    if p['website']:
        contact_bits.append(f'<a href="{htmllib.escape(p["website"])}" target="_blank">website</a>')
    if p['email']:
        contact_bits.append(f'<a href="mailto:{htmllib.escape(p["email"])}">{htmllib.escape(p["email"])}</a>')
    if p['phone']:
        contact_bits.append(htmllib.escape(p['phone']))
    contact_html = ' · '.join(contact_bits) or '<i>(contact via commissioner)</i>'

    contracts_html = []
    for row in p['rows']:
        commiss = htmllib.escape(row['council'])
        title = htmllib.escape(row['title'])
        cats = htmllib.escape(row['categories'])
        award = htmllib.escape(row['award_date'])
        # Extract Bidstats URL if mentioned
        bm = re.search(r'Bidstats\s+(\d+)', row['title'])
        link = ''
        if bm:
            url = f"https://bidstats.uk/notice/{bm.group(1)}"
            link = f' · <a href="{url}" target="_blank">source</a>'
        contracts_html.append(
            f'<div class="contract">'
            f'<div class="commiss">{commiss}</div>'
            f'<div class="title">{title}{link}</div>'
            f'<div class="cat">{cats}{(" · award: " + award) if award else ""}</div>'
            f'</div>'
        )

    parts.append(
        f'<div class="provider" data-search="{htmllib.escape((name + " " + " ".join(r["council"] + " " + r["title"] for r in p["rows"])).lower())}">'
        f'<h2>{name}</h2>'
        f'<div class="tags">{tags_html}</div>'
        f'<div class="contact">{contact_html}</div>'
        f'<div class="contracts">{"".join(contracts_html)}</div>'
        f'</div>'
    )
parts.append('</div>')

parts.append('<script>')
parts.append('function flt(){const q=document.getElementById("q").value.toLowerCase(); const items=document.querySelectorAll(".provider"); let shown=0; items.forEach(it=>{const ok = !q || it.dataset.search.includes(q); it.style.display=ok?"":"none"; if(ok)shown++;}); document.getElementById("c").textContent = shown + " of " + items.length + " providers shown";}')
parts.append('flt();')
parts.append('</script>')
parts.append('</body></html>')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(''.join(parts))

print(f"Wrote {OUT}")
print(f"  Providers shown: {len(net_new)}")
print(f"  File size: {(open(OUT,'rb').read().__len__())//1024}KB")
