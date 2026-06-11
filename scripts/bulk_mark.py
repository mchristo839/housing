"""
Bulk-mark rows in REVIEW_PACK.xlsx by contract source_id or by quality flag.

Usage:
  python scripts/bulk_mark.py --auto-flags
       Pre-fills column A based on the Quality flag column:
         MISCLASSIFIED -> N
         FRAMEWORK     -> ?  (needs supplier extraction first)
         PLACEHOLDER   -> ?

  python scripts/bulk_mark.py --approve <source_id>[,<source_id>...]
       Stamps Y on every row whose Source ID matches.

  python scripts/bulk_mark.py --reject <source_id>[,<source_id>...]
       Stamps N on every row whose Source ID matches.

  python scripts/bulk_mark.py --status
       Prints current Y/N/?/blank counts and a per-contract roll-up.
"""
import openpyxl, sys, io, argparse
from collections import Counter, defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

REVIEW = 'data/scraped/REVIEW_PACK.xlsx'

# Review-pack column indices (1-based, openpyxl style)
# 1 Approve?  2 Quality flag  3 Notes  4 Borough/Body  5 Supplier
# 6 Contract Title  7 Categories  8 Award  9 Value  10 Scope
# 11 Portal  12 URL  13 Source ID  14 Council (raw)  15 Scraped At
COL_APPROVE = 1
COL_FLAG    = 2
COL_BOROUGH = 4
COL_SUPPLIER = 5
COL_TITLE   = 6
COL_SID     = 13


def load():
    wb = openpyxl.load_workbook(REVIEW)
    ws = wb['Review']
    return wb, ws


def auto_flags():
    wb, ws = load()
    n_n = n_q = 0
    for row in ws.iter_rows(min_row=2):
        flag = (row[COL_FLAG-1].value or '').upper()
        cur  = (row[COL_APPROVE-1].value or '').upper()
        if cur:        # don't overwrite user's existing mark
            continue
        if flag.startswith('MISCLASSIFIED'):
            row[COL_APPROVE-1].value = 'N'
            n_n += 1
        elif flag.startswith('FRAMEWORK') or flag.startswith('PLACEHOLDER'):
            row[COL_APPROVE-1].value = '?'
            n_q += 1
    wb.save(REVIEW)
    print(f"Auto-marked: {n_n} as N (misclassified), {n_q} as ? (frameworks / placeholders)")


def mark_by_sid(sids, mark):
    wb, ws = load()
    sids = {s.strip() for s in sids}
    hits = 0
    for row in ws.iter_rows(min_row=2):
        sid = str(row[COL_SID-1].value or '').strip()
        if sid in sids:
            row[COL_APPROVE-1].value = mark
            hits += 1
    wb.save(REVIEW)
    print(f"Set {hits} rows to {mark} for source_id(s): {', '.join(sids)}")


def status():
    wb, ws = load()
    counts = Counter()
    per_contract = defaultdict(lambda: {'Y':0,'N':0,'?':0,'_':0,'title':'','council':''})
    for row in ws.iter_rows(min_row=2):
        m = (row[COL_APPROVE-1].value or '').strip().upper()
        if m in ('Y','YES'):    k='Y'
        elif m in ('N','NO'):   k='N'
        elif m in ('?','MAYBE'):k='?'
        else:                   k='_'
        counts[k] += 1
        sid = str(row[COL_SID-1].value or '').strip()
        ckey = sid if sid else f"NOSID:{row[COL_BOROUGH-1].value}|{row[COL_TITLE-1].value or ''}"[:80]
        per_contract[ckey][k] += 1
        if not per_contract[ckey]['title']:
            per_contract[ckey]['title'] = str(row[COL_TITLE-1].value or '')[:60]
            per_contract[ckey]['council'] = str(row[COL_BOROUGH-1].value or '')[:25]
    total = sum(counts.values())
    print(f"\nREVIEW STATUS ({total} rows)")
    print(f"  Approved (Y) : {counts['Y']:4d}")
    print(f"  Rejected (N) : {counts['N']:4d}")
    print(f"  Maybe (?)    : {counts['?']:4d}")
    print(f"  Unmarked     : {counts['_']:4d}")
    print(f"\nPER-CONTRACT (sorted by undecided row count, top 20):")
    print(f"  {'Y':>3s} {'N':>3s} {'?':>3s} {'_':>3s}  {'Council':<25s}  Contract")
    items = sorted(per_contract.items(), key=lambda kv: (-kv[1]['_'], -(kv[1]['Y']+kv[1]['N']+kv[1]['?'])))
    for sid, v in items[:25]:
        sid_disp = sid if len(sid) <= 12 else sid[:9]+'...'
        print(f"  {v['Y']:>3d} {v['N']:>3d} {v['?']:>3d} {v['_']:>3d}  {v['council']:<25s}  {v['title']}")


def main():
    ap = argparse.ArgumentParser()
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument('--auto-flags', action='store_true')
    g.add_argument('--approve')
    g.add_argument('--reject')
    g.add_argument('--maybe')
    g.add_argument('--status', action='store_true')
    a = ap.parse_args()
    if a.auto_flags:   auto_flags()
    elif a.approve:    mark_by_sid(a.approve.split(','), 'Y')
    elif a.reject:     mark_by_sid(a.reject.split(','), 'N')
    elif a.maybe:      mark_by_sid(a.maybe.split(','), '?')
    elif a.status:     status()

if __name__ == '__main__':
    main()
