"""Backfill the umbrella title onto Waltham Forest CCS rows that only have
the bare Bidstats citation (778567851).

The umbrella is verified:
  "16+ Semi-Independent Accommodation & Support Services to Looked After
   Children and Care Experienced Young Adults"
  London Borough of Waltham Forest — £16M framework, awarded 12 Jul 2022
  Bidstats notice 778567851 / Find a Tender
"""
import openpyxl, re

UMBRELLA = (
    "16+ Semi-Independent Accommodation & Support Services to Looked After "
    "Children and Care Experienced Young Adults — London Borough of Waltham "
    "Forest framework (£16M, 12 Jul 2022 award). "
    "Source: Bidstats 778567851 / Find a Tender."
)

wb = openpyxl.load_workbook('data/manual_contracts.xlsx')
ccs = wb['Company × Council × Sector']
fixed = 0
for r in range(2, ccs.max_row + 1):
    council = (ccs.cell(row=r, column=3).value or '').lower()
    title = (ccs.cell(row=r, column=13).value or '')
    if 'waltham forest' in council and '778567851' in title:
        # Only backfill if title is just the bare citation
        stripped = title.strip()
        if stripped.startswith('(Bidstats') or stripped.startswith('Bidstats'):
            ccs.cell(row=r, column=13).value = UMBRELLA
            fixed += 1

wb.save('data/manual_contracts.xlsx')
print(f"Backfilled {fixed} Waltham Forest CCS rows with umbrella title")
