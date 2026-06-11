"""4-step verification process for housing providers.

USAGE:
  python scripts/verify_provider_v2.py queue              # generate queue CSV
  python scripts/verify_provider_v2.py status             # show progress
  python scripts/verify_provider_v2.py apply              # apply pending decisions

THE 4 STEPS (per provider):
  Step 1 — Website ownership: Fetch stored website. Does the limited company
           name (or close variant) appear anywhere on the homepage / about / footer?
           ✓ Match → Step 3
           ✗ No match → Step 2

  Step 2 — Find the right website: Web search
              "[Company Name Limited]" supported living OR assisted living
              OR asylum OR homelessness OR supported accommodation
           Pick the result whose URL contains the company keyword and
           whose page mentions the company name. Update stored website.

  Step 3 — Service confirmation: Does the verified website mention at least
           one of: supported living, assisted living, supported accommodation,
           extra care, residential care, asylum, homelessness, refuge, hostel,
           sheltered housing, young people 16+, care leavers, mental health
           accommodation? If YES → KEEP. If NO → DROP.

  Step 4 — Contact details: Verify website / phone / email / address.
           - No email? Look for "Contact Us" page URL.
           - No phone? Same.
           Write the verified set to manual_contracts.xlsx Companies sheet
           so they survive every monthly rebuild.

OUTPUTS:
  data/scraped/VERIFICATION_V2_QUEUE.csv  — queue with status per provider
  data/scraped/VERIFICATION_V2_RESULTS.csv — full audit trail with all 4 steps
  data/MANUAL_DROP_LIST.json              — only updated for confirmed Step-3 fails
  data/manual_contracts.xlsx              — Companies sheet updates for fixes
"""
import csv, json, re, sys
from pathlib import Path

QUEUE_CSV   = Path('data/scraped/VERIFICATION_V2_QUEUE.csv')
RESULTS_CSV = Path('data/scraped/VERIFICATION_V2_RESULTS.csv')
DROP_LIST   = Path('data/MANUAL_DROP_LIST.json')

# -------- service keywords that constitute a "housing-related service" --------
HOUSING_TERMS = [
    'supported living', 'assisted living', 'supported accommodation',
    'supported housing', 'extra care', 'residential care', 'care home',
    'nursing home', 'asylum', 'homelessness', 'homeless', 'refuge',
    'hostel', 'sheltered housing', 'care leaver', 'leaving care',
    '16+', 'young people', 'children\'s home', 'children home',
    'mental health accommodation', 'transitional accommodation',
    'semi-independent', 'semi independent', 'rough sleeping',
    'temporary accommodation', 'crisis accommodation', 'housing first',
    'domestic abuse', 'women\'s refuge', 'family accommodation',
]

# -------- terms that mean clearly NOT a housing provider ----------------------
NON_HOUSING_TERMS = [
    'pizza', 'takeaway', 'restaurant', 'cafe', 'food delivery',
    'vet practice', 'veterinary', 'animal hospital',
    'building services', 'plumbing', 'heating', 'electrician',
    'plastics manufacturing', 'packaging', 'recycling bins',
    'department store', 'retail outlet', 'clothing store',
    'recruitment agency', 'staffing solutions', 'temp agency',
    'training academy', 'training courses', 'school of nursing',
    'software development', 'web design', 'digital marketing',
    'procurement consultancy', 'office supplies', 'stationery',
]

def cmd_queue():
    """Generate the verification queue (one row per provider currently in DB)."""
    prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
    # Skip those already verified
    already = set()
    if RESULTS_CSV.exists():
        with open(RESULTS_CSV, encoding='utf-8') as f:
            for r in csv.DictReader(f):
                already.add(r['name'])

    rows = []
    for p in prov:
        if p['name'] in already: continue
        rows.append({
            'name': p['name'],
            'current_website': p.get('website',''),
            'current_phone':   p.get('phone',''),
            'current_email':   p.get('email',''),
            'current_address': p.get('address',''),
            'contracts':       len(p.get('contracts_list') or []),
            'scope':           p.get('scope',''),
            'status':          'pending',
            'step1_url_owner': '',  # Y / N / NO_SITE
            'step2_real_url':  '',  # populated if step1=N and we found one
            'step3_services':  '',  # comma-sep of matched HOUSING_TERMS
            'step4_email':     '',  # verified email
            'step4_phone':     '',  # verified phone
            'step4_contact':   '',  # /contact-us URL fallback
            'verdict':         '',  # KEEP / DROP / UPDATE
            'notes':           '',
        })

    QUEUE_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(QUEUE_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else [])
        w.writeheader()
        w.writerows(rows)

    print(f"Queue written: {QUEUE_CSV}")
    print(f"  Pending:  {len(rows)}")
    print(f"  Already:  {len(already)}")
    print(f"\nNext step: work through the queue 10-15 at a time.")
    print(f"For each provider, the assistant will:")
    print(f"  1. WebFetch its current_website and check for company name match")
    print(f"  2. If no match → WebSearch '[Company Name] supported living' to find real site")
    print(f"  3. Check the verified site lists a housing service (KEEP) or not (DROP)")
    print(f"  4. Extract verified email/phone/contact page")
    print(f"\nResults recorded to: {RESULTS_CSV}")

def cmd_status():
    """Print progress summary."""
    if not RESULTS_CSV.exists():
        print("No results yet.")
        return
    with open(RESULTS_CSV, encoding='utf-8') as f:
        rows = list(csv.DictReader(f))
    from collections import Counter
    verdicts = Counter(r['verdict'].upper() for r in rows)
    print(f"Verified so far: {len(rows)}")
    for v, n in verdicts.most_common():
        print(f"  {v}: {n}")
    # Step-level stats
    s1_y = sum(1 for r in rows if r['step1_url_owner'] == 'Y')
    s1_n = sum(1 for r in rows if r['step1_url_owner'] == 'N')
    s2_fixed = sum(1 for r in rows if r['step2_real_url'])
    print(f"\nStep 1 — website matched company name: {s1_y} Y / {s1_n} N")
    print(f"Step 2 — real URL discovered:           {s2_fixed}")

def cmd_apply():
    """Apply pending decisions from RESULTS_CSV → drop list + manual_contracts."""
    if not RESULTS_CSV.exists():
        print("No results to apply.")
        return
    import openpyxl
    with open(RESULTS_CSV, encoding='utf-8') as f:
        rows = list(csv.DictReader(f))

    drops = [r['name'] for r in rows if r['verdict'].upper() == 'DROP']
    updates = [r for r in rows if r['verdict'].upper() == 'UPDATE']

    # Apply drops
    if drops:
        dl = json.load(open(DROP_LIST, encoding='utf-8'))
        before = len(dl)
        for n in drops:
            if n not in dl: dl.append(n)
        json.dump(dl, open(DROP_LIST, 'w', encoding='utf-8'), indent=2)
        print(f"Drops applied: {before} -> {len(dl)} (+{len(dl)-before})")

    # Apply contact updates to Companies sheet
    if updates:
        wb = openpyxl.load_workbook('data/manual_contracts.xlsx')
        comp = wb['Companies']
        idx_by_name = {}
        for r in range(2, comp.max_row+1):
            nm = (comp.cell(row=r, column=1).value or '').strip().lower()
            if nm: idx_by_name[nm] = r

        n_upd = n_add = 0
        for r in updates:
            nl = r['name'].lower().strip()
            row_idx = idx_by_name.get(nl)
            if row_idx is None:
                row_idx = comp.max_row + 1
                comp.cell(row=row_idx, column=1).value = r['name']
                comp.cell(row=row_idx, column=2).value = 0
                comp.cell(row=row_idx, column=5).value = 1
                comp.cell(row=row_idx, column=7).value = 'No'
                idx_by_name[nl] = row_idx
                n_add += 1
            else:
                n_upd += 1
            # Cols 12=address 13=website 14=phone 15=email
            if r['current_address'] or r.get('step4_address'):
                comp.cell(row=row_idx, column=12).value = r.get('step4_address') or r['current_address']
            if r['step2_real_url']:
                comp.cell(row=row_idx, column=13).value = r['step2_real_url']
            elif r['current_website']:
                comp.cell(row=row_idx, column=13).value = r['current_website']
            if r['step4_phone']:
                comp.cell(row=row_idx, column=14).value = r['step4_phone']
            if r['step4_email']:
                comp.cell(row=row_idx, column=15).value = r['step4_email']
        wb.save('data/manual_contracts.xlsx')
        print(f"Companies sheet: {n_upd} updated, {n_add} added")

    print("\nDone. Now rebuild + deploy:")
    print("  python build_data.py && npm run build && vercel deploy --prod --yes")

# -----------------------------------------------------------------------------
def main():
    cmd = sys.argv[1] if len(sys.argv) > 1 else 'queue'
    if   cmd == 'queue':  cmd_queue()
    elif cmd == 'status': cmd_status()
    elif cmd == 'apply':  cmd_apply()
    else:
        print(__doc__)

if __name__ == '__main__':
    main()
