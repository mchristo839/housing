"""Export the full provider database to:
  1. An Excel file with one row per provider — Provider × Contracts × Contacts
  2. A browseable HTML page you can open offline

Goes to your Desktop so you can review/search/filter as the owner."""
import json, openpyxl, html as htmllib, os
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

DESK = Path(os.path.expanduser('~/OneDrive/Desktop'))
DESK.mkdir(parents=True, exist_ok=True)

prov = json.load(open('api/_data/providers.json', encoding='utf-8'))

# ── 1. Excel export ─────────────────────────────────────────────────────────
xlsx_path = DESK / 'PROVIDER_DATABASE.xlsx'
wb = openpyxl.Workbook()

# Sheet 1: Providers (one row each, with summary)
s1 = wb.active
s1.title = 'Providers'
headers = ['Name', 'Scope', 'Primary Category', 'Sectors',
           'Website', 'Phone', 'Email', 'Address',
           '# Councils', 'In-Network', 'SME',
           'Charity Number', 'Charity Income', 'Companies House']
s1.append(headers)
for c in s1[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F4E79')
    c.alignment = Alignment(horizontal='center')

for p in sorted(prov, key=lambda x: (x.get('scope',''), x['name'].lower())):
    cs = p.get('contracts_list') or []
    s1.append([
        p['name'],
        p.get('scope',''),
        p.get('primary_cat',''),
        ' | '.join(p.get('sector') or []),
        p.get('website','') or '',
        p.get('phone','') or '',
        p.get('email','') or '',
        p.get('address','') or '',
        len(cs),
        'Yes' if p.get('in_network') else '',
        'Yes' if p.get('is_sme') else ('No' if p.get('is_sme') is False else ''),
        (p.get('charity') or {}).get('number','') or '',
        (p.get('charity') or {}).get('income','') or '',
        '',  # CH not in providers.json, leave blank
    ])

# autosize
for col_idx in range(1, len(headers)+1):
    s1.column_dimensions[get_column_letter(col_idx)].width = 22
s1.column_dimensions['A'].width = 40
s1.column_dimensions['D'].width = 40
s1.column_dimensions['E'].width = 35
s1.column_dimensions['G'].width = 32
s1.column_dimensions['H'].width = 50
s1.freeze_panes = 'A2'

# Sheet 2: Contracts (one row per contract = provider × council)
s2 = wb.create_sheet('Contracts')
s2.append(['Provider', 'Council', 'Scope', 'Region', 'Contract Title'])
for c in s2[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F4E79')
    c.alignment = Alignment(horizontal='center')
for p in prov:
    for c in (p.get('contracts_list') or []):
        title = c.get('titles') or ''
        if isinstance(title, list): title = ' | '.join(str(x) for x in title)
        s2.append([
            p['name'],
            c.get('council',''),
            c.get('scope',''),
            c.get('region','') or c.get('county','') or '',
            title,
        ])
for col_idx in range(1, 6):
    s2.column_dimensions[get_column_letter(col_idx)].width = 25
s2.column_dimensions['A'].width = 40
s2.column_dimensions['B'].width = 40
s2.column_dimensions['E'].width = 80
s2.freeze_panes = 'A2'

wb.save(xlsx_path)
print(f"Wrote {xlsx_path}  ({xlsx_path.stat().st_size//1024} KB)")

# ── 2. Browseable HTML ─────────────────────────────────────────────────────
html_path = DESK / 'PROVIDER_DATABASE.html'
parts = ['<!doctype html><html><head><meta charset="utf-8">',
         '<title>Find a Housing Provider — full database export</title>',
         '''<style>
body{font:14px -apple-system,Segoe UI,Roboto,sans-serif;margin:0;background:#f6f7f9;color:#1a1a1a}
.bar{background:#1F4E79;color:#fff;padding:18px 24px;position:sticky;top:0;z-index:9}
.bar input{width:100%;max-width:420px;padding:9px 12px;border:1px solid #fff;border-radius:6px;font-size:14px}
.bar .stats{font-size:13px;margin-top:6px;opacity:0.85}
.wrap{padding:18px 24px}
.tabs{display:flex;gap:8px;margin-bottom:12px}
.tabs button{padding:7px 14px;border:1px solid #1F4E79;background:#fff;color:#1F4E79;border-radius:6px;cursor:pointer;font-weight:600;font-size:13px}
.tabs button.active{background:#1F4E79;color:#fff}
.scope-pill{display:inline-block;padding:2px 8px;border-radius:11px;font-size:10px;font-weight:700;letter-spacing:.05em;text-transform:uppercase}
.scope-Local{background:#d4edda;color:#1e6033}
.scope-County{background:#fff3cd;color:#856404}
.scope-Regional{background:#d1ecf1;color:#0c5460}
.scope-National{background:#e2e3e5;color:#383d41}
table{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06)}
th,td{text-align:left;padding:10px 12px;border-bottom:1px solid #e5e7eb;font-size:13px;vertical-align:top}
th{background:#f4f7fa;font-weight:700;color:#1F4E79;position:sticky;top:115px;cursor:pointer}
th:hover{background:#e7eff8}
tr:hover{background:#fafbfc}
a{color:#0563C1;text-decoration:none}
a:hover{text-decoration:underline}
.tag{display:inline-block;background:#eef2f7;color:#3c4858;padding:1px 6px;border-radius:9px;font-size:10px;margin-right:3px}
.title-cell{font-size:11px;color:#555;max-width:400px}
#contracts{display:none}
.toggle{padding:4px 10px;background:#fff;border:1px solid #1F4E79;color:#1F4E79;border-radius:4px;cursor:pointer;font-size:11px}
</style></head><body>''']
parts.append('<div class="bar"><div><b>Find a Housing Provider</b> — Full database export</div>')
parts.append('<input id="search" placeholder="Search anything (name, council, contract title)…" />')
parts.append(f'<div class="stats">{len(prov):,} providers · {sum(len(p.get("contracts_list") or []) for p in prov):,} contracts</div>')
parts.append('</div>')
parts.append('<div class="wrap">')
parts.append('<div class="tabs"><button id="tab_providers" class="active">Providers</button>'
             '<button id="tab_contracts">Contracts</button></div>')

# Providers table
parts.append('<table id="providers"><thead><tr>')
for h in ['Name','Scope','Sectors','Website','Phone','Email','Councils']:
    parts.append(f'<th>{h}</th>')
parts.append('</tr></thead><tbody>')
for p in sorted(prov, key=lambda x: x['name'].lower()):
    nm = htmllib.escape(p['name'])
    sc = p.get('scope','')
    sectors = ''.join(f'<span class="tag">{htmllib.escape(s)}</span>' for s in (p.get('sector') or [])[:5])
    w = p.get('website','') or ''
    w_html = f'<a href="{htmllib.escape(w)}" target="_blank">{htmllib.escape(w[:42])}</a>' if w else ''
    ph = htmllib.escape(p.get('phone','') or '')
    em = p.get('email','') or ''
    em_html = f'<a href="mailto:{htmllib.escape(em)}">{htmllib.escape(em)}</a>' if em else ''
    n_cs = len(p.get('contracts_list') or [])
    parts.append(f'<tr><td><b>{nm}</b></td>'
                 f'<td><span class="scope-pill scope-{sc}">{sc}</span></td>'
                 f'<td>{sectors}</td>'
                 f'<td>{w_html}</td>'
                 f'<td>{ph}</td>'
                 f'<td>{em_html}</td>'
                 f'<td>{n_cs}</td></tr>')
parts.append('</tbody></table>')

# Contracts table
parts.append('<table id="contracts"><thead><tr>')
for h in ['Provider','Council','Scope','Region','Contract Title']:
    parts.append(f'<th>{h}</th>')
parts.append('</tr></thead><tbody>')
for p in sorted(prov, key=lambda x: x['name'].lower()):
    for c in (p.get('contracts_list') or []):
        title = c.get('titles') or ''
        if isinstance(title, list): title = ' | '.join(str(x) for x in title)
        parts.append(f'<tr><td><b>{htmllib.escape(p["name"])}</b></td>'
                     f'<td>{htmllib.escape(c.get("council","") or "")}</td>'
                     f'<td><span class="scope-pill scope-{c.get("scope","")}">{c.get("scope","")}</span></td>'
                     f'<td>{htmllib.escape(c.get("region","") or c.get("county","") or "")}</td>'
                     f'<td class="title-cell">{htmllib.escape(title[:250])}</td></tr>')
parts.append('</tbody></table>')
parts.append('</div>')

# JS — search + tab toggle
parts.append('''<script>
const search = document.getElementById('search');
const providers = document.getElementById('providers');
const contracts = document.getElementById('contracts');
function filt(tbl) {
  const q = search.value.toLowerCase();
  let n = 0;
  for (const tr of tbl.tBodies[0].rows) {
    const txt = tr.textContent.toLowerCase();
    if (!q || txt.includes(q)) { tr.style.display=''; n++; } else { tr.style.display='none'; }
  }
  return n;
}
search.addEventListener('input', () => { filt(providers); filt(contracts); });

document.getElementById('tab_providers').onclick = () => {
  providers.style.display=''; contracts.style.display='none';
  document.getElementById('tab_providers').classList.add('active');
  document.getElementById('tab_contracts').classList.remove('active');
};
document.getElementById('tab_contracts').onclick = () => {
  providers.style.display='none'; contracts.style.display='';
  document.getElementById('tab_contracts').classList.add('active');
  document.getElementById('tab_providers').classList.remove('active');
};
</script></body></html>''')
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(''.join(parts))
print(f"Wrote {html_path}  ({html_path.stat().st_size//1024} KB)")
