"""
Promote rows the user has marked Approve=Y in REVIEW_PACK.xlsx into the
live manual additions file (data/manual_contracts.xlsx).

Workflow:
  1. User edits data/scraped/REVIEW_PACK.xlsx and sets column A "Approve?"
     to Y for rows they want included on the live site.
  2. User runs:    python scripts/promote_approved.py
  3. This script:
       - reads REVIEW_PACK.xlsx, keeps rows where Approve? in {Y, YES}
       - looks up the full row in curated_scraped.xlsx using
         (Supplier, Council, Source ID) as the key
       - appends those rows (first 17 columns — the schema build_data.py
         reads) to data/manual_contracts.xlsx, creating that file with
         the correct sheet + header if it doesn't exist yet
       - dedups against rows already present in manual_contracts.xlsx
       - stamps approved_at back into curated_scraped.xlsx so we know
         which rows have already been promoted

After this, run:
       python build_data.py
       git add -A && git commit -m "promote N approved providers"
       git push      # Vercel auto-deploys

The live site picks them up on the next build.
"""
import openpyxl
import sys
import io
import os
import datetime
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

REVIEW   = 'data/scraped/REVIEW_PACK.xlsx'
SOURCE   = 'data/scraped/curated_scraped.xlsx'
MANUAL   = 'data/manual_contracts.xlsx'
SHEET    = 'Company × Council × Sector'

# Schema build_data.py expects (first 17 cols of the source workbook).
LIVE_HEADERS = [
    'Company', 'Sector', 'Council',
    'Contracts (this council, this sector)',
    'Company Total (all sectors)',
    'Company Total — Homecare',
    'Company Total — Housing',
    'Companies House', 'Is SME', 'Is VCSE',
    'Categories', 'Most Recent Award', 'Contract Titles',
    'ONS Region', 'Commissioner Type', 'Geographic Scope',
    'Asylum Contractor',
]

# ------------------------------------------------------------------ #
# 1. Load review-pack approvals
# ------------------------------------------------------------------ #
if not os.path.exists(REVIEW):
    print(f"ERR: {REVIEW} does not exist. Run scripts/build_review_pack.py first.")
    sys.exit(1)

wb_r = openpyxl.load_workbook(REVIEW, read_only=True)
ws_r = wb_r['Review']

# Review-pack columns (current schema with Quality flag column):
#   0 Approve?  1 Quality flag  2 Notes  3 Borough/Body  4 Supplier
#   5 Title  6 Categories  7 Award  8 Value  9 Scope
#   10 Portal  11 URL  12 Source ID  13 Council(raw)  14 Scraped At
approvals = {}     # key = (supplier_lc, council_raw_lc, source_id_lc) -> notes
rejections = 0
unmarked  = 0
maybes    = 0
for row in ws_r.iter_rows(min_row=2, values_only=True):
    if not row[4]:
        continue
    flag = str(row[0] or '').strip().upper()
    supplier = str(row[4]).strip()
    council  = str(row[13] or '').strip()
    src_id   = str(row[12] or '').strip()
    key = (supplier.lower(), council.lower(), src_id.lower())
    if flag in ('Y', 'YES'):
        approvals[key] = row[2] or ''
    elif flag in ('N', 'NO'):
        rejections += 1
    elif flag in ('?', 'MAYBE'):
        maybes += 1
    else:
        unmarked += 1

print(f"\nREVIEW PACK STATE")
print(f"  Approved (Y) : {len(approvals)}")
print(f"  Rejected (N) : {rejections}")
print(f"  Maybe (?)    : {maybes}")
print(f"  Unmarked     : {unmarked}")

if not approvals:
    print(f"\nNothing marked Y in column A of {REVIEW}.")
    print(f"Edit that file (Approve? -> Y for rows you want live) and re-run.")
    sys.exit(0)

# ------------------------------------------------------------------ #
# 2. Pull full source rows from curated_scraped.xlsx
# ------------------------------------------------------------------ #
wb_s = openpyxl.load_workbook(SOURCE)
ws_s = wb_s[SHEET]

# build lookup by same triplet
src_rows = {}
src_row_indices = {}     # for stamping approved_at back later
for i, row in enumerate(ws_s.iter_rows(min_row=2, values_only=True), start=2):
    if not row[0] or not row[2]:
        continue
    key = (str(row[0]).strip().lower(),
           str(row[2]).strip().lower(),
           str(row[19] or '').strip().lower())
    src_rows[key] = row
    src_row_indices[key] = i

found    = 0
missing  = 0
to_promote = []
for k in approvals:
    if k in src_rows:
        to_promote.append((k, src_rows[k]))
        found += 1
    else:
        missing += 1

print(f"\nLOOKUP into {SOURCE}")
print(f"  Found    : {found}")
print(f"  Missing  : {missing}  (supplier/council/source-id changed since review pack was generated?)")

# ------------------------------------------------------------------ #
# 3. Append into manual_contracts.xlsx
# ------------------------------------------------------------------ #
if os.path.exists(MANUAL):
    wb_m = openpyxl.load_workbook(MANUAL)
    if SHEET not in wb_m.sheetnames:
        ws_m = wb_m.create_sheet(SHEET)
        ws_m.append(LIVE_HEADERS)
    else:
        ws_m = wb_m[SHEET]
    print(f"\nOpened existing {MANUAL} ({ws_m.max_row-1} existing rows)")
else:
    wb_m = openpyxl.Workbook()
    # remove default sheet
    default = wb_m.active
    wb_m.remove(default)
    ws_m = wb_m.create_sheet(SHEET)
    ws_m.append(LIVE_HEADERS)
    print(f"\nCreating {MANUAL} fresh.")

# dedupe against already-present rows in manual
existing_keys = set()
for row in ws_m.iter_rows(min_row=2, values_only=True):
    if row[0] and row[2]:
        # supplier + council + title — manual file doesn't carry source_id,
        # so use title as the third element of the key
        existing_keys.add((
            str(row[0]).strip().lower(),
            str(row[2]).strip().lower(),
            str(row[12] or '').strip().lower()[:100],
        ))

now = datetime.datetime.now().isoformat()
appended = 0
duplicates = 0

for (key, src) in to_promote:
    # first 17 columns of source ARE the live schema (verified against build_data.py)
    live_row = list(src[:17])
    dedup_key = (str(live_row[0]).strip().lower(),
                 str(live_row[2]).strip().lower(),
                 str(live_row[12] or '').strip().lower()[:100])
    if dedup_key in existing_keys:
        duplicates += 1
        continue
    existing_keys.add(dedup_key)
    ws_m.append(live_row)
    appended += 1
    # stamp approved_at back into source so we don't re-promote later
    src_idx = src_row_indices[key]
    ws_s.cell(row=src_idx, column=24, value=now)  # col 24 = approved_at

wb_m.save(MANUAL)
wb_s.save(SOURCE)

print(f"\nPROMOTION COMPLETE")
print(f"  Appended to {MANUAL} : {appended}")
print(f"  Skipped (duplicate)  : {duplicates}")
print(f"  Total rows now in manual_contracts.xlsx: {ws_m.max_row-1}")

# Per-borough breakdown of what just got promoted
by_borough = Counter()
for (key, src) in to_promote:
    by_borough[str(src[2]).strip()] += 1

if by_borough:
    print(f"\nBy commissioner (top 20):")
    for c, n in by_borough.most_common(20):
        print(f"  +{n:3d}  {c}")

print(f"\nNEXT STEPS")
print(f"  1. python build_data.py")
print(f"  2. git add -A")
print(f"  3. git commit -m 'add {appended} approved London supported-housing providers'")
print(f"  4. git push        # Vercel auto-deploys")
