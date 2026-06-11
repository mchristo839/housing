"""For each of the 10 providers without websites, derive a candidate domain
from their email + name, HTTP-check that the site actually exists at that
address, and add a Companies-sheet override only for the ones that resolve.

Lives in manual_contracts.xlsx so it survives monthly main-DB refreshes.
"""
import openpyxl, json, sys, io, re, urllib.request, ssl, socket
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

MANUAL = 'data/manual_contracts.xlsx'

# Candidates per provider: (name, list of candidate URLs to try in order, phone, email override)
CANDIDATES = [
    # Same trust (post-merger)
    ('2gether NHS Foundation Trust',
        ['https://www.ghc.nhs.uk', 'https://ghc.nhs.uk'],
        '0300 421 8100', 'ghccomms@ghc.nhs.uk'),
    ('Gloucestershire Health and Care NHS Foundation Trust',
        ['https://www.ghc.nhs.uk', 'https://ghc.nhs.uk'],
        '0300 421 8100', 'ghccomms@ghc.nhs.uk'),
    # Councils
    ('Hastings Borough Council',
        ['https://www.hastings.gov.uk'],
        '01424 451066', 'enquiries@hastings.gov.uk'),
    # Charities — derivable from email
    ('The Baca Charity',
        ['https://www.bacacharity.org.uk', 'https://bacacharity.org.uk'],
        '01509 215 444', 'info@bacacharity.org.uk'),
    # Other charities — try canonical names
    ('East Durham Community Initiatives Ltd',
        ['https://www.edci.org.uk', 'https://edci.org.uk',
         'https://www.eastdurhamcommunityinitiatives.org.uk'],
        '0191 581 9499', 'info@edci.org.uk'),
    ('Step by Step',
        ['https://www.stepbystep.org.uk', 'https://stepbystep.org.uk',
         'https://www.stepbystep-uk.org'],
        '01252 346 100', 'info@stepbystep.org.uk'),
    # Council-intake referrers — likely just point at council
    ('Bristol Family Care Ltd',
        ['https://www.bristolfamilycare.co.uk', 'https://bristolfamilycare.co.uk'],
        '0845 129 7217', ''),
    ('Rite Directions Children & Young Peoples Services',
        ['https://www.ritedirections.co.uk', 'https://ritedirections.co.uk',
         'https://www.ritedirections.com'],
        '01723 515562', ''),
    # CQC catch-all (these are likely unfindable typo duplicates)
    ('Supporting Independance',
        ['https://www.supportingindependence.co.uk',
         'https://supportingindependence.org.uk'],
        '01903 944006', ''),
    ('Supporting Independence',
        ['https://www.supportingindependence.co.uk',
         'https://supportingindependence.org.uk'],
        '01903 944006', ''),
]

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

def site_alive(url, timeout=8):
    """HEAD-check (fall back to GET). Returns final URL or '' if dead."""
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/121.0.0.0 Safari/537.36',
            'Accept': 'text/html,*/*'
        }, method='GET')
        with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
            if 200 <= r.getcode() < 400:
                return r.geturl() or url
            return ''
    except urllib.error.HTTPError as e:
        # 403/405/406 from bot detection — site exists, just blocks UA
        if e.code in (403, 405, 406):
            return url
        return ''
    except (urllib.error.URLError, socket.timeout, Exception):
        return ''

# Walk and verify
print(f"{'NAME':50s}  STATUS")
print('=' * 80)
confirmed = []
for name, cands, phone, email in CANDIDATES:
    found = ''
    tried = []
    for c in cands:
        tried.append(c)
        alive = site_alive(c)
        if alive:
            found = c
            break
    if found:
        confirmed.append({'name': name, 'website': found, 'phone': phone, 'email': email})
        print(f"  ✓ {name[:48]:48s}  alive: {found}")
    else:
        print(f"  ✗ {name[:48]:48s}  none of: {', '.join(tried)}")

print(f"\n{len(confirmed)} live websites confirmed of {len(CANDIDATES)} candidates")

# Add Companies entries to manual_contracts.xlsx
if confirmed:
    wb = openpyxl.load_workbook(MANUAL)
    ws_co = wb['Companies']
    existing = {str(r[0].value or '').strip().lower() for r in ws_co.iter_rows(min_row=2) if r[0].value}
    added = 0
    for c in confirmed:
        if c['name'].lower() in existing:
            # already there — update website cell
            for row in ws_co.iter_rows(min_row=2):
                if str(row[0].value or '').strip().lower() == c['name'].lower():
                    row[12].value = c['website']
                    if c['phone']: row[13].value = c['phone']
                    if c['email']: row[14].value = c['email']
                    row[20].value = 'verified'
                    break
            continue
        # Add a fresh entry
        row = [c['name'], 1, 0, '', 1, '', 0, '', 0, 0,
               '', '', c['website'], c['phone'], c['email'], '',
               '', '', '', '', 'verified']
        ws_co.append(row)
        added += 1
    wb.save(MANUAL)
    print(f"\nAdded {added} Companies entries; updated existing for the rest")
    print(f"Total Companies sheet now: {ws_co.max_row - 1} rows")
