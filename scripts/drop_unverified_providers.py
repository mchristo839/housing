"""Add the 55 verified-non-supplier names to build_data.py's drop filter.

Rule: dump anyone who isn't an actual supplier with a contact that relates to them.

Two buckets:
  A. 49 confirmed junk (wrong industry, wrong domain attribution, day services, etc.)
  B. 6 NWADCS-list providers whose only website we have is for a totally
     unrelated business — no verifiable contact relates to them, so by the rule
     they must go until we find a real contact.

Final result: 39 of the original 94 stay (verified on official framework list OR
verified housing/care provider with related contact).
"""

# Confirmed drops — each one verified individually
DROP_LIST = [
    # Wrong attribution: domain belongs to unrelated business
    'Abfallbehalter & Container (A&C) Weber UK Ltd',
    'Biobag Ltd',
    'Bright Avenues Limited',
    'British Polythene Limited',
    'Contenur UK Ltd',
    'Craemer UK Limited',
    'Cromwell Polythene Ltd',
    'ESE World Ltd',
    'Egbert H Taylor & Co Ltd',
    'Fairport Containers Ltd.',
    'H.S Jacksons (Fencing) Limited',
    'J & HM Dickson Ltd',
    'J Hill & Co',
    'J.P. McDougall & Co. Limited t/a Dulux Decorator Centre',
    'James Hargreaves PM Ltd',
    'KB Extruders Ltd',
    'Lesta Packaging PLC',
    'Protec Engineering Services (B,ham) Ltd',
    'Ridley Recycling Ltd T/A Peter Ridley Waste',
    'Smith Bros (Caer Conan) Wholesale Ltd',
    'Stearn Electric Co Ltd',
    'Storm Environmental Ltd',
    'Straight Manufacturing Limited',
    'The Beck Company Ltd',
    'The Premiere Kitchen Company',
    'Travis Perkins',
    'UK Container Maintenance Limited',
    'Unicorn Containers Limited',
    'Valley House',                  # windrushvalleyhouseclearance
    'Vision Gelpack Limited',
    'Wolseley UK Ltd',
    'RS Components Limited',
    'HPC Healthline (UK) Ltd',
    'Wilko Retail Limited',
    'in Out and About Limited',
    'Capital Letters',               # procurement vehicle, not a provider

    # Wrong attribution: domain unrelated, no verifiable real housing operator
    'MCSE Ltd',                      # midcareservices = homecare only
    'Midco Care',                    # same
    'Transition 360 Limited',        # was on NWADCS but check failed
    'Transforming Lives Co',
    'Elite Care UK',                 # recruitment domain, can't verify

    # Tagged NWADCS in our data but NOT on the official NWADCS list
    'ALC Supported Living Ltd',
    'Positive Culture Limited',
    'Pathway forward',
    'Strawberry Fields Training Cic',
    'Trinity Ind. Fostering Organisation Ltd',
    'The Integral Care Hub Ltd',
    'casi care',

    # Not housing services (day care, day service, recruitment, etc.)
    'Roseleigh Day Care',
    'Support and Connections Limited',
    'Care Success Solutions Ltd',

    # On official NWADCS list BUT only website we have is unrelated business —
    # no verifiable contact that relates to them, so per the rule they go until
    # we can find a real direct contact.
    'Altus Social Ltd',                  # altussearch.co.uk = recruitment
    'Avensis Support Ltd',               # avensishospitality.co.uk = hospitality
    'Dynamis Enterprises Ltd',           # dynamiseducation.co.uk = training
    'Horizons Plus',                     # mccarthyandstoneresales.co.uk = wrong
    'Revolve Therapy and Training',      # revolve-leadership.com = leadership co
    'Right Step Limited (RSL)',          # rightsteps.co.uk = workplace mental health
]

if __name__ == '__main__':
    import openpyxl, re
    from pathlib import Path

    # 1. Persist the drop list to a JSON file build_data.py reads at build time
    import json
    Path('data').mkdir(exist_ok=True)
    out = Path('data/MANUAL_DROP_LIST.json')
    out.write_text(json.dumps(DROP_LIST, indent=2), encoding='utf-8')
    print(f"Wrote {len(DROP_LIST)} names → {out}")

    # 2. Also remove from manual_contracts.xlsx rows where we added them via
    #    the GMCA Housing First script (if any of the drops are also in there)
    wb = openpyxl.load_workbook('data/manual_contracts.xlsx')
    companies = wb['Companies']
    ccs = wb['Company × Council × Sector']

    def normalise(s):
        s = (s or '').lower().strip()
        s = re.sub(r'\s+', ' ', s)
        return s

    drop_lc = {normalise(n) for n in DROP_LIST}

    # Companies sheet — delete matching rows
    removed_companies = 0
    rows_to_delete = []
    for r in range(2, companies.max_row + 1):
        v = normalise(companies.cell(row=r, column=1).value or '')
        if v in drop_lc:
            rows_to_delete.append(r)
    for r in reversed(rows_to_delete):
        companies.delete_rows(r)
        removed_companies += 1

    # CCS sheet — delete matching rows
    removed_ccs = 0
    rows_to_delete = []
    for r in range(2, ccs.max_row + 1):
        v = normalise(ccs.cell(row=r, column=1).value or '')
        if v in drop_lc:
            rows_to_delete.append(r)
    for r in reversed(rows_to_delete):
        ccs.delete_rows(r)
        removed_ccs += 1

    wb.save('data/manual_contracts.xlsx')
    print(f"Removed {removed_companies} rows from Companies sheet")
    print(f"Removed {removed_ccs} rows from CCS sheet")
