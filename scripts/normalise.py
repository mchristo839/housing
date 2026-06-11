"""
Read all raw portal JSONs from data/scraped/raw/ and write a single
normalised xlsx at data/scraped/normalised.xlsx with the schema from
scrapers.base.CONTRACT_FIELDS.

Each row keeps its source_portal / source_url / source_id so the
provenance is auditable from the spreadsheet.

Usage:
    python scripts/normalise.py
"""
import glob
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

import openpyxl
from scrapers.base import CONTRACT_FIELDS

RAW_DIR = os.path.join(ROOT, "data", "scraped", "raw")
OUT = os.path.join(ROOT, "data", "scraped", "normalised.xlsx")


def main():
    files = sorted(glob.glob(os.path.join(RAW_DIR, "*.json")))
    if not files:
        print(f"no raw scrape files in {RAW_DIR} — run scripts/scrape_all.py first")
        return

    all_rows = []
    per_portal = {}
    for path in files:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict) or "rows" not in data:
            continue
        portal = data.get("portal", os.path.basename(path).split("_")[0])
        rows = data.get("rows", [])
        # de-duplicate within the same portal by source_id (in case of overlapping runs)
        seen = set()
        kept = []
        for r in rows:
            sid = r.get("source_id") or (r.get("Company"), r.get("Council"), r.get("Contract Titles"))
            if sid in seen:
                continue
            seen.add(sid)
            kept.append(r)
        per_portal[portal] = len(kept)
        all_rows.extend(kept)
        print(f"  {portal}: {len(kept)} rows (from {os.path.basename(path)})")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Company × Council × Sector")
    ws.append(CONTRACT_FIELDS)
    for r in all_rows:
        ws.append([r.get(k, "") for k in CONTRACT_FIELDS])

    # Also a small Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Portal", "Rows"])
    for p, n in per_portal.items():
        ws2.append([p, n])
    ws2.append(["TOTAL", sum(per_portal.values())])

    wb.save(OUT)
    print(f"\nwrote {len(all_rows)} rows → {OUT}")
    print("Next: python scripts/dedup_report.py")


if __name__ == "__main__":
    main()
