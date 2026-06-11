"""Add evidence-backed Greater Manchester ABEN contracts.

A Bed Every Night (ABEN) — established by GM Mayor Andy Burnham 2018, now Phase 5.
Each of GM's 10 contracting authorities runs its local ABEN service.

Verified contracts to add:
  Rochdale Borough Council × WHAG (Women's Housing Action Group)
     - whag.info confirms 6 dispersed ABEN properties via Rochdale BC contract
  Rochdale Borough Council × Stepping Stone Projects
     - stepping-stone.org.uk confirms ABEN Supported Housing Worker coordinating
       Redfearn House, Leopold Court, Sanctuary Trust beds via Rochdale BC
  Rochdale Borough Council × Rochdale Boroughwide Housing (RBH)
     - confirmed ABEN project management role per stepping-stone.org.uk
  Manchester City Council × The Riverside Group (already in DB, add MCC)
     - riverside.org.uk confirms ABEN Withington LGBTQ+ scheme
"""
import openpyxl

PATH = 'data/manual_contracts.xlsx'
wb = openpyxl.load_workbook(PATH)
companies = wb['Companies']
ccs = wb['Company × Council × Sector']

# Companies to add (the 3 new) — Riverside already exists
NEW_COMPANIES = [
    # (Name, Website, Phone, Email, Address, CH#, SME, VCSE, Charity#)
    ('WHAG', 'https://whag.info/', '01706 298 222', 'whag@whag.info',
     '185 Drake Street, Rochdale, OL11 1EF', '03294852', 'Yes', 'Yes', '1029717'),
    ('Stepping Stone Projects', 'https://www.stepping-stone.org.uk/',
     '01706 359 600', 'crt@stepping-stone.org.uk',
     'Stepping Stone Projects, Rochdale, OL16 1FR', None, 'Yes', 'Yes', '1009476'),
    ('Rochdale Boroughwide Housing', 'https://www.rbh.org.uk/',
     '0800 027 7769', 'customer.services@rbh.org.uk',
     'Unique Enterprise Centre, Belfield Road, Rochdale, OL16 2UP',
     None, 'No', 'No', None),
]

# Find existing canonical Riverside name in Companies sheet
RIVERSIDE_CANONICAL = 'The Riverside Group'

# Add Companies rows
existing = set()
for r in range(2, companies.max_row + 1):
    v = (companies.cell(row=r, column=1).value or '').strip().lower()
    if v: existing.add(v)

n_companies_added = 0
for (name, website, phone, email, addr, ch, sme, vcse, charity) in NEW_COMPANIES:
    if name.lower() in existing: continue
    nr = companies.max_row + 1
    companies.cell(row=nr, column=1).value  = name
    companies.cell(row=nr, column=2).value  = 0
    companies.cell(row=nr, column=5).value  = 1
    companies.cell(row=nr, column=7).value  = 'No'
    companies.cell(row=nr, column=8).value  = ch
    companies.cell(row=nr, column=9).value  = sme
    companies.cell(row=nr, column=10).value = vcse
    companies.cell(row=nr, column=12).value = addr
    companies.cell(row=nr, column=13).value = website
    companies.cell(row=nr, column=14).value = phone
    companies.cell(row=nr, column=15).value = email
    if charity:
        companies.cell(row=nr, column=17).value = charity
        companies.cell(row=nr, column=19).value = 'Registered Charity'
    n_companies_added += 1
    print(f"  + Companies: {name}")

# CCS rows — Rochdale ABEN evidence
ABEN_TITLE_ROCHDALE = (
    "A Bed Every Night (ABEN) — Rochdale Borough Council local service. "
    "ABEN established by GM Mayor 2018; each of GM's 10 contracting "
    "authorities commissions a local provider. "
    "Source: whag.info / stepping-stone.org.uk / GMCA gmhousingfirst.org.uk."
)
ABEN_TITLE_MANCHESTER = (
    "A Bed Every Night (ABEN) — Manchester City Council local service. "
    "Riverside ABEN Withington LGBTQ+ scheme. "
    "Source: riverside.org.uk/in-your-neighbourhood/manchester-2/care-and-support/a-bed-every-night-withington/."
)

CCS_ADD = [
    ('WHAG',                          'Rochdale Borough Council', ABEN_TITLE_ROCHDALE),
    ('Stepping Stone Projects',       'Rochdale Borough Council', ABEN_TITLE_ROCHDALE),
    ('Rochdale Boroughwide Housing',  'Rochdale Borough Council', ABEN_TITLE_ROCHDALE),
    (RIVERSIDE_CANONICAL,             'Manchester City Council',  ABEN_TITLE_MANCHESTER),
]
n_ccs_added = 0
for (co, council, title) in CCS_ADD:
    nr = ccs.max_row + 1
    ccs.cell(row=nr, column=1).value  = co
    ccs.cell(row=nr, column=2).value  = 'Housing'
    ccs.cell(row=nr, column=3).value  = council
    ccs.cell(row=nr, column=4).value  = 1
    ccs.cell(row=nr, column=11).value = 'Emergency accommodation | Rough sleepers | Homelessness | Supported accommodation'
    ccs.cell(row=nr, column=13).value = title
    ccs.cell(row=nr, column=14).value = 'North West'
    ccs.cell(row=nr, column=15).value = 'Local Authority'
    ccs.cell(row=nr, column=16).value = 'Local'
    n_ccs_added += 1
    print(f"  + CCS: {co:<38s} × {council}")

wb.save(PATH)
print(f"\n✓ Saved {PATH}")
print(f"  Companies: +{n_companies_added}, CCS: +{n_ccs_added}")
