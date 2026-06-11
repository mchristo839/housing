"""Build contract-first HTML — one card per contract, all suppliers listed under it.

Inverse of NEW_PROVIDERS.html. Same data, regrouped.
"""
import openpyxl, json, re, sys, io, html as htmllib
from collections import defaultdict, Counter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUT = 'data/scraped/NEW_CONTRACTS.html'

def norm(s): return re.sub(r'[^a-z0-9]', '', str(s or '').lower())

# Housing / accommodation category rules (a contract can hit multiple labels)
CAT_RULES = [
    ('Supported living',           r'supported living'),
    ('Supported accommodation',    r'supported accommodation'),
    ('Temporary accommodation',    r'temporary accommodation|nightly purchased'),
    ('Emergency accommodation',    r'emergency accommodation|emergency housing'),
    ('Extra care housing',         r'extra care'),
    ('Mental health accommodation', r'mental health.{0,40}(accommodation|housing|supported|step|crisis|pathway)'),
    ('Learning disability placements', r'learning disab.{0,40}(accommodation|housing|supported|residential|placement)'),
    ('Floating support / HRS',     r'floating support|housing related support|housing[- ]related'),
    ('Step-down accommodation',    r'step[- ]?down|crisis prevention'),
    ('Asylum / refugee',           r'asylum|refugee|aasc'),
    ('Hostels / refuge / DA',      r'\bhostel|refuge\b|domestic abuse'),
    ('Rough sleeping / homelessness', r'rough sleeper|housing first|single homeless|homeless'),
    ('Young people / care leavers', r'young people|youth (housing|accommodation)|care leaver|looked after children|16\s*\+'),
    ('Semi-independent',           r'semi[- ]?independent'),
    ('Pan-London framework',       r'pan[- ]?london|capital letters'),
    ('Sheltered housing',          r'sheltered'),
]

def classify(title, cat_field):
    blob = (str(title or '') + ' ' + str(cat_field or '')).lower()
    return [label for label, pat in CAT_RULES if re.search(pat, blob)]

# Baseline = main DB (so we can mark which suppliers are NEW vs already existed)
wb_main = openpyxl.load_workbook('data/care_housing_database_v2_ENRICHED.xlsx',
                                 read_only=True, data_only=True)
main_suppliers = set()
for r in wb_main['Company × Council × Sector'].iter_rows(min_row=2, values_only=True):
    if r[0]: main_suppliers.add(norm(r[0]))
for r in wb_main['Companies'].iter_rows(min_row=2, values_only=True):
    if r[0]: main_suppliers.add(norm(r[0]))

# Live providers (what survived the build)
prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
live = {norm(p['name']): p for p in prov}

# Walk manual file, group by contract
wb_m = openpyxl.load_workbook('data/manual_contracts.xlsx', read_only=True)
ws_cs = wb_m['Company × Council × Sector']

# Contract identity = source_id (Bidstats notice) if present in title,
# else (council, title-without-supplier-annotation).
def contract_key(council, title):
    """Strip per-supplier annotations from the title to group correctly."""
    t = re.sub(r'\s*\([^)]*?bidstats[^)]*\)\s*$', '', title or '', flags=re.I)
    t = re.sub(r'\s*\([^)]*?supplier\s*@[^)]*\)\s*$', '', t, flags=re.I)
    t = t.strip()
    bm = re.search(r'Bidstats\s+(\d+)', title or '', flags=re.I)
    if bm:
        return ('bidstats', bm.group(1), council)
    return ('title', t[:120].lower(), council.lower() if council else '')

contracts = defaultdict(lambda: {
    'council': '', 'title_clean': '', 'category': '', 'award_date': '',
    'bidstats_id': '', 'suppliers': [], 'cats': set(),
})

for r in ws_cs.iter_rows(min_row=2, values_only=True):
    if not r[0] or not r[2]: continue
    supplier, council, title = str(r[0]), str(r[2]), str(r[12] or '')
    cat = str(r[10] or '')
    award = str(r[11] or '')
    key = contract_key(council, title)
    c = contracts[key]
    # Clean title for the card heading
    t_clean = re.sub(r'\s*\([^)]*?bidstats[^)]*\)\s*$', '', title, flags=re.I)
    t_clean = re.sub(r'\s*\([^)]*?supplier\s*@[^)]*\)\s*$', '', t_clean, flags=re.I)
    t_clean = t_clean.strip()
    if not c['council']: c['council'] = council
    if not c['title_clean'] or len(t_clean) > len(c['title_clean']):
        c['title_clean'] = t_clean
    if not c['category']: c['category'] = cat
    if not c['award_date']: c['award_date'] = award
    if key[0] == 'bidstats' and not c['bidstats_id']:
        c['bidstats_id'] = key[1]
    for L in classify(title, cat): c['cats'].add(L)
    # Extract supplier-office location from the per-supplier annotation
    loc = ''
    lm = re.search(r'supplier\s*@\s*([^)]+)\)', title, re.I)
    if lm: loc = lm.group(1).strip()
    if loc.lower() in ('none', ''): loc = ''
    is_new = norm(supplier) not in main_suppliers
    is_live = norm(supplier) in live
    c['suppliers'].append({
        'name': supplier.strip(), 'location': loc,
        'new': is_new, 'live': is_live,
    })

print(f"Distinct contracts: {len(contracts)}")
print(f"Total supplier awards: {sum(len(c['suppliers']) for c in contracts.values())}")

# London-borough sort order
LONDON = ['Barking and Dagenham','Barnet','Bexley','Brent','Bromley','Camden',
         'City of London','Croydon','Ealing','Enfield','Greenwich','Hackney',
         'Hammersmith and Fulham','Haringey','Harrow','Havering','Hillingdon',
         'Hounslow','Islington','Kensington and Chelsea','Kingston upon Thames',
         'Lambeth','Lewisham','Merton','Newham','Redbridge','Richmond upon Thames',
         'Southwark','Sutton','Tower Hamlets','Waltham Forest','Wandsworth','Westminster']
def borough_of(c):
    if not c: return ''
    cl = c.lower()
    for b in LONDON:
        if b.lower() in cl: return b
    if 'kingston' in cl: return 'Kingston upon Thames'
    if 'richmond' in cl: return 'Richmond upon Thames'
    if 'kensington' in cl or 'rbkc' in cl: return 'Kensington and Chelsea'
    if 'wcc' in cl: return 'Westminster'
    return c

def sort_tuple(item):
    key, c = item
    b = borough_of(c['council'])
    pri = LONDON.index(b) if b in LONDON else 99
    return (pri, b, -len(c['suppliers']), c['title_clean'].lower())

contracts_sorted = sorted(contracts.items(), key=sort_tuple)

# Build HTML
parts = []
parts.append('<!doctype html><html><head><meta charset="utf-8">')
parts.append('<title>Contracts and the suppliers awarded to them</title>')
parts.append('<style>')
parts.append('body{font:14px -apple-system,Segoe UI,Roboto,sans-serif;margin:1.5em;background:#f8f9fa;color:#222;max-width:1200px}')
parts.append('h1{margin:0 0 .2em 0;font-size:22px}')
parts.append('.meta{color:#666;margin-bottom:1.5em;font-size:13px}')
parts.append('.search{position:sticky;top:0;background:#fff;padding:10px 14px;border:1px solid #ddd;border-radius:6px;margin-bottom:1em;box-shadow:0 2px 4px rgba(0,0,0,.05);z-index:5}')
parts.append('.search input{width:340px;padding:6px 10px;font-size:14px;border:1px solid #bbb;border-radius:4px}')
parts.append('.search select{padding:6px 10px;margin-left:6px;font-size:13px}')
parts.append('.counts{color:#444;margin-left:1em;font-size:13px;font-weight:600}')
parts.append('.contract{background:#fff;border:1px solid #ddd;border-radius:8px;padding:14px 18px;margin-bottom:14px;box-shadow:0 1px 2px rgba(0,0,0,.04)}')
parts.append('.contract h2{margin:0 0 4px 0;font-size:15px;color:#1F4E79;font-weight:600}')
parts.append('.contract .commiss{font-size:13px;font-weight:600;color:#444;margin-bottom:4px}')
parts.append('.contract .desc{font-size:13px;color:#555;margin-bottom:8px;line-height:1.4}')
parts.append('.contract .meta-bar{font-size:11px;color:#888;margin-bottom:10px}')
parts.append('.contract .meta-bar a{color:#0563C1;text-decoration:none}')
parts.append('.contract .meta-bar a:hover{text-decoration:underline}')
parts.append('.contract .badges{margin-bottom:8px}')
parts.append('.badge{display:inline-block;padding:2px 9px;border-radius:11px;font-size:11px;font-weight:600;margin-right:5px}')
parts.append('.badge.tot{background:#e7eff8;color:#1F4E79}')
parts.append('.badge.new{background:#d4edda;color:#1e6033}')
parts.append('.badge.live{background:#fff3cd;color:#856404}')
parts.append('.badge.cat{background:#eef2f7;color:#3c4858;font-weight:500}')
parts.append('.suppliers{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:6px;border-top:1px dashed #ddd;padding-top:10px;margin-top:8px}')
parts.append('.sup{padding:6px 10px;background:#fafbfc;border-left:3px solid #ccc;border-radius:0 4px 4px 0;font-size:13px}')
parts.append('.sup.new{border-left-color:#28a745;background:#f0f8f2}')
parts.append('.sup.new-not-live{border-left-color:#fd7e14;background:#fff8f0}')
parts.append('.sup .nm{font-weight:600}')
parts.append('.sup .loc{font-size:11px;color:#777}')
parts.append('.sup .tag{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.3px;padding:1px 5px;border-radius:3px;margin-left:4px}')
parts.append('.sup .tag.new{background:#28a745;color:#fff}')
parts.append('.sup .tag.dropped{background:#fd7e14;color:#fff}')
parts.append('.sup .tag.existing{background:#dee2e6;color:#555}')
parts.append('.borough-divider{font-size:18px;font-weight:700;color:#222;margin:24px 0 10px 0;padding-bottom:4px;border-bottom:2px solid #1F4E79}')
parts.append('</style></head><body>')

total_contracts = len(contracts_sorted)
total_supplier_rows = sum(len(c['suppliers']) for _, c in contracts_sorted)
total_new = sum(1 for _, c in contracts_sorted for s in c['suppliers'] if s['new'])
total_new_live = sum(1 for _, c in contracts_sorted for s in c['suppliers'] if s['new'] and s['live'])
unique_new_suppliers = set()
for _, c in contracts_sorted:
    for s in c['suppliers']:
        if s['new'] and s['live']: unique_new_suppliers.add(norm(s['name']))

parts.append('<h1>Contracts and the suppliers awarded to them</h1>')
parts.append(
    f'<div class="meta">'
    f'<b>{total_contracts}</b> distinct contracts · '
    f'<b>{total_supplier_rows}</b> supplier awards · '
    f'<b>{len(unique_new_suppliers)}</b> unique net-new suppliers live on the site. '
    f'<br><br>'
    f'Grouped by London borough A–Z. Each card shows one contract with all the suppliers awarded to it.'
    f' <span class="badge new">NEW</span> = supplier net-new to the universe this session.'
    f' <span class="badge live">existed</span> = supplier was already in the main DB (we attached this London contract to them).'
    f'</div>'
)

parts.append('<div class="search">')
parts.append('Filter: <input id="q" placeholder="borough, contract title, supplier, source id..." oninput="flt()">')
parts.append('<select id="onlyNew" onchange="flt()">'
             '<option value="">All contracts</option>'
             '<option value="new">Contracts with ≥1 net-new supplier</option>'
             '<option value="allnew">Only fully-new contracts</option>'
             '</select>')
# Category dropdown
cat_counts_local = Counter()
for k2, c2 in contracts_sorted:
    for L in c2['cats']: cat_counts_local[L] += 1
parts.append('<select id="catFilter" onchange="flt()">')
parts.append('<option value="">All categories</option>')
for L, n in cat_counts_local.most_common():
    parts.append(f'<option value="{htmllib.escape(L)}">{htmllib.escape(L)} ({n})</option>')
parts.append('</select>')
parts.append('<span class="counts" id="c"></span>')
parts.append('</div>')

current_borough = None
for key, c in contracts_sorted:
    b = borough_of(c['council'])
    if b != current_borough:
        current_borough = b
        bcount = sum(1 for k2,c2 in contracts_sorted if borough_of(c2['council']) == b)
        parts.append(f'<div class="borough-divider">{htmllib.escape(b or "Other")} <span style="font-size:12px;color:#888;font-weight:400">({bcount} contracts)</span></div>')

    sup_n = len(c['suppliers'])
    new_n = sum(1 for s in c['suppliers'] if s['new'])
    live_n = sum(1 for s in c['suppliers'] if s['new'] and s['live'])
    is_all_new = new_n == sup_n and sup_n > 0
    has_new = new_n > 0

    council = htmllib.escape(c['council'])
    title = htmllib.escape(c['title_clean'] or '(untitled)')
    cat = htmllib.escape(c['category'])
    award = htmllib.escape(c['award_date'])
    bid = c['bidstats_id']
    src_link = f'<a href="https://bidstats.uk/notice/{bid}" target="_blank">Bidstats {bid}</a>' if bid else ''

    badges = (
        f'<span class="badge tot">{sup_n} supplier{"s" if sup_n!=1 else ""}</span>'
        f'{" <span class=\"badge new\">"+str(new_n)+" net-new</span>" if new_n else ""}'
        f'{" <span class=\"badge live\">"+str(live_n)+" live</span>" if live_n != new_n and live_n else ""}'
    )
    cat_badges = ''.join(
        f'<span class="badge cat">{htmllib.escape(L)}</span>'
        for L in sorted(c['cats'])
    )

    search_blob = (council + ' ' + title + ' ' + cat + ' ' + bid + ' ' +
                   ' '.join(s['name'] for s in c['suppliers']) + ' ' +
                   ' '.join(c['cats'])).lower()
    data_attrs = (
        f' data-search="{htmllib.escape(search_blob)}"'
        f' data-allnew="{1 if is_all_new else 0}"'
        f' data-hasnew="{1 if has_new else 0}"'
        f' data-cats="{htmllib.escape("|".join(c["cats"]))}"'
    )

    parts.append(f'<div class="contract"{data_attrs}>')
    parts.append(f'<h2>{title}</h2>')
    parts.append(f'<div class="commiss">{council}</div>')
    if cat: parts.append(f'<div class="desc">{cat}</div>')
    if src_link or award:
        bits = []
        if src_link: bits.append(src_link)
        if award: bits.append(f'awarded {award}')
        parts.append(f'<div class="meta-bar">{" · ".join(bits)}</div>')
    parts.append(f'<div class="badges">{badges}</div>')
    if cat_badges: parts.append(f'<div class="badges" style="margin-top:4px">{cat_badges}</div>')

    parts.append('<div class="suppliers">')
    for s in sorted(c['suppliers'], key=lambda x: x['name'].lower()):
        cls = 'sup'
        tag = ''
        if s['new'] and s['live']:
            cls += ' new'
            tag = '<span class="tag new">NEW</span>'
        elif s['new'] and not s['live']:
            cls += ' new-not-live'
            tag = '<span class="tag dropped">filtered</span>'
        else:
            tag = '<span class="tag existing">existed</span>'
        nm = htmllib.escape(s['name'])
        loc = f'<div class="loc">{htmllib.escape(s["location"])}</div>' if s['location'] else ''
        parts.append(f'<div class="{cls}"><span class="nm">{nm}</span>{tag}{loc}</div>')
    parts.append('</div>')
    parts.append('</div>')

parts.append('<script>')
parts.append('function flt(){const q=document.getElementById("q").value.toLowerCase(); const mode=document.getElementById("onlyNew").value; const cat=document.getElementById("catFilter").value; const items=document.querySelectorAll(".contract"); let shown=0; items.forEach(it=>{let ok=true; if(q && !it.dataset.search.includes(q))ok=false; if(mode==="new" && it.dataset.hasnew!=="1")ok=false; if(mode==="allnew" && it.dataset.allnew!=="1")ok=false; if(cat && !it.dataset.cats.split("|").includes(cat))ok=false; it.style.display=ok?"":"none"; if(ok)shown++;}); document.getElementById("c").textContent=shown+" of "+items.length+" contracts shown";}')
parts.append('flt();')
parts.append('</script></body></html>')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(''.join(parts))

print(f"\nWrote {OUT}")
print(f"  Distinct contracts: {total_contracts}")
print(f"  Supplier awards   : {total_supplier_rows}")
print(f"  Unique net-new live: {len(unique_new_suppliers)}")
print(f"  File size: {len(open(OUT,'rb').read())//1024}KB")
