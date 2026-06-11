"""Final push — for each CSV, ensure every supplier ends up either Verified
or fairly assessed via Firecrawl. Restores wrongly-dropped suppliers along the way.

Stages:
  1. PARSE — pull supplier names from each CSV
  2. RESTORE — remove anyone in the CSVs from MANUAL_DROP_LIST
  3. PROMOTE — make sure each supplier has a Companies row with Housing Contracts>=1
  4. REBUILD — providers.json refresh
  5. VERIFY — Firecrawl /search + /extract on every provider not already in VERIFIED.json
  6. APPLY + REBUILD + DEPLOY

USAGE:
  python scripts/final_push.py csv1 csv2 csv3 ...
"""
import csv, json, re, sys, subprocess
from pathlib import Path
import openpyxl

def norm(name):
    """Strip Ltd/Limited/PLC/LLP/CIC + punctuation for proper duplicate matching.
    Must match the norm() used in ingest_contract.py and the build pipeline."""
    n = (name or '').lower().strip()
    n = re.sub(r'\b(ltd|limited|llp|plc|company|co|cic|t\/a|trading as)\b', '', n)
    n = re.sub(r'[^a-z0-9 ]', ' ', n)
    return ' '.join(n.split())

def parse_suppliers_field(s):
    """Contracts Finder pipe format: [Name|Address|...][...]"""
    if not s: return []
    out = []
    for b in re.split(r'\]\s*\[', s.strip().strip('[]')):
        parts = [p.strip() for p in b.split('|')]
        if parts and parts[0]:
            out.append(re.sub(r'\s+', ' ', parts[0]).strip())
    return out

def names_from_csv(path):
    """Try common name columns: Supplier field (Contracts Finder), Provider, Supplier Name."""
    names = set()
    with open(path, encoding='utf-8-sig', errors='replace') as f:
        rows = list(csv.DictReader(f))
    if not rows: return names
    cols = list(rows[0].keys())
    sup_col = next((c for c in cols if c.startswith('Supplier [Name')), None)
    other_col = next((c for c in cols if c.lower().strip() in
                      {'supplier','supplier name','provider','provider name','company','company name'}), None)
    for r in rows:
        if sup_col:
            for n in parse_suppliers_field(r.get(sup_col,'')):
                if n.lower() not in {'none','na','tbc'}: names.add(n)
        if other_col:
            n = (r.get(other_col,'') or '').strip()
            if n and n.lower() not in {'none','na','tbc'}: names.add(n)
    return names

def main():
    csvs = sys.argv[1:]
    if not csvs:
        print(__doc__); sys.exit(1)

    print("=== STAGE 1: PARSE all CSVs ===", flush=True)
    all_names = set()
    for csv_path in csvs:
        if not Path(csv_path).exists():
            print(f"  SKIP {csv_path}: file not found"); continue
        names = names_from_csv(csv_path)
        print(f"  {Path(csv_path).name}: {len(names)} suppliers")
        all_names |= names
    print(f"\nUnique suppliers across all CSVs: {len(all_names)}", flush=True)

    # Snapshot state — use norm() so 'Foo Care Ltd' matches 'FOO CARE LIMITED'
    prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
    prov_names_norm = {norm(p['name']) for p in prov}
    drops = json.load(open('data/MANUAL_DROP_LIST.json', encoding='utf-8'))
    drops_norm = {norm(n) for n in drops}
    verified = json.load(open('data/verification/VERIFIED.json', encoding='utf-8'))
    verified_norm = {norm(n) for n in verified}

    # === STAGE 2: RESTORE from drops ===
    print("\n=== STAGE 2: RESTORE wrongly-dropped ===", flush=True)
    to_restore = [n for n in all_names if norm(n) in drops_norm]
    print(f"  On drop list: {len(to_restore)} — removing")
    if to_restore:
        restore_norms = {norm(n) for n in to_restore}
        keep = [d for d in drops if norm(d) not in restore_norms]
        json.dump(keep, open('data/MANUAL_DROP_LIST.json','w', encoding='utf-8'), indent=2)
        print(f"  Drop list: {len(drops)} -> {len(keep)}", flush=True)
    else:
        keep = drops

    # === STAGE 3: PROMOTE (ensure Companies row with Housing Contracts>=1) ===
    print("\n=== STAGE 3: PROMOTE Companies rows ===", flush=True)
    wb = openpyxl.load_workbook('data/manual_contracts.xlsx')
    comp = wb['Companies']
    idx_by_norm = {}
    for r in range(2, comp.max_row+1):
        nm = (comp.cell(row=r, column=1).value or '').strip()
        if nm: idx_by_norm[norm(nm)] = r

    upd = added = 0
    for nm in all_names:
        if norm(nm) in verified_norm: continue   # already Verified, leave alone
        nl = norm(nm)
        row_idx = idx_by_norm.get(nl)
        if row_idx:
            h = comp.cell(row=row_idx, column=5).value or 0
            if not h or h == 0:
                comp.cell(row=row_idx, column=5).value = 1
                upd += 1
        else:
            nr = comp.max_row + 1
            comp.cell(row=nr, column=1).value = nm
            comp.cell(row=nr, column=2).value = 0
            comp.cell(row=nr, column=5).value = 1
            comp.cell(row=nr, column=7).value = 'No'
            idx_by_norm[nl] = nr
            added += 1
    wb.save('data/manual_contracts.xlsx')
    print(f"  Companies sheet: {upd} updated, {added} added", flush=True)

    # === STAGE 4: REBUILD ===
    print("\n=== STAGE 4: REBUILD providers.json ===", flush=True)
    subprocess.run(['python', 'build_data.py'], check=True,
                   stdout=subprocess.DEVNULL)
    prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
    print(f"  Total providers in DB: {len(prov)}", flush=True)

    # === STAGE 5: VERIFY everyone not yet Verified ===
    print("\n=== STAGE 5: VERIFY unverified (Firecrawl) ===", flush=True)
    subprocess.run(['python', '-u', 'scripts/firecrawl_search_unverified.py'], check=False)

    # === STAGE 6: APPLY + REBUILD + DEPLOY ===
    print("\n=== STAGE 6: APPLY + REBUILD + DEPLOY ===", flush=True)
    subprocess.run(['python', 'scripts/apply_firecrawl.py'], check=False)
    subprocess.run(['python', 'build_data.py'], check=True,
                   stdout=subprocess.DEVNULL)
    subprocess.run(['npm', 'run', 'build'], check=False,
                   stdout=subprocess.DEVNULL, shell=True)
    subprocess.run(['vercel', 'deploy', '--prod', '--yes'], check=False, shell=True)

    # Final report
    prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
    v = sum(1 for x in prov if x.get('verification',{}).get('verified'))
    print(f"\n=== DONE ===")
    print(f"  Total providers:  {len(prov)}")
    print(f"  ✓ Verified:       {v}  ({v*100//len(prov)}%)")
    print(f"  📋 Listed:         {len(prov)-v}  ({(len(prov)-v)*100//len(prov)}%)")

if __name__ == '__main__':
    main()
