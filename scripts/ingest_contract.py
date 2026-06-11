"""Ingest a new contract-award notice (CSV from Contracts Finder).

Pipeline:
  1. PARSE     — read CSV, extract notice + supplier list
  2. CHECK     — for each notice + (supplier x council), check what's already in DB
  3. CLASSIFY  — bucket into: already-known / new-link-to-existing-supplier / new-supplier
  4. WRITE     — append to manual_contracts.xlsx (Companies sheet + CCS sheet)
  5. VERIFY    — run Firecrawl on new suppliers only (find website + confirm services)
  6. APPLY     — auto-promote verified ones to VERIFIED.json
  7. BUILD     — rebuild providers.json
  8. DEPLOY    — push to Vercel (with --deploy flag)

USAGE:
  python scripts/ingest_contract.py path/to/notice.csv
  python scripts/ingest_contract.py path/to/notice.csv --deploy
  python scripts/ingest_contract.py path/to/notice.csv --no-verify  # skip Firecrawl
"""
import csv, re, json, sys, time, subprocess, os
from pathlib import Path
import openpyxl

MANUAL_XL = Path('data/manual_contracts.xlsx')
PROVIDERS = Path('api/_data/providers.json')
ANALYSED  = Path('data/verification/firecrawl_analysed.json')

# ── 1. PARSE the input CSV ────────────────────────────────────────────────────
# Supplier-name patterns we ALWAYS skip — these are individual facility brand
# names rather than operating companies. We want the parent operator (which
# might own dozens of care homes), not "Wollaton View Care Home" as a separate
# entry.
SKIP_SUPPLIER_PATTERNS = [
    re.compile(r'\bcare\s*home\b', re.I),       # individual care home brands
    re.compile(r'\bnursing\s*home\b', re.I),    # individual nursing home brands
]

def parse_suppliers(s):
    """Contracts Finder format: [Name|Address|Ref type|Ref Number|Is SME|Is VCSE][...][...]
    Skips supplier names matching SKIP_SUPPLIER_PATTERNS (individual facilities).
    """
    if not s: return []
    out = []
    for b in re.split(r'\]\s*\[', s.strip().strip('[]')):
        parts = [p.strip() for p in b.split('|')]
        if not parts or not parts[0]: continue
        name = re.sub(r'\s+', ' ', parts[0]).strip()
        # Skip individual care/nursing home brand names — keep operating companies
        if any(p.search(name) for p in SKIP_SUPPLIER_PATTERNS):
            continue
        out.append({
            'name': name,
            'addr': re.sub(r'\s+', ' ', parts[1] if len(parts) > 1 else '').strip(', '),
            'ch':   parts[3].strip() if len(parts) > 3 else '',
            'sme':  (parts[4].strip().lower() == 'yes') if len(parts) > 4 else None,
            'vcse': (parts[5].strip().lower() == 'yes') if len(parts) > 5 else None,
        })
    return out

REGION_MAP = {
    'east of england': 'East of England', 'east midlands': 'East Midlands',
    'london': 'London', 'north east': 'North East', 'north west': 'North West',
    'south east': 'South East', 'south west': 'South West',
    'west midlands': 'West Midlands', 'yorkshire and the humber': 'Yorkshire & The Humber',
}
PORTAL_ORG_MAP = {
    'IN-TEND LIMITED': 'Hertfordshire County Council',
    'ONESOURCE PARTNERSHIP LTD': 'London Borough of Havering',
    'CAPITAL ESOURCING C.I.C.': 'Royal Borough of Kensington and Chelsea',
}

def is_housing_notice(row):
    """Decide if a Contracts Finder notice is housing-related.

    Catches: supported living, residential care, care homes, nursing homes,
    housing-related support, floating support, refuges, domestic abuse, day
    services with care suppliers, assessment beds, DPS for care, prevention,
    ageing well, and looked-after residential placements — all of which
    operationally involve providers our paying customers care about.

    Excludes: telecare/alarms (assistive tech only), IT services, stationery,
    vehicle hire, training-only, equipment-only, transport-only.
    """
    title = (row.get('Title') or '').lower()
    desc  = ((row.get('Description') or '') or '')[:400].lower()
    text  = title + ' ' + desc

    HOUSING = (
        # core supported-living family
        'supported living', 'supported accommodation', 'supported housing',
        'extra care', 'shared lives', 'independent living',
        # residential / nursing
        'residential care', 'residential home', 'residential service',
        'residential bed', 'residential placement',
        'care home', 'nursing home', 'care and nursing',
        'specialist residential',
        # client groups (almost always supported-living / residential)
        'learning disabilit', 'autism', 'mental health', 'complex need',
        'complex behaviour', 'complex care',
        # community-based housing-with-care
        'community supported', 'community based', 'community support',
        'enablement', 'reenablement', 'placements',
        # 16+ / care-leavers / children in care
        'leaving care', 'care leaver', 'looked after', '16+', '18+',
        "children's home", 'childrens home', 'children home',
        # homelessness + asylum + DA
        'homelessness', 'homeless', 'rough sleeping', 'rough sleeper',
        'temporary accommodation', 'emergency accommodation',
        'asylum', 'refuge', 'refuges', 'domestic abuse', 'domestic violence',
        'modern slavery',
        # housing language used by councils
        'housing related', 'housing-related', 'housing support',
        'housing & support', 'housing and support',
        'floating support', 'tenancy support', 'tenancy sustainment',
        'housing, health', 'health and housing',
        # other commonly-used framework labels
        'accommodation', 'accommodation-based', 'sheltered housing',
        'sheltered accommodation', 'short-term assessment', 'assessment bed',
        'transitional', 'step-down', 'step down', 'pathway',
        'ageing well', 'ageing well framework',
        # DPS / framework wrappers we know hold care
        'dps for care', 'care dynamic purchasing',
        'care and support', 'health and social care',
        # Day services tend to be commissioned alongside supported living
        'day service', 'day opportunit', 'day care',
        # Substance misuse often bundles housing-related support
        'substance misuse',
    )
    NON_HOUSING = (
        # clearly not housing
        'transport only', 'training only', 'equipment only',
        'live in care only', 'home care only',
        # tech / consumables / vehicles / pure IT
        'telecare', 'telehealth', 'lifeline', 'community alarm',
        'pendant alarm', 'stationery', 'office supplies', 'vehicle hire',
        'fleet management', 'fuel card', 'photocopier', 'printing',
        'it services', 'it support', 'broadband', 'telephony',
        'cyber security', 'software licence', 'cloud hosting',
        # works / construction
        'civil engineering', 'highways maintenance', 'street lighting',
        'building maintenance', 'roofing', 'lift maintenance',
        # ancillary, never housing
        'audit services', 'legal services', 'insurance', 'taxi service',
        'pension administration', 'banking services',
    )
    return any(h in text for h in HOUSING) and not any(n in text for n in NON_HOUSING)

def parse_csv(path):
    """Return list of contract dicts:
       { notice_id, title, council, region, awarded_date, value, suppliers: [...] }"""
    rows = []
    with open(path, encoding='utf-8-sig', errors='replace') as f:
        rows = list(csv.DictReader(f))
    print(f"  Read {len(rows)} rows from {path}", flush=True)

    contracts = []
    skipped = 0
    for r in rows:
        nid = (r.get('Notice Identifier') or '').strip()
        if not nid: skipped += 1; continue
        if not is_housing_notice(r): skipped += 1; continue

        org = r.get('Organisation Name', '').strip()
        council = org
        for portal, real in PORTAL_ORG_MAP.items():
            if portal in org.upper(): council = real; break
        if not council: skipped += 1; continue
        council = re.sub(r'\bCOUNCIL\b', 'Council', council, flags=re.I)

        region = REGION_MAP.get((r.get('Region') or '').lower().strip(), '')
        suppliers = parse_suppliers(r.get(
            'Supplier [Name|Address|Ref type|Ref Number|Is SME|Is VCSE]', '') or '')
        if not suppliers: skipped += 1; continue

        contracts.append({
            'notice_id':    nid,
            'title':        r.get('Title', '').strip(),
            'council':      council,
            'region':       region,
            'awarded_date': r.get('Awarded Date', ''),
            'value':        r.get('Awarded Value', ''),
            'description':  r.get('Description', ''),
            'suppliers':    suppliers,
        })
    print(f"  Kept {len(contracts)} housing notices, skipped {skipped} non-housing", flush=True)
    return contracts

# ── 2-3. CHECK existing + CLASSIFY into three buckets ─────────────────────────
def norm(name):
    """Aggressive normalisation for duplicate detection.
    Strips legal suffixes (Ltd/Limited/PLC/LLP/CIC), trading-as prefixes,
    and punctuation. So 'Foo Care Ltd' == 'FOO CARE LIMITED' == 'Foo Care'."""
    n = (name or '').lower().strip()
    n = re.sub(r'\b(ltd|limited|llp|plc|company|co|cic|t\/a|trading as)\b', '', n)
    n = re.sub(r'[^a-z0-9 ]', ' ', n)
    return ' '.join(n.split())

def classify_contracts(contracts):
    prov = json.load(open(PROVIDERS, encoding='utf-8'))
    known_suppliers = {norm(p['name']): p['name'] for p in prov}
    existing_pairs = set()
    for p in prov:
        nl = norm(p['name'])
        for c in (p.get('contracts_list') or []):
            existing_pairs.add((nl, norm(c.get('council', ''))))

    bucket_known    = []   # supplier known + council link known — skip
    bucket_new_link = []   # supplier known + new council link
    bucket_new_supp = []   # brand new supplier

    seen_new_suppliers = {}   # nl -> first contract that introduced them

    for c in contracts:
        for sup in c['suppliers']:
            if not sup['name'] or sup['name'].lower() in {'none', 'na', 'tbc'}:
                continue
            nl = norm(sup['name'])
            existing_name = known_suppliers.get(nl)
            pair = (nl, norm(c['council']))
            row = {**sup, 'contract': c, 'name_norm': nl}
            if existing_name:
                if pair in existing_pairs:
                    bucket_known.append((existing_name, row))
                else:
                    bucket_new_link.append((existing_name, row))
                    existing_pairs.add(pair)
            else:
                if nl in seen_new_suppliers:
                    # same new supplier shows up in another contract — track both
                    bucket_new_link.append((sup['name'], row))
                else:
                    seen_new_suppliers[nl] = row
                    bucket_new_supp.append(row)

    print(f"\n  Classification:")
    print(f"    Already in DB (supplier+council):  {len(bucket_known)}")
    print(f"    New council link to known supplier:{len(bucket_new_link)}")
    print(f"    Brand new supplier:                {len(bucket_new_supp)}")
    return bucket_known, bucket_new_link, bucket_new_supp

# ── 4. WRITE to manual_contracts.xlsx ──────────────────────────────────────────
def write_to_manual(bucket_new_link, bucket_new_supp):
    wb = openpyxl.load_workbook(MANUAL_XL)
    comp = wb['Companies']
    ccs_name = 'Company x Council x Sector' if 'Company x Council x Sector' in wb.sheetnames \
               else 'Company × Council × Sector'
    ccs = wb[ccs_name]

    # 4a. Add new suppliers to Companies sheet
    existing_companies = {(comp.cell(row=r, column=1).value or '').strip().lower()
                          for r in range(2, comp.max_row+1)
                          if comp.cell(row=r, column=1).value}
    n_added = 0
    for sup in bucket_new_supp:
        nl = sup['name'].lower().strip()
        if nl in existing_companies: continue
        nr = comp.max_row + 1
        comp.cell(row=nr, column=1).value  = sup['name']
        comp.cell(row=nr, column=2).value  = 0
        comp.cell(row=nr, column=5).value  = 1
        comp.cell(row=nr, column=7).value  = 'No'
        comp.cell(row=nr, column=8).value  = sup['ch']
        comp.cell(row=nr, column=9).value  = 'Yes' if sup['sme'] else ('No' if sup['sme'] is False else None)
        comp.cell(row=nr, column=10).value = 'Yes' if sup['vcse'] else ('No' if sup['vcse'] is False else None)
        comp.cell(row=nr, column=12).value = sup['addr']
        existing_companies.add(nl)
        n_added += 1

    # 4b. Add contract rows to CCS sheet
    n_ccs = 0
    for canonical_name, row in bucket_new_link:
        c = row['contract']
        nr = ccs.max_row + 1
        citation = (f"{c['title'][:80]} - {c['council']}. "
                    f"Awarded {c['awarded_date']}. Value GBP {c['value']}. "
                    f"Notice {c['notice_id'][:30]}.")
        ccs.cell(row=nr, column=1).value  = canonical_name
        ccs.cell(row=nr, column=2).value  = 'Housing'
        ccs.cell(row=nr, column=3).value  = c['council']
        ccs.cell(row=nr, column=4).value  = 1
        ccs.cell(row=nr, column=9).value  = 'Yes' if row.get('sme') else ('No' if row.get('sme') is False else None)
        ccs.cell(row=nr, column=10).value = 'Yes' if row.get('vcse') else ('No' if row.get('vcse') is False else None)
        ccs.cell(row=nr, column=11).value = 'Supported living | Learning disability | Mental health'
        ccs.cell(row=nr, column=13).value = citation
        ccs.cell(row=nr, column=14).value = c['region']
        ccs.cell(row=nr, column=15).value = 'Local Authority'
        ccs.cell(row=nr, column=16).value = 'Local'
        n_ccs += 1

    # 4c. Also add CCS rows for the brand-new suppliers' first contract
    for sup in bucket_new_supp:
        c = sup['contract']
        nr = ccs.max_row + 1
        citation = (f"{c['title'][:80]} - {c['council']}. "
                    f"Awarded {c['awarded_date']}. Value GBP {c['value']}. "
                    f"Notice {c['notice_id'][:30]}.")
        ccs.cell(row=nr, column=1).value  = sup['name']
        ccs.cell(row=nr, column=2).value  = 'Housing'
        ccs.cell(row=nr, column=3).value  = c['council']
        ccs.cell(row=nr, column=4).value  = 1
        ccs.cell(row=nr, column=9).value  = 'Yes' if sup.get('sme') else ('No' if sup.get('sme') is False else None)
        ccs.cell(row=nr, column=10).value = 'Yes' if sup.get('vcse') else ('No' if sup.get('vcse') is False else None)
        ccs.cell(row=nr, column=11).value = 'Supported living | Learning disability | Mental health'
        ccs.cell(row=nr, column=13).value = citation
        ccs.cell(row=nr, column=14).value = c['region']
        ccs.cell(row=nr, column=15).value = 'Local Authority'
        ccs.cell(row=nr, column=16).value = 'Local'
        n_ccs += 1

    wb.save(MANUAL_XL)
    print(f"\n  Wrote: {n_added} new Companies rows, {n_ccs} CCS rows", flush=True)
    return n_added, n_ccs

# ── 5. VERIFY new suppliers via Firecrawl ─────────────────────────────────────
def verify_new_suppliers(bucket_new_supp, skip_verify=False):
    if skip_verify or not bucket_new_supp:
        print("  Skipping Firecrawl verification", flush=True)
        return
    print(f"\n  Running Firecrawl on {len(bucket_new_supp)} new suppliers...", flush=True)
    # Need to rebuild providers.json first so the search script sees them
    subprocess.run(['python', 'build_data.py'], check=True,
                   stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
    # Then run the search-and-verify pass — it will only process providers
    # missing from VERIFIED.json (which all our new suppliers are)
    subprocess.run(['python', '-u', 'scripts/firecrawl_search_unverified.py'],
                   check=False)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2 or not Path(sys.argv[1]).exists():
        print(__doc__); sys.exit(1)

    csv_path = sys.argv[1]
    skip_verify = '--no-verify' in sys.argv
    deploy     = '--deploy' in sys.argv

    print(f"\n=== STAGE 1: PARSE — {csv_path} ===", flush=True)
    contracts = parse_csv(csv_path)
    if not contracts:
        print("Nothing to ingest."); return

    print(f"\n=== STAGE 2-3: CHECK + CLASSIFY ===", flush=True)
    known, new_link, new_supp = classify_contracts(contracts)

    print(f"\n=== STAGE 4: WRITE to manual_contracts.xlsx ===", flush=True)
    write_to_manual(new_link, new_supp)

    print(f"\n=== STAGE 5: VERIFY new suppliers (Firecrawl) ===", flush=True)
    verify_new_suppliers(new_supp, skip_verify)

    print(f"\n=== STAGE 6-7: APPLY verdicts + REBUILD ===", flush=True)
    if not skip_verify:
        subprocess.run(['python', 'scripts/apply_firecrawl.py'], check=False)
    subprocess.run(['python', 'build_data.py'], check=True)
    subprocess.run(['npm', 'run', 'build'], check=False,
                   stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT, shell=True)

    if deploy:
        print(f"\n=== STAGE 8: DEPLOY ===", flush=True)
        subprocess.run(['vercel', 'deploy', '--prod', '--yes'], check=False, shell=True)

    # Final report
    print(f"\n=== DONE ===", flush=True)
    prov = json.load(open(PROVIDERS, encoding='utf-8'))
    v = sum(1 for x in prov if x.get('verification',{}).get('verified'))
    print(f"  Total providers: {len(prov)}  ({v} Verified)")
    print(f"  New suppliers added: {len(new_supp)}")
    print(f"  New council links:   {len(new_link)}")
    print(f"  Already on file:     {len(known)}")
    if not deploy:
        print(f"\n  NOTE: not deployed. Add --deploy to push live, or run:")
        print(f"    vercel deploy --prod --yes")

if __name__ == '__main__':
    main()
