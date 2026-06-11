"""
Auto-approve rows that meet the deterministic London-supported-housing test.

Decision tree:
  1. Already-marked N (quality-flag misclassifications): keep N.
  2. Already-marked ? (frameworks/placeholders): keep ?.
  3. Council is a London borough / pan-London body / London consortium AND
     supplier is a real named provider (not the framework name itself):  Y
  4. Council is outside London:  N
  5. Anything else:  leave blank (manual review).
"""
import openpyxl, sys, io, re
from collections import Counter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

REVIEW = 'data/scraped/REVIEW_PACK.xlsx'

COL_APPROVE  = 1
COL_FLAG     = 2
COL_SUPPLIER = 5
COL_COUNCIL  = 14   # raw council column in review pack

LONDON_BOROUGH_KEYS = [
    'barking','dagenham','barnet','bexley','brent','bromley','camden',
    'city of london','croydon','ealing','enfield','greenwich','hackney',
    'hammersmith','fulham','haringey','harrow','havering','hillingdon',
    'hounslow','islington','kensington','rbkc','kingston','lambeth',
    'lewisham','merton','newham','redbridge','richmond','southwark',
    'sutton','tower hamlets','waltham','wandsworth','westminster','wcc',
    # pan-London / consortia
    'capital letters','greater london authority','mopac','mayor',
    'london councils','ministry of housing','west london alliance',
    'commissioning alliance','liia','achieving for children',
]

def is_london(council_raw):
    cl = (council_raw or '').lower()
    return any(k in cl for k in LONDON_BOROUGH_KEYS)

def is_framework_placeholder(supplier):
    s = (supplier or '').lower()
    return (
        s.startswith('london borough of')
        or 'pdps (multiple suppliers' in s
        or 'framework (suppliers' in s
        or s in ('tbc','tbd','various','multiple')
    )

wb = openpyxl.load_workbook(REVIEW)
ws = wb['Review']
counts = Counter()

for row in ws.iter_rows(min_row=2):
    current = (row[COL_APPROVE-1].value or '').strip().upper()
    flag = (row[COL_FLAG-1].value or '').strip()
    supplier = str(row[COL_SUPPLIER-1].value or '').strip()
    council = str(row[COL_COUNCIL-1].value or '').strip()

    # Preserve existing decisions
    if current == 'N':
        counts['keep N (quality flag)'] += 1
        continue
    if current == '?':
        counts['keep ? (framework)'] += 1
        continue
    if current == 'Y':
        counts['keep Y'] += 1
        continue

    if not council or not supplier:
        counts['blank — missing data'] += 1
        continue

    if is_framework_placeholder(supplier):
        row[COL_APPROVE-1].value = '?'
        counts['mark ? (framework placeholder)'] += 1
        continue

    if is_london(council):
        row[COL_APPROVE-1].value = 'Y'
        counts['mark Y (London supplier)'] += 1
    else:
        row[COL_APPROVE-1].value = 'N'
        counts['mark N (outside London)'] += 1

wb.save(REVIEW)

print("DECISION SUMMARY")
for k, n in sorted(counts.items(), key=lambda x: -x[1]):
    print(f"  {n:4d}  {k}")

# Final tallies
final = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    v = (row[0] or '').strip().upper()
    if v in ('Y','YES'): k='Y'
    elif v in ('N','NO'): k='N'
    elif v in ('?','MAYBE'): k='?'
    else: k='_'
    final[k] += 1
print(f"\nFINAL: Y:{final['Y']}  N:{final['N']}  ?:{final['?']}  unmarked:{final['_']}  total:{sum(final.values())}")
