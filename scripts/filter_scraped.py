"""
Apply the main pipeline's filters to scraped rows.

The scraper's permissive HOUSING_TITLE matches a lot of false-positives
(legal services on housing programmes, housing repairs, housing software,
housing audits, bin housing(!), etc.). This script applies the same
NONCARE_NAME / IRRELEVANT_TITLE / HOMECARE_OVERRIDE filters the main
pipeline uses, and writes:

    data/scraped/normalised_filtered.xlsx

Run dedup_report.py against THAT instead of the raw normalised file for a
honest "what's actually missing" report.
"""
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl
import build_data as bd
from scrapers.base import CONTRACT_FIELDS

IN = os.path.join(ROOT, "data", "scraped", "normalised.xlsx")
OUT = os.path.join(ROOT, "data", "scraped", "normalised_filtered.xlsx")


def main():
    if not os.path.exists(IN):
        print(f"missing {IN}"); return
    wb_in = openpyxl.load_workbook(IN, read_only=True, data_only=True)
    ws_in = wb_in["Company × Council × Sector"]
    hdr = [c.value for c in next(ws_in.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(hdr, r)) for r in ws_in.iter_rows(min_row=2, values_only=True) if r and r[0]]
    wb_in.close()

    print(f"input: {len(rows)} rows from {IN}")

    kept, dropped_by_reason = [], {"NONCARE_NAME":0, "JUNK_NAME":0, "junk_title":0,
                                    "HOMECARE_OVERRIDE":0, "EXCLUDE":0, "COMMISSIONER_EXCLUDE":0}
    for r in rows:
        company = str(r.get("Company") or "").strip()
        title = str(r.get("Contract Titles") or "").strip()
        council = str(r.get("Council") or "").strip()

        # main-pipeline filters mirrored from build_data.py
        if bd.JUNK_NAME.match(company):
            dropped_by_reason["JUNK_NAME"] += 1; continue
        if bd.NONCARE_NAME.search(company):
            dropped_by_reason["NONCARE_NAME"] += 1; continue
        # soft-noncare (construction/engineering majors) — same rule
        if bd.NONCARE_SOFT.search(company) and not bd.CARE_EXEMPT.search(company):
            dropped_by_reason["NONCARE_NAME"] += 1; continue
        if bd.COMMISSIONER_EXCLUDE.search(council):
            dropped_by_reason["COMMISSIONER_EXCLUDE"] += 1; continue
        if bd.is_junk_title(title):
            dropped_by_reason["junk_title"] += 1; continue
        # homecare-override: care delivered, not housing provided
        if bd.HOMECARE_OVERRIDE.search(title) and not bd.HOUSING_TITLE.search(title):
            dropped_by_reason["HOMECARE_OVERRIDE"] += 1; continue
        kept.append(r)

    print(f"\nkept: {len(kept)} rows")
    for k, n in dropped_by_reason.items():
        if n: print(f"  dropped ({k}): {n}")

    wb_out = openpyxl.Workbook(); wb_out.remove(wb_out.active)
    ws = wb_out.create_sheet("Company × Council × Sector")
    ws.append(CONTRACT_FIELDS)
    for r in kept:
        ws.append([r.get(k, "") for k in CONTRACT_FIELDS])
    wb_out.save(OUT)
    print(f"\nwrote {OUT}")
    print("\nNext: re-run dedup against the filtered file:")
    print("  python scripts/dedup_report.py --scraped data/scraped/normalised_filtered.xlsx")


if __name__ == "__main__":
    main()
