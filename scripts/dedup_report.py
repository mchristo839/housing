"""
Compare data/scraped/normalised.xlsx against the main workbook + manual
additions, and write four audit outputs to data/scraped/audit/:

    new_providers.csv         — companies never seen in main+manual
    new_contracts.csv         — company+council pairs never seen
    potential_duplicates.csv  — fuzzy-title matches needing human review
    coverage_gap.md           — readable summary

Dedup rules (per user choice):
  - "Already known" if norm(company) == norm(existing company)
    AND norm(council) == norm(existing council).
  - Within "Already known", if the title's similarity to any existing title
    is < 0.85, the scrape is flagged as a *potential new contract for an
    existing pair* (often a new framework round). These go to
    potential_duplicates.csv for human review.
  - Anything not "Already known" goes to new_contracts.csv.
  - Companies that don't appear in main+manual at all go to new_providers.csv.

Usage:
    python scripts/dedup_report.py
"""
import csv
import os
import re
import sys
from collections import defaultdict
from difflib import SequenceMatcher

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

import openpyxl

MAIN_XLSX = os.path.join(ROOT, "data", "care_housing_database_v2_ENRICHED.xlsx")
MANUAL_XLSX = os.path.join(ROOT, "data", "manual_contracts.xlsx")
SCRAPED_XLSX = os.path.join(ROOT, "data", "scraped", "normalised.xlsx")
AUDIT_DIR = os.path.join(ROOT, "data", "scraped", "audit")

DROP = {"limited", "ltd", "plc", "cic", "llp", "group", "the", "uk", "t", "a",
        "trust", "services", "service", "association", "ha", "society"}

def norm_company(s):
    s = re.sub(r"[^a-z0-9 ]", " ", str(s or "").lower().replace("&", " and "))
    return " ".join(t for t in s.split() if t and t not in DROP)

# council normalisation: similar but keep more words (so K&C ≠ Westminster etc.)
COUNCIL_DROP = {"council", "borough", "metropolitan", "district", "unitary",
                "authority", "the", "of", "city", "mbc", "lbc", "mdc", "cc"}

def norm_council(s):
    s = re.sub(r"[^a-z0-9 ]", " ", str(s or "").lower().replace("&", " and "))
    return " ".join(t for t in s.split() if t and t not in COUNCIL_DROP)

def title_sim(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def load_known():
    """Build dict company_key → set(council_key) and dict (company_key,council_key) → list(title)
    from main + manual workbooks."""
    company_councils = defaultdict(set)
    pair_titles = defaultdict(list)
    company_display = {}
    for src in (MAIN_XLSX, MANUAL_XLSX):
        if not os.path.exists(src):
            continue
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        if "Company × Council × Sector" not in wb.sheetnames:
            wb.close(); continue
        ws = wb["Company × Council × Sector"]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            ic = hdr.index("Company"); icl = hdr.index("Council"); it = hdr.index("Contract Titles")
        except ValueError:
            wb.close(); continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r[ic]: continue
            ck = norm_company(r[ic])
            lk = norm_council(r[icl]) if icl < len(r) else ""
            t = str(r[it] or "") if it < len(r) else ""
            if ck:
                company_councils[ck].add(lk)
                company_display.setdefault(ck, str(r[ic]).strip())
                if lk:
                    # split on ; for multi-title cells
                    for piece in re.split(r"\s*;\s*", t):
                        if piece.strip():
                            pair_titles[(ck, lk)].append(piece.strip())
        wb.close()
    return company_councils, pair_titles, company_display


def load_scraped():
    if not os.path.exists(SCRAPED_XLSX):
        raise SystemExit(f"missing {SCRAPED_XLSX} — run scripts/normalise.py first")
    wb = openpyxl.load_workbook(SCRAPED_XLSX, read_only=True, data_only=True)
    ws = wb["Company × Council × Sector"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    wb.close()
    return hdr, rows


def main():
    os.makedirs(AUDIT_DIR, exist_ok=True)
    known_company_councils, known_pair_titles, known_display = load_known()
    hdr, scraped = load_scraped()
    print(f"main + manual: {len(known_company_councils)} distinct companies, "
          f"{sum(len(v) for v in known_company_councils.values())} (company,council) pairs")
    print(f"scraped: {len(scraped)} rows")

    new_providers = defaultdict(list)         # company_key → list of scraped rows
    new_contracts = []                         # rows where pair is new (existing company)
    potential_duplicates = []                  # rows where pair exists but title is novel

    for r in scraped:
        company = str(r.get("Company") or "").strip()
        council = str(r.get("Council") or "").strip()
        title = str(r.get("Contract Titles") or "").strip()
        ck = norm_company(company)
        lk = norm_council(council)
        if not ck:
            continue
        if ck not in known_company_councils:
            new_providers[ck].append(r)
            continue
        if lk not in known_company_councils[ck]:
            new_contracts.append(r)
            continue
        # company+council pair known — is the title novel?
        existing_titles = known_pair_titles.get((ck, lk), [])
        max_sim = max((title_sim(title, t) for t in existing_titles), default=0.0)
        if max_sim < 0.85:
            r["_max_title_similarity"] = round(max_sim, 2)
            potential_duplicates.append(r)
        # else: silently considered a duplicate of an existing row

    # ── Write outputs ───────────────────────────────────────────────────────
    def write_csv(path, headers, rows):
        with open(path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for row in rows:
                w.writerow([row.get(h, "") for h in headers])

    AUDIT_FIELDS = ["Company", "Sector", "Council", "Contracts (this council, this sector)",
                    "Categories", "Most Recent Award", "Contract Titles",
                    "ONS Region", "Commissioner Type", "Geographic Scope",
                    "source_portal", "source_url", "contract_value_gbp", "cpv_codes"]

    # new_providers: flatten and write
    flat_new_providers = []
    for ck, rows in new_providers.items():
        for r in rows:
            flat_new_providers.append(r)
    flat_new_providers.sort(key=lambda r: str(r.get("Company") or ""))
    write_csv(os.path.join(AUDIT_DIR, "new_providers.csv"),
              ["scrape_action"] + AUDIT_FIELDS,
              [dict(r, scrape_action="ADD_PROVIDER_AND_CONTRACT") for r in flat_new_providers])

    write_csv(os.path.join(AUDIT_DIR, "new_contracts.csv"),
              ["scrape_action"] + AUDIT_FIELDS,
              [dict(r, scrape_action="ADD_CONTRACT_TO_EXISTING_PROVIDER") for r in new_contracts])

    write_csv(os.path.join(AUDIT_DIR, "potential_duplicates.csv"),
              ["scrape_action", "_max_title_similarity"] + AUDIT_FIELDS,
              [dict(r, scrape_action="REVIEW_POSSIBLE_NEW_AWARD") for r in potential_duplicates])

    # ── Coverage-gap markdown report ────────────────────────────────────────
    md_path = os.path.join(AUDIT_DIR, "coverage_gap.md")
    distinct_new_providers = len(new_providers)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# Scrape coverage gap report\n\n")
        f.write(f"Scraped rows reviewed: **{len(scraped)}**.\n\n")
        f.write("## What I found\n\n")
        f.write(f"| Bucket | Rows | What to do |\n|---|---:|---|\n")
        f.write(f"| New providers (never seen) | {len(flat_new_providers)} across "
                f"{distinct_new_providers} companies | Review `new_providers.csv`, then promote |\n")
        f.write(f"| New contracts (existing provider, new council/region) | {len(new_contracts)} | Review `new_contracts.csv`, then promote |\n")
        f.write(f"| Possible new awards on existing pairs | {len(potential_duplicates)} | Review `potential_duplicates.csv` — likely a re-tender of an existing framework |\n")
        f.write("\n## Top 20 new providers by row count\n\n")
        by_count = sorted(new_providers.items(), key=lambda x: -len(x[1]))[:20]
        if not by_count:
            f.write("_None — every scraped company is already in main+manual._\n")
        else:
            f.write("| Provider | Scraped rows | Sample council |\n|---|---:|---|\n")
            for ck, rs in by_count:
                sample = str(rs[0].get("Council") or "")[:40]
                disp = str(rs[0].get("Company") or "")
                f.write(f"| {disp} | {len(rs)} | {sample} |\n")
        f.write("\n## Recommended next step\n\n")
        f.write("```\npython scripts/promote_to_manual.py "
                "--from data/scraped/audit/new_providers.csv\n```\n")

    print(f"\nWrote:")
    print(f"  {os.path.join(AUDIT_DIR, 'new_providers.csv')}        — {len(flat_new_providers)} rows ({distinct_new_providers} new companies)")
    print(f"  {os.path.join(AUDIT_DIR, 'new_contracts.csv')}        — {len(new_contracts)} rows")
    print(f"  {os.path.join(AUDIT_DIR, 'potential_duplicates.csv')} — {len(potential_duplicates)} rows for review")
    print(f"  {md_path}")
    print(f"\nNext: review the CSVs and run python scripts/promote_to_manual.py")


if __name__ == "__main__":
    main()
