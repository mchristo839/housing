"""Take a decisions CSV exported from ADDING_TO_DATABASE.html and apply
to manual_contracts.xlsx:
  approve -> keep the rows (no change needed — already there)
  delete  -> remove every contract row + every Companies row for that entity
  maybe   -> leave alone (defer)

Usage:
  python scripts/apply_db_decisions.py adding_to_db_decisions_YYYY-MM-DD.csv
"""
import openpyxl, csv, sys, io, os, re, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

MANUAL = 'data/manual_contracts.xlsx'

def norm_id(s):
    return re.sub(r'[^a-z0-9]', '_', str(s or '').lower())[:60]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('csv_path')
    a = ap.parse_args()

    if not os.path.exists(a.csv_path):
        print(f"ERR: {a.csv_path} not found")
        sys.exit(1)

    # Read decisions
    decisions = {}   # entity_id -> ('approve'|'delete'|'maybe', entity_name)
    with open(a.csv_path, encoding='utf-8-sig') as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            eid = row.get('entity_id','').strip()
            dec = row.get('decision','').strip().lower()
            name = row.get('entity_name','').strip().strip('"')
            if eid and dec:
                decisions[eid] = (dec, name)

    n_approve = sum(1 for d,_ in decisions.values() if d == 'approve')
    n_delete  = sum(1 for d,_ in decisions.values() if d == 'delete')
    n_maybe   = sum(1 for d,_ in decisions.values() if d == 'maybe')
    print(f"Loaded decisions from {a.csv_path}:")
    print(f"  Approve : {n_approve}")
    print(f"  Reject  : {n_delete}")
    print(f"  Maybe   : {n_maybe}")
    if n_delete == 0:
        print("\nNo rejected entries — manual_contracts.xlsx already reflects approve/maybe state.")
        return

    # Build the set of entity-IDs to delete
    to_delete = {eid for eid, (dec,_) in decisions.items() if dec == 'delete'}
    names_to_delete = {nm.lower() for eid,(dec,nm) in decisions.items() if dec == 'delete'}

    # Walk manual file — remove every row whose supplier matches a deleted id/name
    wb = openpyxl.load_workbook(MANUAL)

    # 1. Company × Council × Sector — delete contract rows
    ws_cs = wb['Company × Council × Sector']
    rows_to_delete_idx = []
    for i, row in enumerate(ws_cs.iter_rows(min_row=2), start=2):
        nm = str(row[0].value or '').strip()
        if not nm: continue
        if norm_id(nm) in to_delete or nm.lower() in names_to_delete:
            rows_to_delete_idx.append(i)
    print(f"\nDeleting {len(rows_to_delete_idx)} rows from Company × Council × Sector")
    # Delete from bottom up to keep indices stable
    for i in sorted(rows_to_delete_idx, reverse=True):
        ws_cs.delete_rows(i)

    # 2. Companies sheet — delete entries
    ws_co = wb['Companies']
    co_to_delete = []
    for i, row in enumerate(ws_co.iter_rows(min_row=2), start=2):
        nm = str(row[0].value or '').strip()
        if not nm: continue
        if norm_id(nm) in to_delete or nm.lower() in names_to_delete:
            co_to_delete.append(i)
    print(f"Deleting {len(co_to_delete)} rows from Companies")
    for i in sorted(co_to_delete, reverse=True):
        ws_co.delete_rows(i)

    wb.save(MANUAL)
    print(f"\nFinal state:")
    print(f"  Company × Council × Sector : {ws_cs.max_row - 1} rows")
    print(f"  Companies                  : {ws_co.max_row - 1} rows")
    print(f"\nNow run: python build_data.py")


if __name__ == '__main__':
    main()
