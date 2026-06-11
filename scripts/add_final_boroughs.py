"""Final batch — Hackney, Waltham Forest, Bromley, Redbridge, Islington."""
import openpyxl, sys, io, datetime
from collections import Counter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = 'data/scraped/curated_scraped.xlsx'
wb = openpyxl.load_workbook(XLSX)
ws = wb['Company × Council × Sector']
now = datetime.datetime.now().isoformat()

rows = [
    {"co": "London Borough of Hackney AHI Supported Living Framework (suppliers confidential)",
     "council": "London Borough of Hackney",
     "title": "AHI S449 Supported Living Services (SLS) & Residential Open Framework - 3 lots (supported living, residential, block packages). 805.7M / 8 years from October 2025. Suppliers in exempt appendix.",
     "cats": "Supported living | Residential care | Adults",
     "award": "01/10/2025", "val": 805700000,
     "url": "https://hackney.moderngov.co.uk/ieDecisionDetails.aspx?Id=8210"},
    {"co": "London Borough of Waltham Forest 16+ Semi-Independent Framework (suppliers via mini-competition)",
     "council": "London Borough of Waltham Forest",
     "title": "16+ Semi-Independent Accommodation and Support Services Framework for Looked After Children and Care Experienced Young Adults. 16M / 4yr (1 June 2022 - 31 May 2024) + 24mo extension to 31 May 2026.",
     "cats": "Supported accommodation | Care leavers | Looked after children",
     "award": "01/06/2022", "val": 16000000,
     "url": "https://bidstats.uk/tenders/2022/W28/778567851"},
    {"co": "Riverside", "council": "London Borough of Bromley",
     "title": "Bromley Supported Housing - Riverside operates supported housing in Bromley.",
     "cats": "Supported housing | Adults",
     "award": "", "val": "",
     "url": "https://www.riverside.org.uk/in-your-neighbourhood/bromley-2/care-and-support/bromley-supported-housing/"},
    {"co": "London Borough of Bromley Extra Care Housing Re-Procurement",
     "council": "London Borough of Bromley",
     "title": "5 Extra Care Housing schemes - re-procurement of care and support services launched 2025. Find-a-Tender Notice 007384-2025.",
     "cats": "Extra care | Older people",
     "award": "06/03/2025", "val": "",
     "url": "https://www.find-tender.service.gov.uk/Notice/007384-2025"},
    {"co": "London Borough of Redbridge Supported Living (housing associations & voluntary orgs)",
     "council": "London Borough of Redbridge",
     "title": "Supported living operated by mix of housing associations, voluntary organisations, and the council. New commissioning model 1 May 2025 - 30 April 2029.",
     "cats": "Supported living | Learning disabilities | Mental health | Autism",
     "award": "01/05/2025", "val": "",
     "url": "https://bidstats.uk/tenders/2024/W40/831979788"},
    {"co": "London Borough of Islington Mental Health Accommodation Pathway",
     "council": "London Borough of Islington",
     "title": "Mental Health Accommodation Pathway Contract Award (Islington). Plus 2425-0692 Supported Housing and Floating Support Framework. Specific provider in democracy portal documents.",
     "cats": "Supported accommodation | Mental health | Floating support",
     "award": "", "val": "",
     "url": "https://democracy.islington.gov.uk/documents/s25987/Mental%20Health%20Accommodation%20Pathway%20Contract%20Award%20Report.pdf"},
]

added = 0
for r in rows:
    row = [r["co"], "Housing", r["council"], 1, "", "", "", "", "", "",
           r["cats"], r.get("award", ""), r["title"], "London", "Local Council", "Local", ""
    ] + ["council-cabinet-doc", r["url"], "", now, "", str(r.get("val", "")), ""]
    ws.append(row)
    added += 1

wb.save(XLSX)
total = ws.max_row - 1
print(f"Added {added} final rows. Review DB total: {total}\n")

def borough_key(c):
    if not c:
        return None
    cl = c.lower()
    M = {'barking':'Barking and Dagenham','dagenham':'Barking and Dagenham','barnet':'Barnet','bexley':'Bexley','brent':'Brent','bromley':'Bromley','camden':'Camden','city of london':'City of London','croydon':'Croydon','ealing':'Ealing','enfield':'Enfield','greenwich':'Greenwich','hackney':'Hackney','hammersmith':'Hammersmith and Fulham','fulham':'Hammersmith and Fulham','haringey':'Haringey','harrow':'Harrow','havering':'Havering','hillingdon':'Hillingdon','hounslow':'Hounslow','islington':'Islington','kensington':'Kensington and Chelsea','rbkc':'Kensington and Chelsea','kingston':'Kingston upon Thames','lambeth':'Lambeth','lewisham':'Lewisham','merton':'Merton','newham':'Newham','redbridge':'Redbridge','richmond':'Richmond upon Thames','southwark':'Southwark','sutton':'Sutton','tower hamlets':'Tower Hamlets','waltham':'Waltham Forest','wandsworth':'Wandsworth','westminster':'Westminster','wcc':'Westminster'}
    for key, std in M.items():
        if key in cl:
            return std
    return None

all_rows = list(ws.iter_rows(min_row=2, values_only=True))
borough_count = Counter()
for r in all_rows:
    if not r[0]: continue
    b = borough_key(r[2])
    if b: borough_count[b] += 1

ALL_B = ['Barking and Dagenham','Barnet','Bexley','Brent','Bromley','Camden','City of London','Croydon','Ealing','Enfield','Greenwich','Hackney','Hammersmith and Fulham','Haringey','Harrow','Havering','Hillingdon','Hounslow','Islington','Kensington and Chelsea','Kingston upon Thames','Lambeth','Lewisham','Merton','Newham','Redbridge','Richmond upon Thames','Southwark','Sutton','Tower Hamlets','Waltham Forest','Wandsworth','Westminster']
covered = sum(1 for b in ALL_B if b in borough_count)

print(f"LONDON COVERAGE FINAL: {covered}/33 BOROUGHS")
print(f"Total rows: {total}")
print(f"Unique suppliers: {len(set(r[0] for r in all_rows if r[0]))}\n")

print(f"{'BOROUGH':<28s} {'ROWS':>5s}")
print('-' * 38)
for b in ALL_B:
    n = borough_count.get(b, 0)
    mark = "OK" if n > 0 else "--"
    print(f"  {mark} {b[:25]:<26s} {n:>5d}")

missing = [b for b in ALL_B if b not in borough_count]
if missing:
    print(f"\nStill missing: {', '.join(missing)}")
else:
    print(f"\n*** ALL 33 LONDON BOROUGHS COVERED ***")
