"""Add Greenwich + Harrow + Lambeth contract rows + Companies entries."""
import openpyxl, datetime, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = 'data/manual_contracts.xlsx'
wb = openpyxl.load_workbook(XLSX)
ws_cs = wb['Company × Council × Sector']
ws_co = wb['Companies']

CONTRACT_ROWS = [
    # GREENWICH
    ["Alcove Limited", "Housing", "Royal Borough of Greenwich", 1, "", "", "", "09076735", "", "",
     "Supported living | Assistive technology", "01/04/2024",
     "Assistive Technology Enabled Care (ATEC) Service - Royal Borough of Greenwich. Tech and data services for supported living. 17M up to 5+2+2+1 year term.",
     "London", "Local Council", "Local", ""],
    ["Greenwich Living Options", "Housing", "Royal Borough of Greenwich", 1, "", "", "", "", "", "",
     "Supported living | Learning disabilities | Adults", "01/07/2024",
     "Royal Hill Independent Living Service - RB Greenwich in-house provider. Opened July 2024.",
     "London", "Local Council", "Local", ""],
    # HARROW
    ["YMCA St Pauls Group", "Housing", "London Borough of Harrow", 1, "", "", "", "", "", "",
     "Supported accommodation | Care leavers | Looked after children | Homelessness | Young people",
     "01/04/2025",
     "Supported Accommodation for LAC, Care Leavers and Homeless Young People - direct award. 10yr (Apr 2025 to Mar 2035) for 29 young people.",
     "London", "Local Council", "Local", ""],
    # LAMBETH
    ["Cera Care", "Housing", "London Borough of Lambeth", 1, "", "", "", "", "", "",
     "Extra care | Older people | Care and support", "01/06/2023",
     "Extra Care Housing - Care and Support (Lambeth Bidstats 800129262 Award, 15M-36M).",
     "London", "Local Council", "Local", ""],
    ["London Care", "Housing", "London Borough of Lambeth", 1, "", "", "", "", "", "",
     "Extra care | Older people | Care and support", "01/06/2023",
     "Extra Care Housing - Care and Support (Lambeth Bidstats 800129262 Award, 15M-36M).",
     "London", "Local Council", "Local", ""],
    ["Care Outlook", "Housing", "London Borough of Lambeth", 1, "", "", "", "", "", "",
     "Extra care | Older people | Care and support", "01/06/2023",
     "Extra Care Housing - Care and Support (Lambeth Bidstats 800129262 Award, 15M-36M).",
     "London", "Local Council", "Local", ""],
    ["Hestia Housing and Support", "Housing", "London Borough of Lambeth", 1, "", "", "", "", "", "",
     "Floating support | Disabilities | HIV", "01/01/2022",
     "Floating Support Service for People with Physical/Neurological/Sensory Disabilities and HIV/AIDS. 62 people supported in Lambeth. Active per Hestia provider site.",
     "London", "Local Council", "Local", ""],
    # FRAMEWORK PLACEHOLDERS
    ["Greenwich Adults and Children Learning Disability Framework",
     "Housing", "Royal Borough of Greenwich", 1, "", "", "", "", "", "",
     "Supported living | Learning disabilities | Home care | Extra care", "01/01/2022",
     "Adults + Children Learning Disability + Home + Extra Care framework. 169M-180M. Multiple awarded providers per lot (Bidstats 753270910).",
     "London", "Local Council", "Local", ""],
    ["Greenwich Mental Health High Support Accommodation",
     "Housing", "Royal Borough of Greenwich", 1, "", "", "", "", "", "",
     "Supported accommodation | Mental health | Adults", "01/01/2020",
     "Accommodation-Based Care and Support Services for People With Mental Health Needs. 4M+. (Bidstats 718536955).",
     "London", "Local Council", "Local", ""],
    ["Harrow Mental Health Supported Accommodation Block Contract",
     "Housing", "London Borough of Harrow", 1, "", "", "", "", "", "",
     "Supported accommodation | Mental health | Adults", "01/04/2021",
     "Supported Accommodation for Mental Health Recovery - 22 units. Adults 18-65 resident in Harrow. Block contract.",
     "London", "Local Council", "Local", ""],
    ["Lambeth Adults Supported Accommodation Pathway Framework",
     "Housing", "London Borough of Lambeth", 1, "", "", "", "", "", "",
     "Supported accommodation | Mental health | Complex needs | Women services | Dual diagnosis",
     "01/04/2025",
     "Adults Supported Accommodation Pathway - 8 lots: Women Pathway 38 beds, Robertson Street 42 beds, Lambeth High Street 24 beds. 21M+ (Bidstats 851182130).",
     "London", "Local Council", "Local", ""],
    ["Lambeth Housing Related Support for High Risk Young People",
     "Housing", "London Borough of Lambeth", 1, "", "", "", "", "", "",
     "Supported accommodation | Young people | High risk | Care leavers", "01/04/2023",
     "Housing Related Support for High Risk Young People. Routes to Independence Pathway, London Homelessness Award 2024.",
     "London", "Local Council", "Local", ""],
]

added = 0
existing = {(str(r[0].value).strip().lower(), str(r[2].value).strip().lower())
            for r in ws_cs.iter_rows(min_row=2) if r[0].value and r[2].value}
for row in CONTRACT_ROWS:
    key = (row[0].strip().lower(), row[2].strip().lower())
    if key in existing:
        print(f"  skip CS: {row[0]}")
        continue
    existing.add(key)
    ws_cs.append(row)
    added += 1
print(f"Added {added} contract rows")

COMPANY_ROWS = [
    ["Alcove Limited", 1, 0, "", 1, "Royal Borough of Greenwich", 0, "09076735", 1, 0,
     "Royal Borough of Greenwich", "London, UK",
     "https://www.alcove.co.uk", "0203 950 5077", "info@alcove.co.uk",
     "https://www.alcove.co.uk/contact-us", "", "", "", "", "verified"],
    ["Greenwich Living Options", 1, 0, "", 1, "Royal Borough of Greenwich", 0, "", 0, 0,
     "Royal Borough of Greenwich", "Greenwich, London",
     "https://www.royalgreenwich.gov.uk/info/200244/learning_disabilities/507/learning_disabilities",
     "020 8921 4477", "asccentral@royalgreenwich.gov.uk",
     "https://www.royalgreenwich.gov.uk/contact", "", "", "", "", "in-house council"],
    ["YMCA St Pauls Group", 1, 0, "", 1, "London Borough of Harrow", 0, "04129079", 0, 1,
     "London Borough of Harrow", "London, UK",
     "https://www.ymcastpaulsgroup.org", "020 8688 1525", "info@ymcastpaulsgroup.org",
     "https://www.ymcastpaulsgroup.org/contact", "1102266", "", "Active", "Youth housing", "verified"],
    ["Greenwich Adults and Children Learning Disability Framework", 1, 0, "", 1,
     "Royal Borough of Greenwich", 0, "", 0, 0, "Royal Borough of Greenwich", "",
     "https://www.royalgreenwich.gov.uk/info/200244/learning_disabilities",
     "020 8921 4477", "asccentral@royalgreenwich.gov.uk",
     "https://www.royalgreenwich.gov.uk/contact", "", "", "", "", "framework"],
    ["Greenwich Mental Health High Support Accommodation", 1, 0, "", 1,
     "Royal Borough of Greenwich", 0, "", 0, 0, "Royal Borough of Greenwich", "",
     "https://www.royalgreenwich.gov.uk/info/200244/mental_health",
     "020 8921 4477", "asccentral@royalgreenwich.gov.uk",
     "https://www.royalgreenwich.gov.uk/contact", "", "", "", "", "framework"],
    ["Harrow Mental Health Supported Accommodation Block Contract", 1, 0, "", 1,
     "London Borough of Harrow", 0, "", 0, 0, "London Borough of Harrow", "",
     "https://www.harrow.gov.uk/adult-social-care",
     "020 8901 2680", "access.team@harrow.gov.uk",
     "https://www.harrow.gov.uk/contact", "", "", "", "", "framework"],
    ["Lambeth Adults Supported Accommodation Pathway Framework", 1, 0, "", 1,
     "London Borough of Lambeth", 0, "", 0, 0, "London Borough of Lambeth", "",
     "https://www.lambeth.gov.uk/lambeth-data-hub/supported-living",
     "020 7926 1000", "asc@lambeth.gov.uk",
     "https://www.lambeth.gov.uk/contact-us", "", "", "", "", "framework"],
    ["Lambeth Housing Related Support for High Risk Young People", 1, 0, "", 1,
     "London Borough of Lambeth", 0, "", 0, 0, "London Borough of Lambeth", "",
     "https://www.lambeth.gov.uk/support-care-leavers/housing-independent-living",
     "020 7926 1000", "asc@lambeth.gov.uk",
     "https://www.lambeth.gov.uk/contact-us", "", "", "", "", "framework"],
]

added_c = 0
existing_co = {str(r[0].value).strip().lower() for r in ws_co.iter_rows(min_row=2) if r[0].value}
for row in COMPANY_ROWS:
    if row[0].strip().lower() in existing_co:
        print(f"  skip Companies: {row[0]}")
        continue
    existing_co.add(row[0].strip().lower())
    ws_co.append(row)
    added_c += 1

print(f"Added {added_c} Companies rows")

wb.save(XLSX)
print(f"\nManual file now: {ws_cs.max_row-1} contract rows, {ws_co.max_row-1} Companies rows")
