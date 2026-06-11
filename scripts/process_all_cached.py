"""Process every cached raw scrape.
- Bidstats search JSON: 17 records already parsed
- Bidstats cached HTML: 113 notices
- Add net-new London supported-housing supplier rows to manual_contracts.xlsx
  (plus Companies entries with the commissioner-as-contact-route).
- Print a quality report so user can see what's working / what's noise.
"""
import openpyxl, glob, re, html, os, json, sys, io, datetime
from collections import Counter, defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

MANUAL = 'data/manual_contracts.xlsx'

def normname(s): return re.sub(r'[^a-z0-9]', '', str(s or '').lower())

# Load baselines
wb_m = openpyxl.load_workbook(MANUAL)
ws_cs = wb_m['Company × Council × Sector']
ws_co = wb_m['Companies']

import json as _json
prov = _json.load(open('api/_data/providers.json', encoding='utf-8'))
live_norm = {normname(p['name']) for p in prov}

manual_suppliers = set()
manual_keys = set()  # (supplier, council)
for r in ws_cs.iter_rows(min_row=2, values_only=True):
    if r[0]: manual_suppliers.add(normname(r[0]))
    if r[0] and r[2]:
        manual_keys.add((normname(r[0]), normname(r[2])))
existing_companies = {str(r[0].value).strip().lower() for r in ws_co.iter_rows(min_row=2) if r[0].value}

# Filters (same as audit)
LONDON_KW = re.compile(r'london borough|royal borough of greenwich|royal borough of kingston|westminster city|tower hamlets|hackney|islington|lambeth|southwark|lewisham|hounslow|brent|ealing|haringey|enfield|bexley|bromley|sutton|merton|wandsworth|hammersmith|kensington|camden|barnet|harrow|hillingdon|havering|redbridge|newham|waltham forest|barking|dagenham|croydon|capital letters|greater london authority|mopac|liia', re.I)
H_OK = re.compile(r'supported (living|accommodation|housing)|temporary accommodation|emergency accommodation|hostel|refuge|housing related support|extra care|sheltered (housing|accommodation)|young people.{0,40}(housing|accommodation|pathway)|care leavers|homeless|asylum|nightly paid|social housing|mental health.{0,40}(accommodation|housing|supported|step|crisis)|learning disabilit.{0,40}(accommodation|housing|supported|residential|placement)|domestic abuse|rough sleeper|housing first|semi[- ]?independent|floating support|move[- ]?on', re.I)
NOT_H = re.compile(r'\bhome care\b|home support service|home (domiciliary|based) care|aids and adaptations|catering|street lighting|grass cutting|repairs only|lift modernisation|cleaning services|legal services|insurance|civil enforcement|pan[- ]?london domiciliary|reablement only|hired passenger transport', re.I)
CARE_FACILITY = re.compile(r"\b(?:care home(?:s)?|nursing home(?:s)?|residential home(?:s)?|residential care home|care centre|care center|care lodge|care manor)\b", re.I)

COUNCIL_CONTACTS = {
    'london borough of barking and dagenham': ("https://www.lbbd.gov.uk/adult-health-and-social-care", "020 8215 3000", "adult.socialcare@lbbd.gov.uk", "https://www.lbbd.gov.uk/contact-us"),
    'london borough of barnet': ("https://www.barnet.gov.uk/adults-and-health", "020 8359 2000", "first.response@barnet.gov.uk", "https://www.barnet.gov.uk/contact-us"),
    'bexley': ("https://www.bexley.gov.uk/services/adult-social-care", "020 3045 5000", "adultsocialcare@bexley.gov.uk", "https://www.bexley.gov.uk/contact-us"),
    'london borough of brent': ("https://www.brent.gov.uk/adult-social-care-and-health", "020 8937 1234", "adultsocialcare@brent.gov.uk", "https://www.brent.gov.uk/your-council/contact-us"),
    'london borough of bromley': ("https://www.bromley.gov.uk/adult-social-care", "020 8464 3333", "adultcommissioning@bromley.gov.uk", "https://www.bromley.gov.uk/contact"),
    'london borough of camden': ("https://www.camden.gov.uk/adult-social-care", "020 7974 4444", "adultsocialcare@camden.gov.uk", "https://www.camden.gov.uk/contact-us"),
    'london borough of croydon': ("https://www.croydon.gov.uk/adult-social-care", "020 8726 6000", "adult.contact@croydon.gov.uk", "https://www.croydon.gov.uk/contact"),
    'london borough of ealing': ("https://www.ealing.gov.uk/adult-social-care", "020 8825 5000", "ecirs@ealing.gov.uk", "https://www.ealing.gov.uk/contact-us"),
    'london borough of enfield': ("https://www.enfield.gov.uk/services/adult-social-care", "020 8379 1000", "adult.referrals@enfield.gov.uk", "https://www.enfield.gov.uk/contact-us"),
    'royal borough of greenwich': ("https://www.royalgreenwich.gov.uk/info/200244/adult_social_care_and_health", "020 8921 2304", "asccentral@royalgreenwich.gov.uk", "https://www.royalgreenwich.gov.uk/contact"),
    'london borough of hackney': ("https://hackney.gov.uk/adult-social-care", "020 8356 5000", "adultsocialcare@hackney.gov.uk", "https://hackney.gov.uk/contact-us"),
    'london borough of hammersmith and fulham': ("https://www.lbhf.gov.uk/social-care-and-health", "020 8753 1000", "adultsocialcare@lbhf.gov.uk", "https://www.lbhf.gov.uk/contact-us"),
    'london borough of haringey': ("https://www.haringey.gov.uk/social-care-and-health", "020 8489 1400", "first.response.team@haringey.gov.uk", "https://www.haringey.gov.uk/contact-us"),
    'london borough of harrow': ("https://www.harrow.gov.uk/adult-social-care", "020 8901 2680", "access.team@harrow.gov.uk", "https://www.harrow.gov.uk/contact"),
    'london borough of havering': ("https://www.havering.gov.uk/social-care-and-health", "01708 434 000", "adult.socialcare@havering.gov.uk", "https://www.havering.gov.uk/contact-us"),
    'london borough of hillingdon': ("https://www.hillingdon.gov.uk/social-care", "01895 250 111", "socialcare@hillingdon.gov.uk", "https://www.hillingdon.gov.uk/contact-us"),
    'london borough of hounslow': ("https://www.hounslow.gov.uk/adult-social-care", "020 8583 2000", "adultsocialcare@hounslow.gov.uk", "https://www.hounslow.gov.uk/contact-us"),
    'islington council': ("https://www.islington.gov.uk/social-care-and-health", "020 7527 2000", "adult.socialservices@islington.gov.uk", "https://www.islington.gov.uk/contact-us"),
    'royal borough of kensington and chelsea': ("https://www.rbkc.gov.uk/health-and-social-care", "020 7361 3013", "adultsocialcareenquiries@rbkc.gov.uk", "https://www.rbkc.gov.uk/contact-us"),
    'the royal borough of kingston upon thames': ("https://www.kingston.gov.uk/info/200063/adult_care", "020 8547 5005", "asc.contact@kingston.gov.uk", "https://www.kingston.gov.uk/contact-us"),
    'london borough of lambeth': ("https://www.lambeth.gov.uk/lambeth-data-hub/supported-living", "020 7926 1000", "asc@lambeth.gov.uk", "https://www.lambeth.gov.uk/contact-us"),
    'london borough of lewisham': ("https://www.lewisham.gov.uk/myservices/socialcare/adults", "020 8314 6000", "adultsocialcare@lewisham.gov.uk", "https://www.lewisham.gov.uk/contact-us"),
    'london borough of merton': ("https://www.merton.gov.uk/adult-social-care", "020 8545 4218", "adult.socialcare@merton.gov.uk", "https://www.merton.gov.uk/contact-us"),
    'london borough of newham': ("https://www.newham.gov.uk/health-adult-social-care", "020 8430 2000", "asc@newham.gov.uk", "https://www.newham.gov.uk/contact"),
    'london borough of redbridge': ("https://www.redbridge.gov.uk/adult-social-care", "020 8554 5000", "adultsocialservices@redbridge.gov.uk", "https://www.redbridge.gov.uk/contact-us/"),
    'london borough of richmond upon thames': ("https://www.richmond.gov.uk/services/health_and_social_care", "020 8891 1411", "ascontact@richmond.gov.uk", "https://www.richmond.gov.uk/council/contact_us"),
    'london borough of southwark': ("https://www.southwark.gov.uk/health-and-wellbeing/adult-social-care", "020 7525 5000", "adultsocialcare@southwark.gov.uk", "https://www.southwark.gov.uk/contact-us"),
    'london borough of sutton': ("https://www.sutton.gov.uk/adult-social-care", "020 8770 5000", "adult.contact@sutton.gov.uk", "https://www.sutton.gov.uk/contact-us"),
    'tower hamlets': ("https://www.towerhamlets.gov.uk/lgnl/health__social_care/adult_social_care", "020 7364 5000", "ASC.contactteam@towerhamlets.gov.uk", "https://www.towerhamlets.gov.uk/lgnl/help_and_contact"),
    'london borough of waltham forest': ("https://www.walthamforest.gov.uk/adults-and-health", "020 8496 3000", "asc@walthamforest.gov.uk", "https://walthamforest.gov.uk/contact-us"),
    'london borough of wandsworth': ("https://www.wandsworth.gov.uk/social-care", "020 8871 6000", "adultsocialcare@wandsworth.gov.uk", "https://www.wandsworth.gov.uk/contact"),
    'westminster city council': ("https://www.westminster.gov.uk/health-and-social-care", "020 7641 6000", "adultsocialcare@westminster.gov.uk", "https://www.westminster.gov.uk/contact"),
}

# Map a Bidstats borough string (e.g. "London Borough of Hounslow") to council name + contact
def pick_council(borough_strs):
    if not borough_strs: return None, None
    for b in borough_strs:
        lb = b.lower()
        for k, contact in COUNCIL_CONTACTS.items():
            if k in lb or k.replace('london borough of ','') in lb:
                return b, contact
    # try first borough as generic
    b = borough_strs[0]
    return b, COUNCIL_CONTACTS.get(b.lower())

now = datetime.datetime.now().isoformat()

# ============ PROCESS Bidstats search JSON (17 records) ============
print("="*60)
print("PROCESSING Bidstats search JSON (17 records)")
print("="*60)
bv = json.load(open('data/scraped/raw/bidstats_london_via_search_2026-06-07.json', encoding='utf-8'))
records = bv.get('records', []) if isinstance(bv, dict) else bv

added_contracts = added_companies = 0
report = []
for rec in records:
    title = rec.get('title','')
    boroughs = rec.get('boroughs', [])
    suppliers = rec.get('suppliers', [])
    nid = rec.get('notice_id', '')
    status = rec.get('status','')

    classification = []
    if status != 'Award':
        classification.append(f'{status} (no suppliers awarded)')
    if not any(LONDON_KW.search(b) for b in boroughs) and not LONDON_KW.search(title):
        classification.append('not London')
    full = ' '.join(boroughs) + ' ' + title
    if not H_OK.search(full):
        classification.append('not housing')
    if NOT_H.search(full):
        classification.append('excluded (homecare/FM)')
    if not suppliers:
        classification.append('no suppliers parsed')

    if classification:
        report.append((nid, title[:50], 'SKIP', ', '.join(classification), 0))
        continue

    council, contact = pick_council(boroughs)
    if not council:
        report.append((nid, title[:50], 'SKIP', 'no matched council', 0))
        continue

    # Add each supplier as a contract row if not already
    n_added = 0
    for s in suppliers:
        sup = s.get('name','').strip() if isinstance(s, dict) else str(s)
        if not sup: continue
        if CARE_FACILITY.search(sup): continue
        ns = normname(sup)
        key = (ns, normname(council))
        if key in manual_keys: continue
        manual_keys.add(key)
        row = [sup, "Housing", council, 1, "", "", "", "", "", "",
               "Supported accommodation | Supported living",
               (rec.get('published','') or '')[:10],
               f"{title} (Bidstats {nid}; {status})",
               "London", "Local Council", "Local", ""]
        ws_cs.append(row)
        added_contracts += 1
        n_added += 1
        # Add to Companies if net-new
        if ns not in live_norm and ns not in manual_suppliers and sup.lower() not in existing_companies:
            manual_suppliers.add(ns)
            existing_companies.add(sup.lower())
            if contact:
                website, phone, email, contact_page = contact
                ws_co.append([sup, 1, 0, "", 1, council, 0, "", 1, 0,
                              council, "", website, phone, email, contact_page,
                              "", "", "", "", "via-commissioner"])
                added_companies += 1

    report.append((nid, title[:50], 'KEEP', f'{len(suppliers)} suppliers', n_added))

print(f"\nResults of processing 17 Bidstats search records:")
for nid, t, action, reason, n in report:
    print(f"  {action}  {nid}  [{n:2d}]  {t:50s}  | {reason}")
print(f"\nTotal contract rows added: {added_contracts}")
print(f"Total Companies entries added: {added_companies}")

# ============ PROCESS the 2 new Bidstats HTML notices ============
print("\n" + "="*60)
print("PROCESSING 2 new Bidstats HTML notices (778567851 + 800129262)")
print("="*60)
processed_sids = set()
wb_cur = openpyxl.load_workbook('data/scraped/curated_scraped.xlsx', read_only=True)
for r in wb_cur.active.iter_rows(min_row=2, values_only=True):
    if len(r) > 19 and r[19]: processed_sids.add(str(r[19]).strip())

added2 = comp2 = 0
for path in sorted(glob.glob('.scratch/bidstats/*.html')):
    nid = os.path.basename(path).replace('.html','')
    if nid in processed_sids: continue
    h = open(path, encoding='utf-8', errors='replace').read()
    text = re.sub(r'<script.*?</script>', '', h, flags=re.S)
    text = re.sub(r'<style.*?</style>', '', text, flags=re.S)
    text = html.unescape(re.sub(r'<[^>]+>', '\n', text))
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n+', '\n', text)
    if '[Award]' not in text: continue
    if not LONDON_KW.search(text): continue
    if not H_OK.search(text[:3000]): continue
    if NOT_H.search(text[:3000]): continue
    tm = re.search(r'\n\s*([A-Z][^\n]{15,200}?)\s*\[Award\]', text)
    title = tm.group(1).strip() if tm else ""
    # boroughs from text
    boroughs = []
    BR = re.compile(r"(London Borough of [A-Z][a-z]+(?:[ -][A-Z][a-z]+)*|Royal Borough of [A-Za-z &]+|Tower Hamlets|City of London|Westminster City Council)")
    for m in BR.finditer(text[:5000]):
        b = m.group(1).strip()
        if b not in boroughs: boroughs.append(b)
    council, contact = pick_council(boroughs)
    if not council: continue
    # suppliers
    suppliers = []
    ai = text.lower().find('award detail')
    if ai > 0:
        section = text[ai:ai+15000]
        for m in re.finditer(r"\n\s*([A-Z][A-Za-z0-9 &'\.\,/-]{4,80}?)\s*\n\s*\(([A-Za-z ,]+?)\)", section):
            name = m.group(1).strip()
            if name.lower() in {'award detail','awards','contractors','supplier analysis','quality','price','award criteria','reference','status','history','categories','domains','indicators','social value','published','published category','unnamed'}: continue
            if name not in suppliers: suppliers.append(name)
    n_here = 0
    for sup in suppliers:
        if CARE_FACILITY.search(sup): continue
        ns = normname(sup)
        key = (ns, normname(council))
        if key in manual_keys: continue
        manual_keys.add(key)
        ws_cs.append([sup, "Housing", council, 1, "", "", "", "", "", "",
                      "Supported accommodation | Supported living", "",
                      f"{title} (Bidstats {nid}; Award)",
                      "London", "Local Council", "Local", ""])
        added2 += 1; n_here += 1
        if ns not in live_norm and ns not in manual_suppliers and sup.lower() not in existing_companies:
            manual_suppliers.add(ns); existing_companies.add(sup.lower())
            if contact:
                website, phone, email, contact_page = contact
                ws_co.append([sup, 1, 0, "", 1, council, 0, "", 1, 0, council, "",
                              website, phone, email, contact_page,
                              "", "", "", "", "via-commissioner"])
                comp2 += 1
    print(f"  +{n_here} from {nid} ({title[:50]}) → {council}")

wb_m.save(MANUAL)
print(f"\nFrom 2 cached HTML notices: {added2} contract rows, {comp2} new Companies entries")
print(f"\nGRAND TOTAL THIS RUN:")
print(f"  Contract rows added : {added_contracts + added2}")
print(f"  Companies added     : {added_companies + comp2}")
print(f"  Manual file totals  : {ws_cs.max_row-1} contracts, {ws_co.max_row-1} Companies")
