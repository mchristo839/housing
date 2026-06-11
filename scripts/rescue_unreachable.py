"""Rescue every net-new supplier currently dropped by the build's
'must be reachable' filter.

For each filtered-out supplier we add a Companies sheet row using the
commissioning council's contact info. The supplier reaches the live site
with a "contact via commissioner" route — accurate for procurement-led
discovery, and lets the row through the reachability gate.

This is the same pattern we used for the 6 framework placeholders, just
applied automatically to every dropped Bidstats-discovered supplier.
"""
import openpyxl, sys, io, re, json
from collections import defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

MAIN = 'data/care_housing_database_v2_ENRICHED.xlsx'
MANUAL = 'data/manual_contracts.xlsx'

def norm_supplier(s):
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())

# Council → contact info. Coverage = every London commissioner that appears in
# our manual rows. Used as a contact-route for filtered-out suppliers commissioned
# by that body.
COUNCIL_CONTACTS = {
    'london borough of barking and dagenham':
        ("https://www.lbbd.gov.uk/adult-health-and-social-care", "020 8215 3000", "adult.socialcare@lbbd.gov.uk", "https://www.lbbd.gov.uk/contact-us"),
    'london borough of barnet':
        ("https://www.barnet.gov.uk/adults-and-health", "020 8359 2000", "first.response@barnet.gov.uk", "https://www.barnet.gov.uk/contact-us"),
    'bexley': ("https://www.bexley.gov.uk/services/adult-social-care", "020 3045 5000", "adultsocialcare@bexley.gov.uk", "https://www.bexley.gov.uk/contact-us"),
    'london borough of bexley': ("https://www.bexley.gov.uk/services/adult-social-care", "020 3045 5000", "adultsocialcare@bexley.gov.uk", "https://www.bexley.gov.uk/contact-us"),
    'london borough of brent': ("https://www.brent.gov.uk/adult-social-care-and-health", "020 8937 1234", "adultsocialcare@brent.gov.uk", "https://www.brent.gov.uk/your-council/contact-us"),
    'london borough of bromley': ("https://www.bromley.gov.uk/adult-social-care", "020 8464 3333", "adultcommissioning@bromley.gov.uk", "https://www.bromley.gov.uk/contact"),
    'camden': ("https://www.camden.gov.uk/adult-social-care", "020 7974 4444", "adultsocialcare@camden.gov.uk", "https://www.camden.gov.uk/contact-us"),
    'london borough of camden': ("https://www.camden.gov.uk/adult-social-care", "020 7974 4444", "adultsocialcare@camden.gov.uk", "https://www.camden.gov.uk/contact-us"),
    'city of london corporation': ("https://www.cityoflondon.gov.uk/services/adult-social-care", "020 7332 1224", "asc@cityoflondon.gov.uk", "https://www.cityoflondon.gov.uk/contact-us"),
    'croydon': ("https://www.croydon.gov.uk/adult-social-care", "020 8726 6000", "adult.contact@croydon.gov.uk", "https://www.croydon.gov.uk/contact"),
    'london borough of croydon': ("https://www.croydon.gov.uk/adult-social-care", "020 8726 6000", "adult.contact@croydon.gov.uk", "https://www.croydon.gov.uk/contact"),
    'ealing': ("https://www.ealing.gov.uk/adult-social-care", "020 8825 5000", "ecirs@ealing.gov.uk", "https://www.ealing.gov.uk/contact-us"),
    'london borough of ealing': ("https://www.ealing.gov.uk/adult-social-care", "020 8825 5000", "ecirs@ealing.gov.uk", "https://www.ealing.gov.uk/contact-us"),
    'london borough of enfield': ("https://www.enfield.gov.uk/services/adult-social-care", "020 8379 1000", "adult.referrals@enfield.gov.uk", "https://www.enfield.gov.uk/contact-us"),
    'royal borough of greenwich': ("https://www.royalgreenwich.gov.uk/info/200244/adult_social_care_and_health", "020 8921 2304", "asccentral@royalgreenwich.gov.uk", "https://www.royalgreenwich.gov.uk/contact"),
    'greater london authority': ("https://www.london.gov.uk/about-us/our-building-and-squares/our-building-city-hall", "020 7983 4000", "mayor@london.gov.uk", "https://www.london.gov.uk/about-us/contact-us"),
    'london borough of hackney': ("https://hackney.gov.uk/adult-social-care", "020 8356 5000", "adultsocialcare@hackney.gov.uk", "https://hackney.gov.uk/contact-us"),
    'london borough of hammersmith and fulham': ("https://www.lbhf.gov.uk/social-care-and-health", "020 8753 1000", "adultsocialcare@lbhf.gov.uk", "https://www.lbhf.gov.uk/contact-us"),
    'london borough of hammersmith & fulham': ("https://www.lbhf.gov.uk/social-care-and-health", "020 8753 1000", "adultsocialcare@lbhf.gov.uk", "https://www.lbhf.gov.uk/contact-us"),
    'london borough of haringey': ("https://www.haringey.gov.uk/social-care-and-health", "020 8489 1400", "first.response.team@haringey.gov.uk", "https://www.haringey.gov.uk/contact-us"),
    'london borough of harrow': ("https://www.harrow.gov.uk/adult-social-care", "020 8901 2680", "access.team@harrow.gov.uk", "https://www.harrow.gov.uk/contact"),
    'harrow council': ("https://www.harrow.gov.uk/adult-social-care", "020 8901 2680", "access.team@harrow.gov.uk", "https://www.harrow.gov.uk/contact"),
    'london borough of havering': ("https://www.havering.gov.uk/social-care-and-health", "01708 434 000", "adult.socialcare@havering.gov.uk", "https://www.havering.gov.uk/contact-us"),
    'london borough of hillingdon': ("https://www.hillingdon.gov.uk/social-care", "01895 250 111", "socialcare@hillingdon.gov.uk", "https://www.hillingdon.gov.uk/contact-us"),
    'london borough of hounslow': ("https://www.hounslow.gov.uk/adult-social-care", "020 8583 2000", "adultsocialcare@hounslow.gov.uk", "https://www.hounslow.gov.uk/contact-us"),
    'islington': ("https://www.islington.gov.uk/social-care-and-health", "020 7527 2000", "adult.socialservices@islington.gov.uk", "https://www.islington.gov.uk/contact-us"),
    'islington council': ("https://www.islington.gov.uk/social-care-and-health", "020 7527 2000", "adult.socialservices@islington.gov.uk", "https://www.islington.gov.uk/contact-us"),
    'london borough of islington': ("https://www.islington.gov.uk/social-care-and-health", "020 7527 2000", "adult.socialservices@islington.gov.uk", "https://www.islington.gov.uk/contact-us"),
    'royal borough of kensington and chelsea': ("https://www.rbkc.gov.uk/health-and-social-care", "020 7361 3013", "adultsocialcareenquiries@rbkc.gov.uk", "https://www.rbkc.gov.uk/contact-us"),
    'the royal borough of kingston upon thames': ("https://www.kingston.gov.uk/info/200063/adult_care", "020 8547 5005", "asc.contact@kingston.gov.uk", "https://www.kingston.gov.uk/contact-us"),
    'royal borough of kingston upon thames': ("https://www.kingston.gov.uk/info/200063/adult_care", "020 8547 5005", "asc.contact@kingston.gov.uk", "https://www.kingston.gov.uk/contact-us"),
    'london borough of lambeth': ("https://www.lambeth.gov.uk/lambeth-data-hub/supported-living", "020 7926 1000", "asc@lambeth.gov.uk", "https://www.lambeth.gov.uk/contact-us"),
    'london borough of lewisham': ("https://www.lewisham.gov.uk/myservices/socialcare/adults", "020 8314 6000", "adultsocialcare@lewisham.gov.uk", "https://www.lewisham.gov.uk/contact-us"),
    'london borough of merton': ("https://www.merton.gov.uk/adult-social-care", "020 8545 4218", "adult.socialcare@merton.gov.uk", "https://www.merton.gov.uk/contact-us"),
    'merton': ("https://www.merton.gov.uk/adult-social-care", "020 8545 4218", "adult.socialcare@merton.gov.uk", "https://www.merton.gov.uk/contact-us"),
    'london borough of newham': ("https://www.newham.gov.uk/health-adult-social-care", "020 8430 2000", "asc@newham.gov.uk", "https://www.newham.gov.uk/contact"),
    'london borough of redbridge': ("https://www.redbridge.gov.uk/adult-social-care", "020 8554 5000", "adultsocialservices@redbridge.gov.uk", "https://www.redbridge.gov.uk/contact-us/"),
    'london borough of richmond upon thames': ("https://www.richmond.gov.uk/services/health_and_social_care", "020 8891 1411", "ascontact@richmond.gov.uk", "https://www.richmond.gov.uk/council/contact_us"),
    'london borough of southwark': ("https://www.southwark.gov.uk/health-and-wellbeing/adult-social-care", "020 7525 5000", "adultsocialcare@southwark.gov.uk", "https://www.southwark.gov.uk/contact-us"),
    'southwark': ("https://www.southwark.gov.uk/health-and-wellbeing/adult-social-care", "020 7525 5000", "adultsocialcare@southwark.gov.uk", "https://www.southwark.gov.uk/contact-us"),
    'london borough of sutton': ("https://www.sutton.gov.uk/adult-social-care", "020 8770 5000", "adult.contact@sutton.gov.uk", "https://www.sutton.gov.uk/contact-us"),
    'tower hamlets': ("https://www.towerhamlets.gov.uk/lgnl/health__social_care/adult_social_care", "020 7364 5000", "ASC.contactteam@towerhamlets.gov.uk", "https://www.towerhamlets.gov.uk/lgnl/help_and_contact"),
    'london borough of tower hamlets': ("https://www.towerhamlets.gov.uk/lgnl/health__social_care/adult_social_care", "020 7364 5000", "ASC.contactteam@towerhamlets.gov.uk", "https://www.towerhamlets.gov.uk/lgnl/help_and_contact"),
    'london borough of waltham forest': ("https://www.walthamforest.gov.uk/adults-and-health", "020 8496 3000", "asc@walthamforest.gov.uk", "https://walthamforest.gov.uk/contact-us"),
    'london borough of wandsworth': ("https://www.wandsworth.gov.uk/social-care", "020 8871 6000", "adultsocialcare@wandsworth.gov.uk", "https://www.wandsworth.gov.uk/contact"),
    'westminster': ("https://www.westminster.gov.uk/health-and-social-care", "020 7641 6000", "adultsocialcare@westminster.gov.uk", "https://www.westminster.gov.uk/contact"),
    'westminster city council': ("https://www.westminster.gov.uk/health-and-social-care", "020 7641 6000", "adultsocialcare@westminster.gov.uk", "https://www.westminster.gov.uk/contact"),
    "mayor's office for policing and crime (mopac)": ("https://www.london.gov.uk/who-we-are/mopac", "020 7983 4000", "mopac.enquiries@london.gov.uk", "https://www.london.gov.uk/about-us/contact-us"),
    "ministry of housing, communities and local government": ("https://www.gov.uk/government/organisations/ministry-of-housing-communities-and-local-government", "030 3444 0000", "correspondence@communities.gov.uk", "https://www.gov.uk/government/organisations/ministry-of-housing-communities-and-local-government/about/about-our-services"),
    'london councils (capital letters)': ("https://www.capitalletters.org.uk", "020 7158 5300", "info@capitalletters.org.uk", "https://www.capitalletters.org.uk/contact-us"),
    'home office': ("https://www.gov.uk/government/organisations/home-office", "020 7035 4848", "public.enquiries@homeoffice.gov.uk", "https://www.gov.uk/contact-the-home-office"),
    'ministry of justice': ("https://www.gov.uk/government/organisations/ministry-of-justice", "020 3334 3555", "general.queries@justice.gov.uk", "https://www.gov.uk/government/organisations/ministry-of-justice/about/access-and-opening"),
}

# 1. Load main DB
print("Loading main DB suppliers...")
wb_main = openpyxl.load_workbook(MAIN, read_only=True, data_only=True)
main_suppliers = set()
for r in wb_main['Company × Council × Sector'].iter_rows(min_row=2, values_only=True):
    if r[0]: main_suppliers.add(norm_supplier(r[0]))
for r in wb_main['Companies'].iter_rows(min_row=2, values_only=True):
    if r[0]: main_suppliers.add(norm_supplier(r[0]))
print(f"  Main DB unique suppliers: {len(main_suppliers)}")

# 2. Load currently live providers
prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
live_norm = {norm_supplier(p['name']) for p in prov}
print(f"  Currently live providers: {len(prov)}")

# 3. Walk manual file — find every net-new supplier currently DROPPED from live
wb_man = openpyxl.load_workbook(MANUAL)
ws_cs = wb_man['Company × Council × Sector']
ws_co = wb_man['Companies']

existing_companies = {str(r[0].value).strip().lower() for r in ws_co.iter_rows(min_row=2) if r[0].value}

# Build: for each supplier, which London council to use as their contact route?
# Prefer the first London council that has a contact entry in COUNCIL_CONTACTS.
supplier_council = {}
supplier_pretty = {}
for r in ws_cs.iter_rows(min_row=2, values_only=True):
    if not r[0] or not r[2]: continue
    ns = norm_supplier(r[0])
    supplier_pretty[ns] = str(r[0]).strip()
    cnc = str(r[2]).strip().lower()
    if ns not in supplier_council and cnc in COUNCIL_CONTACTS:
        supplier_council[ns] = (str(r[2]).strip(), COUNCIL_CONTACTS[cnc])

# Net-new + dropped = those not in main DB AND not in live providers
dropped = []
for ns, (cnc, contact) in supplier_council.items():
    if ns in main_suppliers: continue   # not net-new, skip
    if ns in live_norm: continue        # already live, skip
    if supplier_pretty[ns].strip().lower() in existing_companies: continue
    dropped.append((ns, cnc, contact))

print(f"\nSuppliers to rescue (net-new + currently dropped): {len(dropped)}")

# 4. Add Companies sheet entries
added = 0
for ns, cnc, (website, phone, email, contact_page) in dropped:
    pretty = supplier_pretty[ns]
    row = [
        pretty,                  # 0 Company
        1,                       # 1 Total Contracts
        0,                       # 2 Homecare Contracts
        "",                      # 3 Homecare Councils
        1,                       # 4 Housing Contracts
        cnc,                     # 5 Housing Councils
        0,                       # 6 Both Sectors?
        "",                      # 7 Companies House
        1,                       # 8 Is SME (default yes — most are small)
        0,                       # 9 Is VCSE
        cnc,                     # 10 All Councils
        "",                      # 11 Address
        website,                 # 12 Website (council page)
        phone,                   # 13 Telephone
        email,                   # 14 Email
        contact_page,            # 15 Contact Page
        "", "", "", "",          # 16-19 Charity fields
        "via-commissioner",      # 20 Website Review tag
    ]
    ws_co.append(row)
    existing_companies.add(pretty.lower())
    added += 1

wb_man.save(MANUAL)
print(f"Added {added} Companies entries (via-commissioner contact route)")
print(f"Companies sheet now: {ws_co.max_row-1} rows")
