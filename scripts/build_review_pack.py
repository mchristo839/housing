"""
Build the human-review pack from data/scraped/curated_scraped.xlsx.

Produces TWO outputs intended for the user to work through:

  1. data/scraped/REVIEW_PACK.xlsx
       One row per (supplier x council x contract) with a leading
       "Approve?" column for the user to mark Y / N / ? as they go.
       Rows are sorted by borough then supplier so the user can review
       a borough at a time. Source URL is hyperlinked.

  2. data/scraped/REVIEW_PACK.md
       Markdown summary grouped by borough showing:
         - Borough name + total proposed providers
         - Bulleted list of (supplier - contract title - value - source)
       For quick scan-reading in a browser.

The review pack is NEVER read by build_data.py. It's purely a staging
view of curated_scraped.xlsx. When the user has marked rows Approve=Y
in REVIEW_PACK.xlsx, run scripts/promote_approved.py (next step) to
copy those rows into data/manual/manual_contracts.xlsx, which IS read
by build_data.py and goes live on the next build/deploy.
"""
import openpyxl
import sys
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import defaultdict, Counter
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC = 'data/scraped/curated_scraped.xlsx'
OUT_XLSX = 'data/scraped/REVIEW_PACK.xlsx'
OUT_MD = 'data/scraped/REVIEW_PACK.md'

# ------------------------------------------------------------------ #
# Load source
# ------------------------------------------------------------------ #
wb_src = openpyxl.load_workbook(SRC, read_only=True)
ws_src = wb_src.active
rows = list(ws_src.iter_rows(min_row=2, values_only=True))
print(f"Loaded {len(rows)} rows from {SRC}")

# Source column index (matches curated_scraped.xlsx layout)
C_COMPANY      = 0
C_SECTOR       = 1
C_COUNCIL      = 2
C_CATEGORIES   = 10
C_AWARD        = 11
C_TITLE        = 12
C_REGION       = 13
C_COMM_TYPE    = 14
C_SCOPE        = 15
C_PORTAL       = 17
C_URL          = 18
C_SOURCE_ID    = 19
C_SCRAPED_AT   = 20
C_VALUE        = 22
C_APPROVED_AT  = 23


# ------------------------------------------------------------------ #
# Borough classification (so we can sort/group)
# ------------------------------------------------------------------ #
LONDON = [
    'Barking and Dagenham', 'Barnet', 'Bexley', 'Brent', 'Bromley',
    'Camden', 'City of London', 'Croydon', 'Ealing', 'Enfield',
    'Greenwich', 'Hackney', 'Hammersmith and Fulham', 'Haringey',
    'Harrow', 'Havering', 'Hillingdon', 'Hounslow', 'Islington',
    'Kensington and Chelsea', 'Kingston upon Thames', 'Lambeth',
    'Lewisham', 'Merton', 'Newham', 'Redbridge', 'Richmond upon Thames',
    'Southwark', 'Sutton', 'Tower Hamlets', 'Waltham Forest',
    'Wandsworth', 'Westminster',
]

PAN_LONDON_BODIES = {
    'capital letters': 'Pan-London (Capital Letters)',
    'greater london authority': 'Pan-London (GLA)',
    'mopac': 'Pan-London (MOPAC)',
    "mayor's office for policing": 'Pan-London (MOPAC)',
    'mayor.s office for policing': 'Pan-London (MOPAC)',
    'ministry of housing': 'Pan-London (MHCLG)',
    'west london alliance': 'Pan-London (WLA)',
    'london councils': 'Pan-London (London Councils)',
    'commissioning alliance': 'Consortium (Commissioning Alliance)',
    'liia': 'Consortium (LIIA)',
    'achieving for children': 'Consortium (Achieving for Children)',
}

BOROUGH_KEYS = {
    'barking': 'Barking and Dagenham', 'dagenham': 'Barking and Dagenham',
    'barnet': 'Barnet', 'bexley': 'Bexley', 'brent': 'Brent',
    'bromley': 'Bromley', 'camden': 'Camden',
    'city of london': 'City of London', 'croydon': 'Croydon',
    'ealing': 'Ealing', 'enfield': 'Enfield', 'greenwich': 'Greenwich',
    'hackney': 'Hackney',
    'hammersmith': 'Hammersmith and Fulham', 'fulham': 'Hammersmith and Fulham',
    'haringey': 'Haringey', 'harrow': 'Harrow', 'havering': 'Havering',
    'hillingdon': 'Hillingdon', 'hounslow': 'Hounslow',
    'islington': 'Islington',
    'kensington': 'Kensington and Chelsea', 'rbkc': 'Kensington and Chelsea',
    'kingston': 'Kingston upon Thames', 'lambeth': 'Lambeth',
    'lewisham': 'Lewisham', 'merton': 'Merton', 'newham': 'Newham',
    'redbridge': 'Redbridge',
    'richmond': 'Richmond upon Thames',
    'southwark': 'Southwark', 'sutton': 'Sutton',
    'tower hamlets': 'Tower Hamlets',
    'waltham': 'Waltham Forest', 'wandsworth': 'Wandsworth',
    'westminster': 'Westminster', 'wcc': 'Westminster',
}


def classify(council_raw):
    if not council_raw:
        return ('?', 'Unknown', False)
    cl = str(council_raw).lower()
    # Pan-London / consortium first (they take precedence — they fan out)
    for k, label in PAN_LONDON_BODIES.items():
        if k in cl:
            return ('Z', label, True)
    # Then individual borough
    for k, std in BOROUGH_KEYS.items():
        if k in cl:
            return (std, std, False)
    return ('Other', council_raw, False)


# ------------------------------------------------------------------ #
# Enrich rows
# ------------------------------------------------------------------ #
enriched = []
for r in rows:
    if not r[C_COMPANY] or not r[C_COUNCIL]:
        continue
    sort_key, group, is_pan = classify(r[C_COUNCIL])
    enriched.append({
        'sort_key': sort_key,
        'group': group,
        'is_pan_london': is_pan,
        'company': str(r[C_COMPANY]).strip(),
        'council': str(r[C_COUNCIL]).strip(),
        'sector': r[C_SECTOR] or '',
        'categories': r[C_CATEGORIES] or '',
        'award_date': r[C_AWARD] or '',
        'title': r[C_TITLE] or '',
        'scope': r[C_SCOPE] or '',
        'portal': r[C_PORTAL] or '',
        'url': r[C_URL] or '',
        'source_id': r[C_SOURCE_ID] or '',
        'value': r[C_VALUE] or '',
        'scraped_at': r[C_SCRAPED_AT] or '',
        'approved_at': r[C_APPROVED_AT] or '',
    })

# Sort: London boroughs A-Z, then pan-London/consortia, then other
def sort_tuple(e):
    pri = 0 if e['sort_key'] in LONDON else (1 if e['sort_key'] == 'Z' else 2)
    return (pri, e['group'], e['company'].lower())

enriched.sort(key=sort_tuple)

# ------------------------------------------------------------------ #
# Build REVIEW_PACK.xlsx
# ------------------------------------------------------------------ #
wb = openpyxl.Workbook()

# === Sheet 1: Review (main worksheet) ===
ws = wb.active
ws.title = 'Review'

headers = [
    'Approve? (Y/N/?)',
    'Quality flag',
    'Notes',
    'Borough / Body',
    'Supplier',
    'Contract Title',
    'Categories',
    'Award Date',
    'Value (GBP)',
    'Scope',
    'Source Portal',
    'Source URL',
    'Source ID',
    'Council (raw)',
    'Scraped At',
]

# Header row styling
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F4E79')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin = Side(border_style='thin', color='B4B4B4')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border
ws.row_dimensions[1].height = 32

# Alternating fill for borough groups
GROUP_FILLS = [
    PatternFill('solid', fgColor='F2F2F2'),
    PatternFill('solid', fgColor='FFFFFF'),
]
PAN_FILL = PatternFill('solid', fgColor='FFF2CC')  # yellow tint for pan-London

current_group = None
fill_idx = 0
for row_i, e in enumerate(enriched, 2):
    if e['group'] != current_group:
        current_group = e['group']
        fill_idx = 1 - fill_idx
    fill = PAN_FILL if e['is_pan_london'] else GROUP_FILLS[fill_idx]

    # Format value
    val = e['value']
    try:
        val_n = float(str(val).replace(',', '').strip()) if val not in ('', None) else None
        val_str = f"£{val_n:,.0f}" if val_n else ''
    except Exception:
        val_str = str(val) if val else ''

    # Quality flag — auto-detect known classification issues
    flag = ''
    title_lc = (e['title'] or '').lower()
    company_lc = e['company'].lower()
    if ('wnc' in title_lc or 'west northamptonshire' in title_lc) and e['sort_key'] != 'Westminster':
        flag = 'MISCLASSIFIED — actually West Northamptonshire (not London)'
    elif 'london borough of' in company_lc or company_lc.startswith('london borough'):
        flag = 'FRAMEWORK — supplier name = buyer; suppliers TBD via mini-comp'
    elif 'suppliers confidential' in (e['title'] or '').lower() or 'exempt appendix' in (e['title'] or '').lower():
        flag = 'FRAMEWORK — suppliers exempt/confidential in source doc'
    elif e['company'].strip().lower() in {'tbc', 'tbd', 'various', 'multiple'}:
        flag = 'PLACEHOLDER — no named supplier'

    cells = [
        '',                              # Approve?
        flag,                            # Quality flag (auto)
        '',                              # Notes
        e['group'],
        e['company'],
        e['title'],
        e['categories'],
        e['award_date'],
        val_str,
        e['scope'],
        e['portal'],
        e['url'],
        e['source_id'],
        e['council'],
        e['scraped_at'],
    ]
    for col_i, val in enumerate(cells, 1):
        c = ws.cell(row=row_i, column=col_i, value=val)
        c.fill = fill
        c.alignment = left_wrap
        c.border = border
        if col_i == 11 and val:           # URL → hyperlink
            try:
                c.hyperlink = str(val)
                c.font = Font(color='0563C1', underline='single')
            except Exception:
                pass

# Freeze header + Approve column
ws.freeze_panes = 'D2'

# Column widths tuned for readability
widths = [14, 38, 28, 28, 32, 50, 30, 12, 14, 18, 18, 50, 14, 32, 20]
for col_i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col_i)].width = w

# Data validation for Approve? (Y / N / ?)
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type='list', formula1='"Y,N,?,YES,NO,MAYBE"', allow_blank=True)
dv.error = 'Use Y, N or ? to mark this row.'
dv.errorTitle = 'Invalid value'
dv.prompt = 'Y = approve, N = reject, ? = come back to it'
dv.promptTitle = 'Approval'
ws.add_data_validation(dv)
dv.add(f'A2:A{len(enriched)+1}')


# === Sheet 2: Borough summary ===
ws2 = wb.create_sheet('Per-Borough Summary')

# count proposed direct adds per borough + pan-London count
direct = defaultdict(set)
pan_suppliers = set()
for e in enriched:
    if e['is_pan_london']:
        pan_suppliers.add(e['company'].lower())
    elif e['sort_key'] in LONDON:
        direct[e['sort_key']].add(e['company'].lower())

sum_headers = [
    'Borough',
    'Live count (current)',
    'Proposed direct adds',
    'Pan-London fan-out',
    'Projected after promotion',
]
for col_i, h in enumerate(sum_headers, 1):
    c = ws2.cell(row=1, column=col_i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border
ws2.row_dimensions[1].height = 30

# load live db for current counts
import json
db = json.load(open('api/_data/db.json', encoding='utf-8'))


def norm(s):
    s = re.sub(r'[^a-z0-9 ]', ' ', str(s or '').lower().replace('&', ' and '))
    drop = {'council','borough','county','city','district','the','of','metropolitan','unitary','authority','corporation','mbc','lbc','mdc','cc'}
    return ' '.join(t for t in s.split() if t and t not in drop)


def live_count(borough):
    nb = norm(borough)
    found = set()
    for k, ids in db['c'].items():
        if norm(k) == nb:
            found.update(ids)
    return len(found)


for row_i, b in enumerate(LONDON, 2):
    cur = live_count(b)
    add = len(direct[b])
    pan = len(pan_suppliers)
    proj = cur + add + pan
    for col_i, val in enumerate([b, cur, add, pan, proj], 1):
        c = ws2.cell(row=row_i, column=col_i, value=val)
        c.alignment = left_wrap
        c.border = border
        if col_i == 5:
            c.font = Font(bold=True)
ws2.column_dimensions['A'].width = 28
for col in 'BCDE':
    ws2.column_dimensions[col].width = 22
ws2.freeze_panes = 'A2'


# === Sheet 3: How to review ===
ws3 = wb.create_sheet('How to review')
instructions = [
    ('REVIEW PACK — INSTRUCTIONS', None),
    ('', None),
    ('1. Open the "Review" tab.', None),
    ('2. For each row, set column A (Approve?) to:', None),
    ('     Y = include this supplier in the live site', None),
    ('     N = reject (out of scope, wrong sector, duplicate, etc.)', None),
    ('     ? = come back to it (paul to check source)', None),
    ('3. Use column B (Notes) for any reasoning.', None),
    ('4. Hover over column K (Source URL) and click to verify against the original notice / PDF.', None),
    ('5. Save the file. Do NOT rename it.', None),
    ('6. When you have at least some Y rows, ask Claude to "promote approved rows".', None),
    ('     -> That copies them into data/manual/manual_contracts.xlsx', None),
    ('     -> Which is what build_data.py reads on the next build', None),
    ('     -> Which is what then ships on the next Vercel deploy', None),
    ('', None),
    ('IMPORTANT: this file is the REVIEW BUFFER, never read by the live site.', None),
    ('The live site reads only:', None),
    ('   - main provider source (Excel)', None),
    ('   - data/manual/manual_contracts.xlsx (curated, signed-off additions)', None),
    ('', None),
    ('Yellow rows are PAN-LONDON / CONSORTIUM contracts. Approving one of these', None),
    ('fans the supplier out to every member borough automatically via CONSORTIA in', None),
    ('build_data.py — so a Y on a Capital Letters row reaches 27 boroughs.', None),
    ('', None),
    ('Row count breakdowns are on the "Per-Borough Summary" tab.', None),
]
for row_i, (text, _) in enumerate(instructions, 1):
    c = ws3.cell(row=row_i, column=1, value=text)
    if row_i == 1:
        c.font = Font(bold=True, size=14)
ws3.column_dimensions['A'].width = 110

# Order: Review first
wb.active = wb['Review']
wb.save(OUT_XLSX)
print(f"Wrote {OUT_XLSX} — {len(enriched)} rows for review")


# ------------------------------------------------------------------ #
# Build REVIEW_PACK.md
# ------------------------------------------------------------------ #
lines = []
lines.append('# Review pack — supported-housing contracts (London)\n')
lines.append(f'_{len(enriched)} proposed rows from `data/scraped/curated_scraped.xlsx`._\n')
lines.append('Open `REVIEW_PACK.xlsx` to mark Approve? = Y / N / ? per row.\n')
lines.append('When done, ask Claude to **promote approved rows** to `manual_contracts.xlsx`.\n\n')
lines.append('---\n\n')

# Group by borough/body
grouped = defaultdict(list)
for e in enriched:
    grouped[e['group']].append(e)

# London boroughs first, alphabetically
ordered_groups = [b for b in LONDON if b in grouped]
ordered_groups += sorted(g for g in grouped if g not in LONDON)

for g in ordered_groups:
    items = grouped[g]
    tag = ' (PAN-LONDON / fans out to all London boroughs)' if items[0]['is_pan_london'] else ''
    lines.append(f'## {g}{tag} — {len(items)} proposed\n\n')
    for e in items:
        val = ''
        try:
            v = float(str(e['value']).replace(',', '').strip())
            if v >= 1_000_000:
                val = f' · £{v/1_000_000:.1f}M'
            elif v > 0:
                val = f' · £{int(v):,}'
        except Exception:
            pass
        title = (e['title'] or '').replace('\n', ' ').strip()
        if len(title) > 180:
            title = title[:177] + '...'
        url = e['url'] or ''
        src = f' · [source]({url})' if url else ''
        lines.append(f'- **{e["company"]}** — {title}{val}{src}\n')
    lines.append('\n')

with open(OUT_MD, 'w', encoding='utf-8') as f:
    f.writelines(lines)
print(f"Wrote {OUT_MD}")


# ------------------------------------------------------------------ #
# Final summary
# ------------------------------------------------------------------ #
group_counts = Counter(e['group'] for e in enriched)
print(f"\n--- Review pack ready ---")
print(f"Rows for review : {len(enriched)}")
print(f"Distinct groups : {len(group_counts)}")
print(f"Pan-London rows : {sum(1 for e in enriched if e['is_pan_london'])}")
print(f"\nTop groups by proposed-row count:")
for g, n in group_counts.most_common(15):
    print(f"  {n:4d}  {g}")
