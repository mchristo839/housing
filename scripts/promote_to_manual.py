"""
Promote human-approved scraped rows to a SEPARATE curated-scrape database.

This script intentionally does NOT modify:
    data/manual_contracts.xlsx       (hand-verified curated CAS / AASC rows)
    data/care_housing_database_v2_ENRICHED.xlsx (the main source workbook)

Approved rows instead land in:
    data/scraped/curated_scraped.xlsx   (separate, NOT loaded by build_data.py)

This keeps three tiers of data clearly separated:
  1. Main source         — refreshed monthly, untouched
  2. manual_contracts    — hand-verified named contracts (CAS-2/3, AASC, etc.)
  3. curated_scraped     — scraper findings the human approved, kept separate
                           for review/reference but NOT served to users.

If you later decide a curated_scraped row is high-confidence enough to surface
on the live site, COPY it manually into manual_contracts.xlsx.

Workflow:
  1. Open data/scraped/audit/new_providers.csv (or new_contracts.csv) in Excel.
  2. Set the `scrape_action` column to one of:
        APPROVE   — copy this row into curated_scraped.xlsx
        SKIP      — leave it alone (default if you change nothing)
        REJECT    — explicit no (kept for audit)
  3. Save the CSV.
  4. Run:  python scripts/promote_to_manual.py --from <csv-path>

Usage:
    python scripts/promote_to_manual.py --from data/scraped/audit/new_providers.csv
    python scripts/promote_to_manual.py --from data/scraped/audit/new_contracts.csv
"""
import argparse
import csv
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

import openpyxl

# ── NOTE: this is the SEPARATE curated-scrape database, not manual_contracts ──
CURATED_SCRAPED_XLSX = os.path.join(ROOT, "data", "scraped", "curated_scraped.xlsx")

MAIN_COLS = ["Company", "Sector", "Council", "Contracts (this council, this sector)",
             "Company Total (all sectors)", "Company Total — Homecare",
             "Company Total — Housing", "Companies House", "Is SME", "Is VCSE",
             "Categories", "Most Recent Award", "Contract Titles", "ONS Region",
             "Commissioner Type", "Geographic Scope", "Asylum Contractor"]

# extra columns we keep ONLY in the curated-scraped db so provenance survives
EXTRA_COLS = ["source_portal", "source_url", "source_id", "scraped_at",
              "cpv_codes", "contract_value_gbp", "approved_at"]


def load_csv(path):
    with open(path, encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--from", dest="csv_path", required=True,
                   help="Audit CSV (new_providers.csv / new_contracts.csv / ...)")
    args = p.parse_args()

    rows = load_csv(args.csv_path)
    approved = [r for r in rows
                if str(r.get("scrape_action", "")).strip().upper() == "APPROVE"]
    if not approved:
        print(f"No rows marked APPROVE in {args.csv_path}.")
        print("Edit the CSV, set scrape_action=APPROVE on rows you want, save, re-run.")
        return

    # bootstrap curated_scraped.xlsx if needed
    os.makedirs(os.path.dirname(CURATED_SCRAPED_XLSX), exist_ok=True)
    if not os.path.exists(CURATED_SCRAPED_XLSX):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Company × Council × Sector")
        ws.append(MAIN_COLS + EXTRA_COLS)
        wb.save(CURATED_SCRAPED_XLSX)

    wb = openpyxl.load_workbook(CURATED_SCRAPED_XLSX)
    ws = wb["Company × Council × Sector"]

    import datetime
    approved_at = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

    added = 0
    for r in approved:
        out = [r.get(col, "") for col in MAIN_COLS] + [
            r.get("source_portal", ""), r.get("source_url", ""),
            r.get("source_id", ""), r.get("scraped_at", ""),
            r.get("cpv_codes", ""), r.get("contract_value_gbp", ""),
            approved_at,
        ]
        ws.append(out)
        added += 1

    wb.save(CURATED_SCRAPED_XLSX)
    print(f"Appended {added} approved rows to:")
    print(f"  {CURATED_SCRAPED_XLSX}")
    print()
    print("This file is NOT loaded by the main build pipeline. It's a reference /")
    print("review database. If you decide any of these rows are high-confidence")
    print("enough to surface on the live site, copy them MANUALLY into")
    print("data/manual_contracts.xlsx.")


if __name__ == "__main__":
    main()
