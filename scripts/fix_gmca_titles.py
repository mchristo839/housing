"""Fix: contract titles containing 'Greater Manchester' get re-routed to Manchester
by classify_title, ignoring the per-borough CCS council. Strip city-name strings
from our GMCA HF titles so the council column drives the assignment.
"""
import openpyxl
PATH = 'data/manual_contracts.xlsx'
wb = openpyxl.load_workbook(PATH)
ccs = wb['Company × Council × Sector']

NEW_TITLE = ("Housing First pilot — Apr 2019, extended Mar 2029. MHCLG-funded, "
             "GMCA-led 10-organisation partnership delivered across the ten "
             "constituent boroughs in zone-based teams. Source: gmhousingfirst.org.uk.")

GMCA_FIRMS = {
    'petrus community', 'jigsaw support', 'stockport homes group',
    'manchester action on street health (mash)', 'early break',
    'community-led initiatives',
    'the riverside group', 'great places housing association',
    'humankind charity',
}

fixed = 0
for r in range(2, ccs.max_row + 1):
    co = (ccs.cell(row=r, column=1).value or '').strip().lower()
    title = (ccs.cell(row=r, column=13).value or '')
    if co in GMCA_FIRMS and 'Greater Manchester' in title:
        ccs.cell(row=r, column=13).value = NEW_TITLE
        fixed += 1

wb.save(PATH)
print(f"Rewrote {fixed} GMCA HF titles to avoid 'Greater Manchester' string-match")
