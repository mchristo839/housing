"""Cross-reference manual_contracts.xlsx against the main DB.

For every supplier in manual_contracts.xlsx, classify it as:
  - EXISTING in main DB (was already a provider, we added contracts to them)
  - NET-NEW (didn't exist in main DB before this session)

Report: how many net-new providers did we add overall, and which London
boroughs gained them.
"""
import openpyxl, sys, io, re, json
from collections import Counter, defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

MAIN = 'data/care_housing_database_v2_ENRICHED.xlsx'
MANUAL = 'data/manual_contracts.xlsx'

def norm_supplier(s):
    """Same normalisation build_data.py uses to merge suppliers."""
    s = re.sub(r'[^a-z0-9]', '', str(s or '').lower())
    return s

# 1. Load every supplier from main DB
print("Loading main DB suppliers...")
wb_main = openpyxl.load_workbook(MAIN, read_only=True, data_only=True)
main_suppliers = set()
main_company_names = []

for r in wb_main['Company × Council × Sector'].iter_rows(min_row=2, values_only=True):
    if r[0]:
        main_suppliers.add(norm_supplier(r[0]))
        main_company_names.append(str(r[0]).strip())

for r in wb_main['Companies'].iter_rows(min_row=2, values_only=True):
    if r[0]:
        main_suppliers.add(norm_supplier(r[0]))

print(f"  Main DB unique suppliers: {len(main_suppliers)}")

# 2. Load every supplier from manual_contracts.xlsx + which London council
print("\nLoading manual additions...")
wb_man = openpyxl.load_workbook(MANUAL, read_only=True, data_only=True)
ws_cs = wb_man['Company × Council × Sector']

LONDON_KEYS = ['barking','dagenham','barnet','bexley','brent','bromley','camden',
               'city of london','croydon','ealing','enfield','greenwich','hackney',
               'hammersmith','fulham','haringey','harrow','havering','hillingdon',
               'hounslow','islington','kensington','rbkc','kingston','lambeth',
               'lewisham','merton','newham','redbridge','richmond','southwark',
               'sutton','tower hamlets','waltham','wandsworth','westminster','wcc']

def london_borough_of(c):
    cl = str(c or '').lower()
    M = {'barking':'Barking and Dagenham','dagenham':'Barking and Dagenham',
         'barnet':'Barnet','bexley':'Bexley','brent':'Brent','bromley':'Bromley',
         'camden':'Camden','city of london':'City of London','croydon':'Croydon',
         'ealing':'Ealing','enfield':'Enfield','greenwich':'Greenwich',
         'hackney':'Hackney','hammersmith':'Hammersmith and Fulham',
         'fulham':'Hammersmith and Fulham','haringey':'Haringey','harrow':'Harrow',
         'havering':'Havering','hillingdon':'Hillingdon','hounslow':'Hounslow',
         'islington':'Islington','kensington':'Kensington and Chelsea',
         'rbkc':'Kensington and Chelsea','kingston':'Kingston upon Thames',
         'lambeth':'Lambeth','lewisham':'Lewisham','merton':'Merton','newham':'Newham',
         'redbridge':'Redbridge','richmond':'Richmond upon Thames',
         'southwark':'Southwark','sutton':'Sutton','tower hamlets':'Tower Hamlets',
         'waltham':'Waltham Forest','wandsworth':'Wandsworth','westminster':'Westminster',
         'wcc':'Westminster'}
    for key, std in M.items():
        if key in cl: return std
    return None

manual_suppliers = set()
supplier_to_councils = defaultdict(set)
supplier_pretty = {}

for r in ws_cs.iter_rows(min_row=2, values_only=True):
    if not r[0] or not r[2]: continue
    norm_s = norm_supplier(r[0])
    manual_suppliers.add(norm_s)
    supplier_pretty[norm_s] = str(r[0]).strip()
    b = london_borough_of(r[2])
    if b:
        supplier_to_councils[norm_s].add(b)

print(f"  Manual unique suppliers: {len(manual_suppliers)}")

# 3. Classify each
existing = manual_suppliers & main_suppliers
net_new  = manual_suppliers - main_suppliers

print(f"\n=== CROSS-REFERENCE RESULTS ===")
print(f"  Existing in main DB (added contracts only) : {len(existing)}")
print(f"  Net-new providers (added to universe)      : {len(net_new)}")

# 4. Net-new per London borough
print(f"\n=== NET-NEW PROVIDERS PER LONDON BOROUGH ===")
borough_net_new = defaultdict(set)
for ns in net_new:
    for b in supplier_to_councils[ns]:
        borough_net_new[b].add(ns)

LONDON_ORDER = ['Barking and Dagenham','Barnet','Bexley','Brent','Bromley','Camden',
                'City of London','Croydon','Ealing','Enfield','Greenwich','Hackney',
                'Hammersmith and Fulham','Haringey','Harrow','Havering','Hillingdon',
                'Hounslow','Islington','Kensington and Chelsea','Kingston upon Thames',
                'Lambeth','Lewisham','Merton','Newham','Redbridge','Richmond upon Thames',
                'Southwark','Sutton','Tower Hamlets','Waltham Forest','Wandsworth','Westminster']
for b in LONDON_ORDER:
    n = len(borough_net_new[b])
    if n > 0:
        print(f"  +{n:3d}  {b}")

# 5. Sample of each kind
print(f"\n=== SAMPLE OF NET-NEW PROVIDERS (first 30, A-Z) ===")
for ns in sorted(net_new, key=lambda x: supplier_pretty.get(x, x).lower())[:30]:
    pretty = supplier_pretty.get(ns, ns)
    cncls = ', '.join(sorted(supplier_to_councils[ns])) or '(non-London)'
    print(f"  {pretty[:55]:55s}  -> {cncls}")

print(f"\n=== SAMPLE OF EXISTING PROVIDERS WE ADDED LONDON CONTRACTS TO (first 20) ===")
for ns in sorted(existing, key=lambda x: supplier_pretty.get(x, x).lower())[:20]:
    pretty = supplier_pretty.get(ns, ns)
    cncls = ', '.join(sorted(supplier_to_councils[ns])) or '(non-London)'
    print(f"  {pretty[:45]:45s}  -> {cncls[:60]}")

# 6. Of the net-new, how many actually survived into providers.json?
print(f"\n=== HOW MANY NET-NEW MADE IT INTO LIVE providers.json ? ===")
prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
live_norm = {norm_supplier(p['name']) for p in prov}
new_live = net_new & live_norm
new_dropped = net_new - live_norm
print(f"  Net-new in providers.json (live)  : {len(new_live)}")
print(f"  Net-new dropped by filter         : {len(new_dropped)}")
if new_dropped:
    print(f"\n  Dropped suppliers (filtered out as homecare-only/non-care/unreachable):")
    for ns in sorted(new_dropped, key=lambda x: supplier_pretty.get(x, x).lower())[:15]:
        print(f"    {supplier_pretty.get(ns, ns)}")

# 7. Headline
print(f"\n{'='*55}")
print(f"  HEADLINE")
print(f"{'='*55}")
print(f"  Total UK provider universe (live now)  : {len(prov)}")
print(f"  Suppliers we touched (this session)    : {len(manual_suppliers)}")
print(f"  - of which existing in main DB         : {len(existing)} (just added London contracts)")
print(f"  - of which NET-NEW providers           : {len(net_new)}")
print(f"  - of which net-new that SURVIVED live  : {len(new_live)}")
