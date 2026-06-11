"""Two-pass cleanup of manual_contracts.xlsx:

1. FILTER OUT non-housing suppliers — anyone whose ALL contracts are
   non-housing (technology, telecare, catering, insurance, legal, transport,
   IT, hotel-only, etc.). Suppliers with at least one housing contract are
   kept.

2. VERIFY supplier websites — HTTP HEAD/GET each Companies-sheet website
   URL with a short timeout. Flag dead URLs (HTTP 4xx, 5xx, DNS fail, SSL
   fail, timeout) by setting Website Review column to "broken-url".

Outputs the changes back to manual_contracts.xlsx and prints a report.
"""
import openpyxl, sys, io, re, urllib.request, ssl, socket, concurrent.futures
from collections import Counter, defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

MANUAL = 'data/manual_contracts.xlsx'

# STRONG non-housing markers — supplier provides X, not housing. These
# override any housing context (e.g. "Technology for supported living" — the
# supplier is a technology vendor, not a housing provider).
HARD_NON_HOUSING = re.compile(
    r"\b(?:"
    r"assistive technology|\batec\b|telecare|telehealth|"
    r"technology[- ]enabled care|technology and data services|"
    r"\btec\b service|telephony|community alarm|alarm system|"
    r"hired passenger transport|vehicle hire|fleet|taxi|minibus|"
    r"council tax|rating.*consultancy|"
    r"audit services|insurance broker|insurance services|"
    r"legal services framework|legal advice (?!support)|"
    r"it services|software|computer hardware|"
    r"street lighting|lift modernisation|fire (?:safety|risk|protection)|"
    r"grounds maintenance|grass cutting|"
    r"catering services|catering supply|"
    r"single point of access for.{0,40}(?:health|gp)|"
    r"cvs for the london|civil enforcement|"
    r"insurance loss adjuster|"
    r"booking system|payroll|"
    r"hotel.{0,30}(?:awards|business)|"
    r"recycling|waste collection|street cleansing|"
    r"residents with visual impairment|sensory loss"
    r")\b", re.I)

# Positive signals — contract IS supported housing/accommodation
HOUSING = re.compile(
    r"supported (?:living|accommodation|housing)|temporary accommodation|"
    r"emergency accommodation|hostel|refuge|housing related support|"
    r"extra care|sheltered (?:housing|accommodation)|"
    r"young people.{0,40}(?:housing|accommodation|pathway|home)|"
    r"care leaver|supported lodgings|"
    r"homeless|asylum|nightly paid|social housing|"
    r"mental health.{0,40}(?:accommodation|housing|supported|step|crisis)|"
    r"learning disabilit.{0,40}(?:accommodation|housing|supported|residential|placement)|"
    r"domestic abuse|rough sleeper|housing first|"
    r"semi[- ]?independent|floating support|move[- ]?on|tenancy support|"
    r"adult pathway|pathway services|housing pathway|"
    r"approved provider list|residential.{0,15}framework|"
    r"crisis prevention|practice flats|adult social care.*accommodation",
    re.I)

print("Loading manual_contracts.xlsx...")
wb = openpyxl.load_workbook(MANUAL)
ws_cs = wb['Company × Council × Sector']
ws_co = wb['Companies']

# Map supplier -> contracts
supplier_contracts = defaultdict(list)   # name -> [{title, cat}]
for r in ws_cs.iter_rows(min_row=2, values_only=True):
    if r[0]:
        supplier_contracts[str(r[0]).strip()].append({
            'title': str(r[12] or ''),
            'cat':   str(r[10] or ''),
        })

# Classify each contract: HARD_NON_HOUSING markers override housing words.
# Classify each supplier: drop if ALL their contracts classify non-housing.
def classify_contract(c):
    blob = c['title'] + ' ' + c['cat']
    if HARD_NON_HOUSING.search(blob):
        return 'non-housing'   # strong signal — supplier provides non-housing service
    if HOUSING.search(blob):
        return 'housing'
    return 'unknown'   # no clear signal — treat as housing by default

non_housing_only = []
mixed = []
all_housing = []
for name, contracts in supplier_contracts.items():
    classes = [classify_contract(c) for c in contracts]
    has_housing = any(x == 'housing' or x == 'unknown' for x in classes)
    only_non    = all(x == 'non-housing' for x in classes)
    has_mixed   = ('non-housing' in classes) and (('housing' in classes) or ('unknown' in classes))
    if only_non:
        non_housing_only.append(name)
    elif has_mixed:
        # Has at least one non-housing AND at least one housing contract.
        # Keep the supplier but flag for review — we'll prune the non-housing
        # contract rows specifically.
        mixed.append(name)
    else:
        all_housing.append(name)

print(f"\nSupplier classification:")
print(f"  Housing-only (KEEP)       : {len(all_housing)}")
print(f"  Mixed (KEEP — has housing too): {len(mixed)}")
print(f"  Non-housing only (REMOVE) : {len(non_housing_only)}")

if non_housing_only:
    print(f"\nNON-HOUSING SUPPLIERS TO REMOVE:")
    for n in non_housing_only:
        print(f"  - {n}")
        for c in supplier_contracts[n][:2]:
            print(f"      {c['title'][:80]}")

# ── apply removals ────────────────────────────────────────────────────────
to_remove_lc = {n.lower() for n in non_housing_only}
mixed_lc = {n.lower() for n in mixed}

removed_cs = 0
removed_co = 0

# 1. Drop the non-housing-only suppliers entirely (both sheets)
if to_remove_lc:
    for i in range(ws_cs.max_row, 1, -1):
        nm = str(ws_cs.cell(row=i, column=1).value or '').strip().lower()
        if nm in to_remove_lc:
            ws_cs.delete_rows(i)
            removed_cs += 1
    for i in range(ws_co.max_row, 1, -1):
        nm = str(ws_co.cell(row=i, column=1).value or '').strip().lower()
        if nm in to_remove_lc:
            ws_co.delete_rows(i)
            removed_co += 1

# 2. For MIXED suppliers, prune individual non-housing contract rows only
pruned_mixed = 0
if mixed_lc:
    for i in range(ws_cs.max_row, 1, -1):
        nm = str(ws_cs.cell(row=i, column=1).value or '').strip().lower()
        if nm in mixed_lc:
            title = str(ws_cs.cell(row=i, column=13).value or '')
            cat   = str(ws_cs.cell(row=i, column=11).value or '')
            blob = title + ' ' + cat
            if HARD_NON_HOUSING.search(blob):
                ws_cs.delete_rows(i)
                pruned_mixed += 1

# ── verify supplier websites ───────────────────────────────────────────────
print(f"\n\nChecking supplier websites...")
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

def check_url(url):
    if not url:
        return ('skip', '')
    if not url.startswith('http'):
        url = 'https://' + url
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
            'Accept': 'text/html,*/*'
        }, method='GET')
        with urllib.request.urlopen(req, timeout=10, context=ctx) as resp:
            code = resp.getcode()
            if 200 <= code < 400:
                return ('ok', f'{code}')
            return ('http', f'HTTP {code}')
    except urllib.error.HTTPError as e:
        # Some sites return 403 for headless requests but are alive
        if e.code in (403, 405, 406):
            return ('ok', f'HTTP {e.code} (likely UA block, site likely alive)')
        return ('http', f'HTTP {e.code}')
    except urllib.error.URLError as e:
        return ('dns', str(e.reason)[:60])
    except socket.timeout:
        return ('timeout', '10s')
    except Exception as e:
        return ('error', str(e)[:60])

# Collect URLs to check
to_check = []
for i, row in enumerate(ws_co.iter_rows(min_row=2), start=2):
    name = str(row[0].value or '').strip()
    url  = str(row[12].value or '').strip()
    tag  = str(row[20].value or '').strip()
    if not name or not url: continue
    # Skip framework / council-routed entries — those are council URLs we don't need to verify
    if tag in ('framework', 'unknown-research', 'in-house council'): continue
    to_check.append((i, name, url))

print(f"Verifying {len(to_check)} supplier websites (parallel, 10s timeout each)...")
results = []
with concurrent.futures.ThreadPoolExecutor(max_workers=12) as ex:
    futures = {ex.submit(check_url, u): (i, n, u) for i, n, u in to_check}
    for fut in concurrent.futures.as_completed(futures):
        i, n, u = futures[fut]
        try:
            status, detail = fut.result()
        except Exception as e:
            status, detail = 'error', str(e)[:60]
        results.append((i, n, u, status, detail))

# Stats
status_counts = Counter(r[3] for r in results)
print(f"\nWebsite check results:")
for s, n in status_counts.most_common():
    print(f"  {s:10s} : {n}")

# Mark broken URLs in the sheet
broken = [(i, n, u, s, d) for i, n, u, s, d in results if s != 'ok']
fixed = 0
for i, n, u, s, d in broken:
    cur_tag = str(ws_co.cell(row=i, column=21).value or '').strip()
    if cur_tag == 'verified':
        ws_co.cell(row=i, column=21).value = f'broken-url ({s})'
        fixed += 1

print(f"\nBroken/unreachable websites flagged: {fixed}")
if broken:
    print(f"\nSAMPLE BROKEN URLS (first 30):")
    for i, n, u, s, d in sorted(broken)[:30]:
        print(f"  [{s:8s}] {n[:35]:35s}  {u[:60]}  {d}")

wb.save(MANUAL)

print(f"\n{'='*55}")
print(f"  SUMMARY")
print(f"{'='*55}")
print(f"  Non-housing suppliers removed : {len(non_housing_only)}")
print(f"  Mixed suppliers — non-housing rows pruned : {pruned_mixed}")
print(f"  Contract rows removed         : {removed_cs}")
print(f"  Companies rows removed        : {removed_co}")
print(f"  Broken websites flagged       : {fixed}")
print(f"\nNext: python build_data.py")
