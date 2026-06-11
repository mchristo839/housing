"""Build HTML showing every entity being added to the main database
via manual_contracts.xlsx — full contact + which London council they serve."""
import openpyxl, json, re, sys, io, html as htmllib
from collections import defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUT = 'data/scraped/ADDING_TO_DATABASE.html'

def norm(s): return re.sub(r'[^a-z0-9]', '', str(s or '').lower())

# Baseline = main DB suppliers (to flag who's net-new vs existing)
wb_main = openpyxl.load_workbook('data/care_housing_database_v2_ENRICHED.xlsx',
                                  read_only=True, data_only=True)
main_suppliers = set()
for r in wb_main['Company × Council × Sector'].iter_rows(min_row=2, values_only=True):
    if r[0]: main_suppliers.add(norm(r[0]))
for r in wb_main['Companies'].iter_rows(min_row=2, values_only=True):
    if r[0]: main_suppliers.add(norm(r[0]))

# Read every supplier from BOTH sheets — every entity being added to the live DB
wb = openpyxl.load_workbook('data/manual_contracts.xlsx', read_only=True)
ws_co = wb['Companies']
ws_cs = wb['Company × Council × Sector']

# Live providers.json — to confirm what actually shipped after build
prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
live_by_norm = {norm(p['name']): p for p in prov}

# Walk contracts sheet — every unique supplier + their councils + every contract
supplier_councils = defaultdict(set)
supplier_contracts = defaultdict(list)
for r in ws_cs.iter_rows(min_row=2, values_only=True):
    if r[0] and r[2]:
        nm = str(r[0]).strip()
        council = str(r[2]).strip()
        title = str(r[12] or '').strip()
        cats = str(r[10] or '').strip()
        award = str(r[11] or '').strip()
        supplier_councils[nm].add(council)
        # Extract Bidstats id if any
        bm = re.search(r'Bidstats\s+(\d+)', title, re.I)
        bid = bm.group(1) if bm else ''
        # Clean title — strip per-supplier annotation
        t = re.sub(r'\s*\([^)]*?bidstats[^)]*\)\s*$', '', title, flags=re.I)
        t = re.sub(r'\s*\([^)]*?supplier\s*@[^)]*\)\s*$', '', t, flags=re.I).strip()
        supplier_contracts[nm].append({
            'council': council, 'title': t, 'categories': cats,
            'award_date': award, 'bidstats_id': bid,
        })

# Walk Companies sheet — get any direct contact data we have
companies_data = {}
for r in ws_co.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    companies_data[str(r[0]).strip()] = {
        'companies_house': str(r[7] or '').strip(),
        'website': str(r[12] or '').strip(),
        'phone': str(r[13] or '').strip(),
        'email': str(r[14] or '').strip(),
        'contact_page': str(r[15] or '').strip(),
        'charity_no': str(r[16] or '').strip(),
        'review_tag': str(r[20] or '').strip(),
    }

# Build the unified entity list — everyone being added
all_names = set(supplier_councils) | set(companies_data)
entries = []
for name in all_names:
    cd = companies_data.get(name, {})
    # If no companies-sheet entry, fall back to live provider's own contact info
    nn = norm(name)
    live = live_by_norm.get(nn)
    if not cd and live:
        cd = {
            'website': live.get('website',''),
            'phone': live.get('phone',''),
            'email': live.get('email',''),
            'contact_page': live.get('contact_page',''),
            'review_tag': 'main-db-existing',
            'companies_house': '', 'charity_no': '',
        }
    entries.append({
        'name': name,
        'companies_house': cd.get('companies_house',''),
        'website': cd.get('website',''),
        'phone': cd.get('phone',''),
        'email': cd.get('email',''),
        'contact_page': cd.get('contact_page',''),
        'charity_no': cd.get('charity_no',''),
        'review_tag': cd.get('review_tag','main-db-existing'),
        'councils': sorted(supplier_councils.get(name, set())),
        'contracts': supplier_contracts.get(name, []),
        'is_net_new': nn not in main_suppliers,
        'is_live': nn in live_by_norm,
        'n_contracts': len(supplier_contracts.get(name, [])),
    })

# Sort: net-new first, then verified status, then name
def sort_key(e):
    tag_order = {'verified': 0, 'main-db-existing': 1, 'partial': 2,
                 'in-house council': 3, 'framework': 4, 'OK': 5, 'unknown-research': 6}
    return (0 if e['is_net_new'] else 1, tag_order.get(e['review_tag'], 9), e['name'].lower())

entries.sort(key=sort_key)

# Stats
total = len(entries)
net_new = sum(1 for e in entries if e['is_net_new'])
existing = total - net_new
status_counts = defaultdict(int)
for e in entries: status_counts[e['review_tag'] or '(blank)'] += 1

# Build HTML
parts = []
parts.append('<!doctype html><html><head><meta charset="utf-8">')
parts.append('<title>Entities being added to the main database</title>')
parts.append('<style>')
parts.append('body{font:14px -apple-system,Segoe UI,Roboto,sans-serif;margin:1.5em;background:#f8f9fa;color:#222;max-width:1200px}')
parts.append('h1{margin:0 0 .2em 0;font-size:22px}')
parts.append('.meta{color:#666;margin-bottom:1.5em;font-size:13px}')
parts.append('.search{position:sticky;top:0;background:#fff;padding:10px 14px;border:1px solid #ddd;border-radius:6px;margin-bottom:1em;box-shadow:0 2px 4px rgba(0,0,0,.05);z-index:5}')
parts.append('.search input{width:330px;padding:6px 10px;font-size:14px;border:1px solid #bbb;border-radius:4px}')
parts.append('.search select{padding:6px 10px;margin-left:6px;font-size:13px}')
parts.append('.counts{color:#444;margin-left:1em;font-size:13px;font-weight:600}')
parts.append('.stat-row{display:flex;gap:10px;margin-bottom:1em;flex-wrap:wrap}')
parts.append('.stat{background:#fff;padding:8px 14px;border-radius:6px;border:1px solid #ddd;font-size:13px}')
parts.append('.stat .n{font-size:18px;font-weight:700;color:#1F4E79}')
parts.append('.entry{background:#fff;border:1px solid #ddd;border-radius:8px;padding:12px 16px;margin-bottom:10px;box-shadow:0 1px 2px rgba(0,0,0,.04);display:grid;grid-template-columns:1fr 280px;gap:14px}')
parts.append('.entry.net-new{border-left:4px solid #28a745}')
parts.append('.entry.existing{border-left:4px solid #dee2e6}')
parts.append('.entry h2{margin:0 0 6px 0;font-size:15px;color:#1F4E79;font-weight:700}')
parts.append('.tag{display:inline-block;padding:2px 8px;border-radius:11px;font-size:11px;font-weight:600;margin-left:5px;vertical-align:middle}')
parts.append('.tag.netnew{background:#d4edda;color:#1e6033}')
parts.append('.tag.existing{background:#e9ecef;color:#555}')
parts.append('.tag.verified{background:#d1ecf1;color:#0c5460}')
parts.append('.tag.framework{background:#fff3cd;color:#856404}')
parts.append('.tag.unknown{background:#f8d7da;color:#721c24}')
parts.append('.tag.partial{background:#ffeeba;color:#856404}')
parts.append('.tag.inhouse{background:#cce5ff;color:#004085}')
parts.append('.tag.ok{background:#e2e3e5;color:#383d41}')
parts.append('.contact{font-size:12px;color:#444;margin-top:6px}')
parts.append('.contact a{color:#0563C1;word-break:break-all}')
parts.append('.contact .ic{display:inline-block;width:18px;text-align:center;margin-right:4px;color:#777}')
parts.append('.councils{display:flex;flex-wrap:wrap;gap:4px;font-size:11px}')
parts.append('.council{background:#eef2f7;color:#3c4858;padding:2px 8px;border-radius:11px;white-space:nowrap}')
parts.append('.cards-side{font-size:11px;color:#666;display:flex;flex-direction:column;gap:2px;border-left:1px dashed #ddd;padding-left:10px}')
parts.append('.contracts{border-top:1px dashed #ddd;padding-top:8px;margin-top:10px}')
parts.append('.contracts-label{font-size:11px;color:#666;margin-bottom:4px;text-transform:uppercase;letter-spacing:.3px;font-weight:600}')
parts.append('.contract-block{background:#fafbfc;border-left:3px solid #1F4E79;padding:6px 10px;margin-bottom:5px;border-radius:0 4px 4px 0}')
parts.append('.contract-block .ctit{font-size:13px;color:#222;line-height:1.35;margin-bottom:2px}')
parts.append('.contract-block .cmeta{font-size:11px;color:#666}')
parts.append('.contract-block .cmeta .commiss{font-weight:600;color:#1F4E79}')
parts.append('.contract-block .cmeta a{color:#0563C1;text-decoration:none}')
parts.append('.contract-block .cmeta a:hover{text-decoration:underline}')
# Decision buttons
parts.append('.decision-bar{margin-top:10px;display:flex;gap:6px;flex-wrap:wrap;align-items:center}')
parts.append('.dec-btn{padding:6px 14px;font-size:12px;border:1px solid #bbb;border-radius:4px;cursor:pointer;background:#fff;font-weight:600;transition:all .15s}')
parts.append('.dec-btn:hover{background:#f3f5f8}')
parts.append('.dec-btn.approve{background:#28a745;border-color:#28a745;color:#fff}')
parts.append('.dec-btn.approve:hover{background:#218838}')
parts.append('.dec-btn.delete{background:#fff;border-color:#dc3545;color:#dc3545}')
parts.append('.dec-btn.delete:hover{background:#dc3545;color:#fff}')
parts.append('.dec-btn.maybe{background:#fff;border-color:#ffc107;color:#856404}')
parts.append('.dec-btn.maybe:hover{background:#ffc107;color:#212529}')
parts.append('.dec-btn.active.approve{background:#218838;color:#fff;box-shadow:0 0 0 2px rgba(40,167,69,.2)}')
parts.append('.dec-btn.active.delete{background:#dc3545;color:#fff;box-shadow:0 0 0 2px rgba(220,53,69,.2)}')
parts.append('.dec-btn.active.maybe{background:#ffc107;color:#212529;box-shadow:0 0 0 2px rgba(255,193,7,.2)}')
parts.append('.decision-label{font-size:11px;color:#777;font-weight:600;text-transform:uppercase;letter-spacing:.3px;margin-right:4px}')
parts.append('.entry.dec-approve{border-left:4px solid #28a745;background:#f4fbf6}')
parts.append('.entry.dec-delete{opacity:.55;border-left:4px solid #dc3545;background:#fff5f5}')
parts.append('.entry.dec-maybe{border-left:4px solid #ffc107;background:#fffbef}')
parts.append('.toolbar-actions{margin-left:auto;display:flex;gap:8px}')
parts.append('.toolbar-actions button{padding:6px 12px;font-size:12px;border:1px solid #bbb;border-radius:4px;cursor:pointer;background:#fff;font-weight:600}')
parts.append('.toolbar-actions button.primary{background:#1F4E79;color:#fff;border-color:#1F4E79}')
parts.append('.toolbar-actions button.warn{background:#fff;color:#dc3545;border-color:#dc3545}')
parts.append('.tally{display:inline-block;padding:3px 9px;border-radius:11px;font-size:12px;font-weight:600;margin-right:4px}')
parts.append('.tally.t-a{background:#d4edda;color:#1e6033}')
parts.append('.tally.t-d{background:#f8d7da;color:#721c24}')
parts.append('.tally.t-m{background:#fff3cd;color:#856404}')
parts.append('.tally.t-u{background:#e9ecef;color:#555}')
parts.append('@media (max-width:760px){.entry{grid-template-columns:1fr}.cards-side{border-left:none;border-top:1px dashed #ddd;padding-left:0;padding-top:8px;margin-top:8px}}')
parts.append('</style></head><body>')

parts.append('<h1>Entities being added to the main database</h1>')
parts.append(
    f'<div class="meta">'
    f'<b>{total}</b> entities will merge into the main database when you deploy. '
    f'Of those, <b>{net_new}</b> are <span class="tag netnew">NEW</span> (didn&rsquo;t exist in the main DB before this session) '
    f'and <b>{existing}</b> are <span class="tag existing">existing</span> (already in main DB — we&rsquo;re attaching London contracts to them).'
    f'</div>'
)

# Status stats
parts.append('<div class="stat-row">')
parts.append(f'<div class="stat"><span class="n">{total}</span> total entities</div>')
parts.append(f'<div class="stat"><span class="n">{net_new}</span> net-new providers</div>')
parts.append(f'<div class="stat"><span class="n">{status_counts["verified"]}</span> verified direct contacts</div>')
parts.append(f'<div class="stat"><span class="n">{status_counts["framework"]}</span> framework placeholders</div>')
parts.append(f'<div class="stat"><span class="n">{status_counts["partial"]}</span> partial contacts</div>')
parts.append(f'<div class="stat"><span class="n">{status_counts["unknown-research"]}</span> unknown</div>')
parts.append('</div>')

parts.append('<div class="search">')
parts.append('Filter: <input id="q" placeholder="name, borough, website, source id..." oninput="flt()">')
parts.append('<select id="netNew" onchange="flt()"><option value="">All entities</option><option value="new">Net-new only</option><option value="existing">Existing only</option></select>')
parts.append('<select id="status" onchange="flt()">'
             '<option value="">Any status</option>'
             '<option value="verified">Verified direct contact</option>'
             '<option value="framework">Framework placeholder</option>'
             '<option value="partial">Partial contact</option>'
             '<option value="unknown-research">Unknown / needs research</option>'
             '<option value="in-house council">In-house council</option>'
             '<option value="OK">Pre-existing OK</option>'
             '</select>')
parts.append('<select id="decision" onchange="flt()">'
             '<option value="">All decisions</option>'
             '<option value="approve">Approved only</option>'
             '<option value="delete">Rejected only</option>'
             '<option value="maybe">Maybe only</option>'
             '<option value="undecided">Undecided only</option>'
             '</select>')
parts.append('<span class="counts" id="c"></span>')
parts.append('<div class="toolbar-actions">')
parts.append('<span class="tally t-a" id="ta">0 approve</span>')
parts.append('<span class="tally t-d" id="td">0 reject</span>')
parts.append('<span class="tally t-m" id="tm">0 maybe</span>')
parts.append('<span class="tally t-u" id="tu">0 undecided</span>')
parts.append('<button onclick="approveAllVisible()">✓ Approve all visible</button>')
parts.append('<button onclick="deleteAllVisible()" class="warn">✗ Reject all visible</button>')
parts.append('<button onclick="exportDecisions()" class="primary">⬇ Download decisions (CSV)</button>')
parts.append('<button onclick="resetAll()">Reset</button>')
parts.append('</div>')
parts.append('</div>')

for e in entries:
    name = htmllib.escape(e['name'])
    cls = 'net-new' if e['is_net_new'] else 'existing'
    new_tag = '<span class="tag netnew">NEW</span>' if e['is_net_new'] else '<span class="tag existing">existing</span>'

    status_classes = {
        'verified': 'verified', 'framework': 'framework', 'partial': 'partial',
        'unknown-research': 'unknown', 'in-house council': 'inhouse', 'OK': 'ok',
        'main-db-existing': 'ok',
    }
    sc = status_classes.get(e['review_tag'], 'ok')
    status_label = e['review_tag'] or '—'
    status_tag = f'<span class="tag {sc}">{htmllib.escape(status_label)}</span>'

    # Contact bits
    contact_lines = []
    if e['website']:
        url = e['website']
        if not url.startswith('http'): url = 'https://' + url
        contact_lines.append(f'<div><span class="ic">🌐</span><a href="{htmllib.escape(url)}" target="_blank">{htmllib.escape(e["website"])}</a></div>')
    if e['phone']:
        contact_lines.append(f'<div><span class="ic">📞</span>{htmllib.escape(e["phone"])}</div>')
    if e['email']:
        contact_lines.append(f'<div><span class="ic">✉️</span><a href="mailto:{htmllib.escape(e["email"])}">{htmllib.escape(e["email"])}</a></div>')
    if not contact_lines:
        contact_lines.append('<div><i>(no direct contact — engage via commissioning council)</i></div>')

    side_bits = []
    if e['companies_house']:
        ch = e['companies_house']
        ch_link = f'https://find-and-update.company-information.service.gov.uk/company/{ch}'
        side_bits.append(f'CH: <a href="{ch_link}" target="_blank">{ch}</a>')
    if e['charity_no']:
        side_bits.append(f'Charity: {htmllib.escape(e["charity_no"])}')
    if e['contact_page']:
        side_bits.append(f'<a href="{htmllib.escape(e["contact_page"])}" target="_blank">Contact page</a>')

    councils_html = ''.join(f'<span class="council">{htmllib.escape(c)}</span>' for c in e['councils'])

    # Contracts block — one per contract row
    contracts_html = ''
    if e['contracts']:
        rows_html = []
        for c in e['contracts']:
            tit = htmllib.escape(c['title']) if c['title'] else '<i>(untitled)</i>'
            cm = htmllib.escape(c['council'])
            cats = htmllib.escape(c['categories']) if c['categories'] else ''
            award = htmllib.escape(c['award_date']) if c['award_date'] else ''
            src_link = ''
            if c['bidstats_id']:
                src_link = f' · <a href="https://bidstats.uk/notice/{c["bidstats_id"]}" target="_blank">Bidstats {c["bidstats_id"]}</a>'
            meta_bits = []
            if cats: meta_bits.append(cats)
            if award: meta_bits.append(f'awarded {award}')
            meta = ' · '.join(meta_bits)
            rows_html.append(
                f'<div class="contract-block">'
                f'<div class="ctit">{tit}</div>'
                f'<div class="cmeta"><span class="commiss">{cm}</span>'
                f'{(" · " + meta) if meta else ""}{src_link}</div>'
                f'</div>'
            )
        contracts_html = (
            f'<div class="contracts">'
            f'<div class="contracts-label">Contract{"s" if len(e["contracts"])!=1 else ""} ({len(e["contracts"])})</div>'
            f'{"".join(rows_html)}'
            f'</div>'
        )

    search_blob = (e['name'] + ' ' + e['website'] + ' ' + e['email'] + ' ' +
                   e['phone'] + ' ' + ' '.join(e['councils']) + ' ' + e['review_tag'] + ' ' +
                   ' '.join(c['title'] + ' ' + c['bidstats_id'] for c in e['contracts'])).lower()

    entity_id = re.sub(r'[^a-z0-9]', '_', e['name'].lower())[:60]
    parts.append(f'<div class="entry {cls}" data-id="{htmllib.escape(entity_id)}" data-name="{htmllib.escape(e["name"])}" data-search="{htmllib.escape(search_blob)}" data-new="{1 if e["is_net_new"] else 0}" data-status="{htmllib.escape(e["review_tag"])}">')
    parts.append('<div>')
    parts.append(f'<h2>{name}{new_tag}{status_tag}</h2>')
    parts.append(f'<div class="contact">{"".join(contact_lines)}</div>')
    if contracts_html:
        parts.append(contracts_html)
    parts.append('<div class="decision-bar">')
    parts.append('<span class="decision-label">Decision:</span>')
    parts.append(f'<button class="dec-btn approve" onclick="decide(this,\'approve\')">✓ Approve</button>')
    parts.append(f'<button class="dec-btn maybe" onclick="decide(this,\'maybe\')">? Maybe</button>')
    parts.append(f'<button class="dec-btn delete" onclick="decide(this,\'delete\')">✗ Reject</button>')
    parts.append('</div>')
    parts.append('</div>')
    parts.append(f'<div class="cards-side">{"<br>".join(side_bits) or "&nbsp;"}</div>')
    parts.append('</div>')

parts.append('<script>')
# Decision logic with localStorage
parts.append('const STORAGE_KEY="adding_to_db_decisions_v1";')
parts.append('let decisions = {}; try { decisions = JSON.parse(localStorage.getItem(STORAGE_KEY)||"{}"); } catch(e){}')
parts.append('function saveDecisions(){ try { localStorage.setItem(STORAGE_KEY, JSON.stringify(decisions)); } catch(e){ console.error(e); } }')
parts.append('function applyDecisionClass(entry){ entry.classList.remove("dec-approve","dec-delete","dec-maybe"); const d = decisions[entry.dataset.id]; if(d==="approve") entry.classList.add("dec-approve"); else if(d==="delete") entry.classList.add("dec-delete"); else if(d==="maybe") entry.classList.add("dec-maybe"); entry.querySelectorAll(".dec-btn").forEach(b=>{b.classList.remove("active"); if(b.textContent.toLowerCase().includes(d||"")) {} }); /* highlight active */ const map={"approve":"approve","delete":"delete","maybe":"maybe"}; entry.querySelectorAll(".dec-btn").forEach(b=>{ const k = b.classList.contains("approve")?"approve":b.classList.contains("delete")?"delete":b.classList.contains("maybe")?"maybe":""; if(k===d) b.classList.add("active"); else b.classList.remove("active"); }); }')
parts.append('function decide(btn, action){ const entry = btn.closest(".entry"); const id = entry.dataset.id; if(decisions[id]===action){ delete decisions[id]; } else { decisions[id] = action; } saveDecisions(); applyDecisionClass(entry); cnt(); }')
parts.append('function approveAllVisible(){ if(!confirm("Mark every currently VISIBLE entity as APPROVED?")) return; document.querySelectorAll(".entry").forEach(e=>{ if(e.style.display!=="none"){ decisions[e.dataset.id]="approve"; applyDecisionClass(e); } }); saveDecisions(); cnt(); }')
parts.append('function deleteAllVisible(){ if(!confirm("Mark every currently VISIBLE entity as REJECTED?")) return; document.querySelectorAll(".entry").forEach(e=>{ if(e.style.display!=="none"){ decisions[e.dataset.id]="delete"; applyDecisionClass(e); } }); saveDecisions(); cnt(); }')
parts.append('function resetAll(){ if(!confirm("Wipe ALL decisions (cannot undo)?")) return; decisions = {}; saveDecisions(); document.querySelectorAll(".entry").forEach(applyDecisionClass); cnt(); }')
parts.append('function cnt(){ let a=0,d=0,m=0,u=0; document.querySelectorAll(".entry").forEach(e=>{ const x=decisions[e.dataset.id]; if(x==="approve")a++; else if(x==="delete")d++; else if(x==="maybe")m++; else u++; }); document.getElementById("ta").textContent=a+" approve"; document.getElementById("td").textContent=d+" reject"; document.getElementById("tm").textContent=m+" maybe"; document.getElementById("tu").textContent=u+" undecided"; }')
parts.append('function exportDecisions(){ const rows=[["entity_id","entity_name","decision","review_tag","is_net_new","councils","website","phone","email"]]; document.querySelectorAll(".entry").forEach(e=>{ const id=e.dataset.id; const dec=decisions[id]||""; if(!dec) return; /* only export decided */ const name=e.dataset.name; const tag=e.dataset.status; const isNew=e.dataset.new; const councilEls=e.querySelectorAll(".commiss"); const councils=Array.from(new Set(Array.from(councilEls).map(c=>c.textContent))).join(" | "); const cn=e.querySelector(".contact"); const wsLink=cn?cn.querySelector("a[href^=http]"):null; const ws=wsLink?wsLink.href:""; const phMatch=cn?cn.textContent.match(/[0-9 ]{6,}/):""; const ph=phMatch?phMatch[0].trim():""; const emLink=cn?cn.querySelector("a[href^=mailto]"):null; const em=emLink?emLink.href.replace(/^mailto:/,""):""; rows.push([id,JSON.stringify(name),dec,tag,isNew,JSON.stringify(councils),ws,JSON.stringify(ph),em]); }); if(rows.length===1){ alert("No decisions yet — mark some entries Approve / Reject / Maybe first."); return; } const csv=rows.map(r=>r.join(",")).join("\\n"); const blob=new Blob([csv],{type:"text/csv"}); const a=document.createElement("a"); a.href=URL.createObjectURL(blob); const ts=new Date().toISOString().slice(0,19).replace(/[:.]/g,"-"); a.download=`adding_to_db_decisions_${ts}.csv`; a.click(); }')
parts.append('function flt(){const q=document.getElementById("q").value.toLowerCase(); const nn=document.getElementById("netNew").value; const st=document.getElementById("status").value; const dec=document.getElementById("decision").value; const items=document.querySelectorAll(".entry"); let shown=0; items.forEach(it=>{let ok=true; if(q && !it.dataset.search.includes(q))ok=false; if(nn==="new" && it.dataset.new!=="1")ok=false; if(nn==="existing" && it.dataset.new!=="0")ok=false; if(st && it.dataset.status!==st)ok=false; if(dec){ const cur = decisions[it.dataset.id] || ""; if(dec==="undecided" && cur) ok=false; if(dec!=="undecided" && cur!==dec) ok=false; } it.style.display=ok?"":"none"; if(ok)shown++;}); document.getElementById("c").textContent=shown+" of "+items.length+" entities shown"; cnt(); }')
parts.append('document.querySelectorAll(".entry").forEach(applyDecisionClass);')
parts.append('flt();')
parts.append('</script></body></html>')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(''.join(parts))

print(f"Wrote {OUT}")
print(f"  {total} total entities ({net_new} net-new + {existing} existing)")
print(f"  {status_counts['verified']} verified · {status_counts['framework']} framework · "
      f"{status_counts['partial']} partial · {status_counts['unknown-research']} unknown")
print(f"  Size: {len(open(OUT,'rb').read())//1024}KB")
