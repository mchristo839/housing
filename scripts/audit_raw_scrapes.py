"""Walk every cached raw scrape on disk. For each:
   - Classify: London? housing-relevant? has named suppliers?
   - Dedupe vs current providers.json + manual_contracts.xlsx
   - Count new vs known
   - Tag the source quality

Output: audit report telling us if our scraping targets are right + what's
new to process.
"""
import openpyxl, glob, re, html, os, json, sys, io
from collections import Counter, defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ───────────────────────────────────────────────────────────────────────────
# Load current live data + manual additions for dedup
# ───────────────────────────────────────────────────────────────────────────
prov = json.load(open('api/_data/providers.json', encoding='utf-8'))
def normname(s): return re.sub(r'[^a-z0-9]', '', str(s or '').lower())
live_norm = {normname(p['name']) for p in prov}

wb_m = openpyxl.load_workbook('data/manual_contracts.xlsx', read_only=True)
manual_suppliers = set()
for r in wb_m['Company × Council × Sector'].iter_rows(min_row=2, values_only=True):
    if r[0]: manual_suppliers.add(normname(r[0]))
all_known = live_norm | manual_suppliers

# Bidstats notice IDs already in our review_db
wb_c = openpyxl.load_workbook('data/scraped/curated_scraped.xlsx', read_only=True)
processed_sids = set()
for r in wb_c.active.iter_rows(min_row=2, values_only=True):
    if len(r) > 19 and r[19]:
        processed_sids.add(str(r[19]).strip())

print(f"Baseline (already in DB):")
print(f"  Live providers       : {len(prov)}")
print(f"  Manual file suppliers: {len(manual_suppliers)}")
print(f"  Bidstats notice IDs already processed: {len(processed_sids)}")

# ───────────────────────────────────────────────────────────────────────────
# Filters — same as build_data.py so we count what would actually land
# ───────────────────────────────────────────────────────────────────────────
LONDON_KW = re.compile(
    r'london borough|royal borough of greenwich|royal borough of kingston|'
    r'westminster city|tower hamlets|hackney|islington|lambeth|southwark|'
    r'lewisham|hounslow|brent|ealing|haringey|enfield|bexley|bromley|sutton|'
    r'merton|wandsworth|hammersmith|kensington|camden|barnet|harrow|hillingdon|'
    r'havering|redbridge|newham|waltham forest|barking|dagenham|croydon|'
    r'capital letters|greater london authority|mopac|liia', re.I)

H_OK = re.compile(
    r'supported (living|accommodation|housing)|temporary accommodation|'
    r'emergency accommodation|hostel|refuge|housing related support|extra care|'
    r'sheltered (housing|accommodation)|young people.{0,40}(housing|accommodation|pathway)|'
    r'care leavers|homeless|asylum|nightly paid|social housing|'
    r'mental health.{0,40}(accommodation|housing|supported|step|crisis)|'
    r'learning disabilit.{0,40}(accommodation|housing|supported|residential|placement)|'
    r'domestic abuse|rough sleeper|housing first|semi[- ]?independent|'
    r'floating support|move[- ]?on|tenancy', re.I)

NOT_H = re.compile(
    r'\bhome care\b|home support service|home (domiciliary|based) care|'
    r'aids and adaptations|catering|street lighting|grass cutting|repairs only|'
    r'lift modernisation|cleaning services|legal services|insurance|'
    r'civil enforcement|pan[- ]?london domiciliary|reablement only', re.I)

CARE_FACILITY = re.compile(
    r"\b(?:care home(?:s)?|nursing home(?:s)?|residential home(?:s)?|"
    r"residential care home|care centre|care center|care lodge|care manor)\b", re.I)

# ───────────────────────────────────────────────────────────────────────────
# Source 1: cached Bidstats HTML notices
# ───────────────────────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("SOURCE 1: Cached Bidstats HTML notices")
print('='*65)

files = sorted(glob.glob('.scratch/bidstats/*.html')) + \
        sorted(glob.glob('.scratch/bidstats_3boroughs/*.html'))
print(f"Cached notices on disk: {len(files)}")

bidstats_stats = Counter()
bidstats_new_contracts = []
bidstats_new_suppliers = set()
borough_to_new = defaultdict(set)

for path in files:
    nid = os.path.basename(path).replace('.html','')
    h = open(path, encoding='utf-8', errors='replace').read()
    text = re.sub(r'<script.*?</script>', '', h, flags=re.S)
    text = re.sub(r'<style.*?</style>', '', text, flags=re.S)
    text = html.unescape(re.sub(r'<[^>]+>', '\n', text))
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n+', '\n', text)

    # status
    if '[Award]' in text:    status = 'Award'
    elif '[Tender]' in text: status = 'Tender'
    elif '[Notice]' in text: status = 'Notice'
    else:                    status = 'Other'

    bidstats_stats[f'total {status}'] += 1

    if not LONDON_KW.search(text):
        bidstats_stats['out-of-London'] += 1
        continue
    if not H_OK.search(text[:3000]):
        bidstats_stats['London but not housing'] += 1
        continue
    if NOT_H.search(text[:3000]):
        bidstats_stats['London housing but excluded (homecare/FM/etc)'] += 1
        continue

    tm = re.search(r'\n\s*([A-Z][^\n]{15,200}?)\s*\[(Tender|Award|Notice|Modification|PIN|Pipeline|Addendum)\]', text)
    title = tm.group(1).strip() if tm else ""

    suppliers = []
    ai = text.lower().find('award detail')
    if ai > 0:
        section = text[ai:ai+15000]
        for m in re.finditer(r"\n\s*([A-Z][A-Za-z0-9 &'\.\,/-]{4,80}?)\s*\n\s*\(([A-Za-z ,]+?)\)", section):
            name = m.group(1).strip()
            if name.lower() in {'award detail','awards','contractors','supplier analysis',
                                'quality','price','award criteria','reference','status','history',
                                'categories','domains','indicators','social value','published',
                                'published category','unnamed'}: continue
            if name not in suppliers: suppliers.append(name)

    # Which borough(s) does the title/text reference?
    boroughs = set()
    for bm in LONDON_KW.finditer(text[:5000]):
        boroughs.add(bm.group(0).lower())

    if status != 'Award':
        bidstats_stats[f'London housing {status} (no suppliers possible)'] += 1
        continue
    if not suppliers:
        bidstats_stats['London housing Award but no parseable suppliers'] += 1
        continue

    bidstats_stats['London housing Award WITH suppliers'] += 1

    already_processed = nid in processed_sids
    if already_processed:
        bidstats_stats['London housing Award (already in review db)'] += 1
        continue

    # New notice — record it
    bidstats_stats['London housing Award (NEW to add)'] += 1
    new_sup_here = [s for s in suppliers
                    if normname(s) not in all_known and not CARE_FACILITY.search(s)]
    new_dupes = [s for s in suppliers
                 if normname(s) in all_known]

    bidstats_new_contracts.append({
        'nid': nid, 'title': title, 'all_suppliers': len(suppliers),
        'new_suppliers': len(new_sup_here), 'dupes': len(new_dupes),
        'boroughs': sorted(boroughs)[:3],
    })
    for s in new_sup_here:
        bidstats_new_suppliers.add(normname(s))
        for b in boroughs:
            borough_to_new[b].add(normname(s))

print(f"\nDistribution of cached Bidstats notices:")
for k, n in bidstats_stats.most_common():
    print(f"  {n:4d}  {k}")

print(f"\nNEW Bidstats Award notices to process: {len(bidstats_new_contracts)}")
print(f"NEW unique suppliers discoverable from them: {len(bidstats_new_suppliers)}")
print(f"\nTop new-contract notices by new-supplier count:")
for n in sorted(bidstats_new_contracts, key=lambda x: -x['new_suppliers'])[:10]:
    print(f"  {n['nid']}  +{n['new_suppliers']:3d} new (+{n['dupes']:2d} dupes)  | {n['title'][:55]}  | {', '.join(n['boroughs'])[:30]}")

# ───────────────────────────────────────────────────────────────────────────
# Source 2: data/scraped/raw/contractsfinder JSON
# ───────────────────────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("SOURCE 2: Contracts Finder JSON")
print('='*65)
cf_path = 'data/scraped/raw/contractsfinder_2026-06-06.json'
if os.path.exists(cf_path) and os.path.getsize(cf_path) > 100:
    cf = json.load(open(cf_path, encoding='utf-8'))
    if isinstance(cf, dict):
        records = cf.get('records', cf.get('notices', list(cf.values())[0] if cf else []))
    else:
        records = cf
    if not isinstance(records, list): records = [records]
    print(f"CF records on disk: {len(records)}")

    cf_stats = Counter()
    cf_new_suppliers = set()
    for r in records[:5000]:
        if not isinstance(r, dict): continue
        buyer = ((r.get('buyer') or r.get('publisher') or r.get('organisationName') or '') if isinstance(r.get('buyer'), str) else (r.get('buyer',{}).get('name','') if isinstance(r.get('buyer'),dict) else ''))
        title = r.get('title') or r.get('notice', {}).get('title','') if isinstance(r.get('notice'),dict) else r.get('title','')
        title = str(title or '')
        buyer = str(buyer or '')
        blob = buyer + ' ' + title
        if not LONDON_KW.search(blob):
            cf_stats['out-of-London'] += 1
            continue
        if not H_OK.search(blob):
            cf_stats['London but not housing'] += 1
            continue
        if NOT_H.search(blob):
            cf_stats['London housing but excluded'] += 1
            continue
        cf_stats['London housing relevant'] += 1
        # suppliers?
        sups = r.get('awardedSuppliers') or r.get('awards') or r.get('contractAwards') or []
        if isinstance(sups, dict): sups = [sups]
        n_new = 0
        for s in sups[:30]:
            if isinstance(s, dict):
                nm = s.get('name') or s.get('supplierName') or ''
            else:
                nm = str(s)
            if nm and normname(nm) not in all_known and not CARE_FACILITY.search(nm):
                cf_new_suppliers.add(normname(nm))
                n_new += 1
        if n_new:
            cf_stats['London housing Award with new suppliers'] += 1
    print(f"\nDistribution:")
    for k, n in cf_stats.most_common():
        print(f"  {n:5d}  {k}")
    print(f"\nNet-new suppliers discoverable: {len(cf_new_suppliers)}")
else:
    print("  (no usable Contracts Finder data on disk)")

# ───────────────────────────────────────────────────────────────────────────
# Source 3: data/scraped/raw/bidstats_london_via_search JSON
# ───────────────────────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("SOURCE 3: Bidstats London via Google search JSON")
print('='*65)
bv_path = 'data/scraped/raw/bidstats_london_via_search_2026-06-07.json'
if os.path.exists(bv_path) and os.path.getsize(bv_path) > 100:
    bv = json.load(open(bv_path, encoding='utf-8'))
    if isinstance(bv, dict): bv = bv.get('urls', bv.get('records', []))
    print(f"  Items: {len(bv) if isinstance(bv, list) else 'dict'}")
    if isinstance(bv, list):
        # urls only — these would feed into Bidstats fetch if we ran again
        already = sum(1 for url in bv if isinstance(url, str) and any(p in url for p in processed_sids))
        print(f"  Items pointing at notices we already processed: {already}")
        print(f"  Items not yet processed: {len(bv) - already}")
else:
    print("  (file empty or missing)")

# ───────────────────────────────────────────────────────────────────────────
# Verdict
# ───────────────────────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("VERDICT — IS OUR SCRAPING TARGETING THE RIGHT THINGS?")
print('='*65)
total_cached = sum(v for k,v in bidstats_stats.items() if k.startswith('total'))
landing = bidstats_stats.get('London housing Award (NEW to add)', 0) + bidstats_stats.get('London housing Award (already in review db)', 0)
landing_pct = (landing*100/total_cached) if total_cached else 0
print(f"  Bidstats cache total: {total_cached}")
print(f"  Of those, London-housing-Award (the bullseye): {landing} ({landing_pct:.0f}%)")
print(f"  Hit rate: {'GOOD' if landing_pct > 30 else 'MEDIOCRE — too much noise' if landing_pct > 10 else 'POOR — wrong source'}")
