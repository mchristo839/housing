"""Set Companies.Housing Contracts = count-from-CCS for the GMCA Housing First
providers I just added. The build_data filter requires housing > 0 or the row
gets dropped at the 'homecare_only' filter."""
import openpyxl
from collections import Counter

PATH = 'data/manual_contracts.xlsx'
wb = openpyxl.load_workbook(PATH)
companies = wb['Companies']
ccs = wb['Company × Council × Sector']

# Count housing contracts per company in CCS
counts = Counter()
for r in range(2, ccs.max_row + 1):
    co = (ccs.cell(row=r, column=1).value or '').strip()
    sec = (ccs.cell(row=r, column=2).value or '').strip().lower()
    if co and sec == 'housing':
        counts[co.lower()] += int(ccs.cell(row=r, column=4).value or 1)

# Update Companies sheet for the 6 new GMCA providers
NEW_GMCA = [
    'Petrus Community', 'Jigsaw Support', 'Stockport Homes Group',
    'Manchester Action on Street Health (MASH)', 'Early Break',
    'Community-Led Initiatives',
]

for target in NEW_GMCA:
    tl = target.lower()
    n_housing = counts.get(tl, 0)
    if n_housing == 0:
        print(f"  ! no CCS rows found for {target}")
        continue
    # find Companies row
    for r in range(2, companies.max_row + 1):
        if (companies.cell(row=r, column=1).value or '').strip().lower() == tl:
            # column 5 = "Housing Contracts", column 2 = "Total Contracts"
            companies.cell(row=r, column=5).value = n_housing
            companies.cell(row=r, column=2).value = n_housing
            # Also set "Housing Councils" (col 6) to a semicolon-list of councils
            # (so the source-of-truth display matches the CCS reality)
            councils = []
            for rr in range(2, ccs.max_row + 1):
                co = (ccs.cell(row=rr, column=1).value or '').strip().lower()
                if co == tl and (ccs.cell(row=rr, column=2).value or '').lower() == 'housing':
                    c = (ccs.cell(row=rr, column=3).value or '').strip()
                    if c and c not in councils: councils.append(c)
            companies.cell(row=r, column=6).value = '; '.join(councils)
            companies.cell(row=r, column=11).value = '; '.join(councils)  # All Councils
            print(f"  ✓ {target}: Housing={n_housing}  Councils=[{', '.join(councils)}]")
            break

wb.save(PATH)
print(f"\nSaved {PATH}")
