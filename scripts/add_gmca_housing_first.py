"""Add evidence-backed GMCA Housing First contracts + Wigan ABEN to manual_contracts.xlsx.

Evidence: GMCA Housing First pilot (April 2019, extended to March 2029).
Lead partner Great Places, 10-organisation partnership, 4-zone delivery model.
Sources:
  - https://www.greatermanchester-ca.gov.uk/what-we-do/planning-and-housing/housing/housing-first-greater-manchester/
  - https://www.greatplaces.org.uk/about-us/greater-manchester-housing-first/
  - https://www.regenda.org.uk/news/greater-manchester-housing-first-project-launches-in-bolton-bury-and-rochdale-2487
  - https://www.gmhousingfirst.org.uk/case-studies
  - Inside Housing: https://www.insidehousing.co.uk/news/housing-associations-selected-to-deliver-housing-first-pilot-in-manchester-for-two-more-years-76206

Plus Wigan ABEN (Queens Hall Action on Poverty) Contract Award - Wigan Council, £597K, Dec 2022:
  - Bidstats 789119320 / Contracts Finder 10ea5135-36c3-4711-884d-9ea278edf186
"""
import openpyxl
from pathlib import Path

PATH = Path('data/manual_contracts.xlsx')
wb = openpyxl.load_workbook(PATH)
companies = wb['Companies']
ccs = wb['Company × Council × Sector']

# --- Zone → councils mapping for GMCA Housing First ---
ZONES = {
    'A': ['Manchester City Council'],
    'B': ['Bolton Council', 'Bury Council', 'Rochdale Borough Council'],
    'C': ['Oldham Metropolitan Borough Council',
          'Stockport Metropolitan Borough Council',
          'Tameside Metropolitan Borough Council'],
    'D': ['Salford City Council', 'Trafford Council', 'Wigan Council'],
}

# --- Provider × zones (verified from GMCA + Great Places + partner sources) ---
HF_PROVIDERS = [
    # (Company, zones-they-deliver-in, website, phone, email, ch_number, is_sme, is_vcse, charity_no)
    ('Petrus Community',            ['B'],        'https://www.petrus.org.uk/',           '01706 341 121', 'info@petrus.org.uk',         '04330870', 'No', 'Yes', '1095636'),
    ('Jigsaw Support',              ['C'],        'https://support.jigsawhomes.org.uk/',  '0300 111 1133', 'enquiries@jigsawhomes.org.uk', '04598378', 'No', 'No', None),
    ('Stockport Homes Group',       ['C'],        'https://www.stockporthomesgroup.com/', '0161 217 6016', 'customerservices@stockporthomes.org', '04026317', 'No', 'No', None),
    ('Manchester Action on Street Health (MASH)', ['A'], 'https://mash.org.uk/',        '0161 273 4555', 'admin@mash.org.uk',          '03301466', 'Yes', 'Yes', '1051754'),
    ('Early Break',                 ['B'],        'https://www.earlybreak.co.uk/',        '0161 723 3880', 'info@earlybreak.co.uk',      '02942506', 'Yes', 'Yes', '1031022'),
    ('Community-Led Initiatives',   ['C'],        'https://www.communityledinitiatives.org.uk/', '0161 393 3008', 'info@cli.org.uk', None, 'Yes', 'Yes', '1184056'),
    # Existing companies — add their HF zones (no Companies sheet change needed since the row exists)
    ('Riverside',                   ['A'],        None, None, None, None, None, None, None),       # The Riverside Group
    ('Great Places Housing',        ['D'],        None, None, None, None, None, None, None),
    ('Humankind',                   ['D'],        None, None, None, None, None, None, None),       # Humankind Charity
]

# Use exact existing-company names where we already have entries (to allow dedup at build time)
NAME_ALIASES = {
    'Riverside':              'The Riverside Group',
    'Great Places Housing':   'Great Places Housing Association',
    'Humankind':              'Humankind Charity',
}

def find_company_row(name):
    """Return row index for an existing Companies sheet entry matching name, else None."""
    target = name.lower().strip()
    for r in range(2, companies.max_row + 1):
        v = (companies.cell(row=r, column=1).value or '').strip().lower()
        if v == target:
            return r
    return None

def add_company(name, website, phone, email, ch, is_sme, is_vcse, charity_no):
    """Append a new Companies row. Skip if already present."""
    if find_company_row(name):
        print(f"  - Companies: '{name}' already present, skipping")
        return
    next_row = companies.max_row + 1
    # Header order: Company | Total | Homecare | HomecareCouncils | Housing | HousingCouncils | Both?
    #               | CH | SME | VCSE | AllCouncils | Address | Website | Phone | Email | ContactPage | Charity# | Income | Status | Activities | Review
    companies.cell(row=next_row, column=1).value  = name
    companies.cell(row=next_row, column=2).value  = 0          # Total Contracts (build_data will compute)
    companies.cell(row=next_row, column=5).value  = 0          # Housing Contracts
    companies.cell(row=next_row, column=7).value  = 'No'
    companies.cell(row=next_row, column=8).value  = ch
    companies.cell(row=next_row, column=9).value  = is_sme
    companies.cell(row=next_row, column=10).value = is_vcse
    companies.cell(row=next_row, column=13).value = website
    companies.cell(row=next_row, column=14).value = phone
    companies.cell(row=next_row, column=15).value = email
    if charity_no:
        companies.cell(row=next_row, column=17).value = charity_no
        companies.cell(row=next_row, column=19).value = 'Registered Charity'
    print(f"  + Companies: added '{name}'")

def add_ccs_row(company_name, council, contract_title, region='North West'):
    """Append a row to Company × Council × Sector."""
    next_row = ccs.max_row + 1
    ccs.cell(row=next_row, column=1).value  = company_name
    ccs.cell(row=next_row, column=2).value  = 'Housing'
    ccs.cell(row=next_row, column=3).value  = council
    ccs.cell(row=next_row, column=4).value  = 1
    ccs.cell(row=next_row, column=9).value  = 'No'
    ccs.cell(row=next_row, column=10).value = 'No'
    ccs.cell(row=next_row, column=11).value = 'Supported accommodation | Housing First'
    ccs.cell(row=next_row, column=12).value = '01/04/2019'
    ccs.cell(row=next_row, column=13).value = contract_title
    ccs.cell(row=next_row, column=14).value = region
    ccs.cell(row=next_row, column=15).value = 'Local Authority'
    ccs.cell(row=next_row, column=16).value = 'Local'
    print(f"  + CCS: {company_name:38s} × {council:50s}")

# 1. Add new Companies entries
print("=== Companies sheet ===")
for (name, zones, website, phone, email, ch, sme, vcse, charity) in HF_PROVIDERS:
    if name in NAME_ALIASES: continue  # existing entry, skip Companies add
    if website is None: continue       # passthrough for existing (Riverside, Great Places, Humankind)
    add_company(name, website, phone, email, ch, sme, vcse, charity)

# 2. Add Company × Council × Sector rows (the evidence-backed contracts)
print("\n=== Company × Council × Sector (Housing First) ===")
HF_TITLE = ("Greater Manchester Housing First pilot — pan-GM partnership (Apr 2019, extended Mar 2029). "
            "Lead partner Great Places. 10-org partnership funded by MHCLG via GMCA. "
            "Source: greatermanchester-ca.gov.uk / gmhousingfirst.org.uk")
for (name, zones, *_rest) in HF_PROVIDERS:
    company_name = NAME_ALIASES.get(name, name)
    for z in zones:
        for council in ZONES[z]:
            add_ccs_row(company_name, council, HF_TITLE)

# 3. Wigan ABEN (already in DB, but make sure)
print("\n=== Wigan ABEN ===")
add_ccs_row('Queens Hall Action on Poverty', 'Wigan Council',
            "A Bed Every Night (ABEN) — Wigan local service, £597K 3-year contract (Jan 2023 – Mar 2026). "
            "Source: Contracts Finder 10ea5135-36c3-4711-884d-9ea278edf186 / Bidstats 789119320")

wb.save(PATH)
print(f"\n✓ Wrote {PATH}")
print(f"  Companies sheet: {companies.max_row} rows")
print(f"  CCS sheet:       {ccs.max_row} rows")
